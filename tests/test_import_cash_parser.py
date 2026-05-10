"""Smoke test for the Касса (1C cash receipts) parser.

Pins the parsing rules captured in memory `finances_1c_parsing_rules`:
  - cp1251 + skip rows without 'V'/marker in column 0
  - Касса client key = Субконто1, NOT 'Принято от' (received_from)
  - corr_account 40.10 → UZS leg uses Сумма; 40.11 → USD leg uses ВалСумма
  - Totals/separator rows are dropped silently
  - parse returns {ok, payments[], stats{}} on success

Per the project's "zero data loss" red line (memory: no_data_loss),
importer parsers are the highest-risk untested code in the tree —
this test catches the next regression that would otherwise ship silently.

Synthetic .xlsx fixtures only — no PII committed.
"""
import io

import openpyxl

from backend.services.import_cash import parse_cash_xls


def _build_cash_xlsx(rows: list[dict]) -> bytes:
    """Construct a synthetic Касса .xlsx as bytes.

    Layout mirrors what 1C exports (one of the variants — the parser is
    tolerant of column shuffling because it builds a column map by header
    name):

      Row 1: report title (ignored)
      Row 2: blank (ignored)
      Row 3: header row — column captions in Cyrillic
      Row 4+: data rows. A 'V' marker in column 0 means it's a payment row.
              Empty marker means a separator / totals row that the parser
              must drop.

    `rows` is a list of dicts; each dict gets one data row. Required keys:
      marker (str, '' or 'V'), doc_number, doc_date, author, received_from,
      basis, attachment, corr_account, subconto1, subconto2, subconto3,
      currency_code, fx_rate, amount_local, amount_currency, cashflow_category
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Касса"

    ws.cell(row=1, column=1, value="Отчёт кассира")

    headers = [
        (1, ""),  # V-marker column — no header
        (2, "Номер"),
        (3, "Дата"),
        (4, "Время"),
        (5, "Автор"),
        (6, "Принято от"),
        (7, "Основание"),
        (8, "Приложение"),
        (9, "Корреспондирующий счет"),
        (10, "Субконто1"),
        (11, "Субконто2"),
        (12, "Субконто3"),
        (13, "Сумма"),
        (14, "Валюта"),
        (15, "Движение денежных средств"),
        (16, "Курс"),
        (17, "ВалСумма"),
    ]
    for col, val in headers:
        ws.cell(row=3, column=col, value=val)

    for i, row in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=row.get("marker", ""))
        ws.cell(row=r, column=2, value=row.get("doc_number", ""))
        ws.cell(row=r, column=3, value=row.get("doc_date", ""))
        ws.cell(row=r, column=4, value=row.get("doc_time", ""))
        ws.cell(row=r, column=5, value=row.get("author", ""))
        ws.cell(row=r, column=6, value=row.get("received_from", ""))
        ws.cell(row=r, column=7, value=row.get("basis", ""))
        ws.cell(row=r, column=8, value=row.get("attachment", ""))
        ws.cell(row=r, column=9, value=row.get("corr_account", ""))
        ws.cell(row=r, column=10, value=row.get("subconto1", ""))
        ws.cell(row=r, column=11, value=row.get("subconto2", ""))
        ws.cell(row=r, column=12, value=row.get("subconto3", ""))
        ws.cell(row=r, column=13, value=row.get("amount_local", 0))
        ws.cell(row=r, column=14, value=row.get("currency_code", ""))
        ws.cell(row=r, column=15, value=row.get("cashflow_category", ""))
        ws.cell(row=r, column=16, value=row.get("fx_rate", 0))
        ws.cell(row=r, column=17, value=row.get("amount_currency", 0))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_cash_dual_currency():
    """Two UZS rows + one USD row + a totals/separator row that must be ignored."""
    fixture = [
        # UZS payment from a real client — corr_account 40.10
        {
            "marker": "V",
            "doc_number": "K-001",
            "doc_date": "2026-04-01",
            "author": "Кассир",
            "received_from": "Иванов Иван",  # physical deliverer (must NOT be the client key)
            "basis": "Оплата по договору",
            "corr_account": "40.10",
            "subconto1": "ООО АЛЬФА",  # actual client
            "amount_local": 1500000.0,
            "currency_code": "UZS",
            "fx_rate": 0,
            "amount_currency": 0,
            "cashflow_category": "Поступление",
        },
        # USD payment — corr_account 40.11
        {
            "marker": "V",
            "doc_number": "K-002",
            "doc_date": "2026-04-01",
            "author": "Кассир",
            "received_from": "Курьер",
            "basis": "Оплата",
            "corr_account": "40.11",
            "subconto1": "ИП БЕТА",
            "amount_local": 1280000.0,  # in UZS column too (1C convention)
            "currency_code": "USD",
            "fx_rate": 12800.0,
            "amount_currency": 100.0,  # actual USD amount
            "cashflow_category": "Поступление",
        },
        # Another UZS payment same day
        {
            "marker": "V",
            "doc_number": "K-003",
            "doc_date": "2026-04-01",
            "author": "Кассир",
            "received_from": "Сам клиент",
            "corr_account": "40.10",
            "subconto1": "ООО АЛЬФА",  # same client as K-001 — should aggregate uniquely
            "amount_local": 500000.0,
            "currency_code": "UZS",
        },
        # Totals row — no V marker, must be dropped
        {
            "marker": "",
            "doc_number": "",
            "doc_date": "",
            "amount_local": 3280000.0,  # totals
        },
    ]
    file_bytes = _build_cash_xlsx(fixture)

    result = parse_cash_xls(file_bytes, filename_hint="kassa_2026-04-01.xlsx")

    assert result["ok"] is True, f"parse failed: {result.get('error')}"
    payments = result["payments"]
    assert len(payments) == 3, f"expected 3 payments, got {len(payments)}: {payments}"

    # First payment — UZS leg from corr_account 40.10
    p1 = payments[0]
    assert p1["doc_number_1c"] == "K-001"
    assert p1["currency"] == "UZS"
    assert p1["amount_local"] == 1500000.0
    # CRITICAL: client key is Субконто1, not 'Принято от' (received_from)
    assert p1["client_name_1c"] == "ООО АЛЬФА"
    assert p1["received_from"] == "Иванов Иван"

    # Second payment — USD leg from corr_account 40.11
    p2 = payments[1]
    assert p2["doc_number_1c"] == "K-002"
    assert p2["currency"] == "USD"
    assert p2["amount_currency"] == 100.0
    assert p2["fx_rate"] == 12800.0
    assert p2["client_name_1c"] == "ИП БЕТА"

    # Third payment — UZS to same client as p1
    p3 = payments[2]
    assert p3["doc_number_1c"] == "K-003"
    assert p3["currency"] == "UZS"
    assert p3["amount_local"] == 500000.0
    assert p3["client_name_1c"] == "ООО АЛЬФА"

    # Stats
    stats = result["stats"]
    assert stats["row_count"] == 3
    assert stats["client_count"] == 2  # АЛЬФА + БЕТА
    assert stats["total_uzs"] == 2000000.0  # 1.5M + 0.5M
    assert stats["total_usd"] == 100.0


def test_parse_cash_rejects_unmarked_rows():
    """Rows without a marker in column 0 (totals, separators, blank) must
    be silently dropped — not parsed as payments and not raised as errors.
    Regression guard: an earlier bug parsed totals rows and double-counted them.
    """
    fixture = [
        {  # No marker → drop
            "marker": "",
            "doc_number": "totals",
            "doc_date": "2026-04-01",
            "corr_account": "40.10",
            "subconto1": "TOTALS",
            "amount_local": 999999.0,
        },
        {  # Marker present → keep
            "marker": "V",
            "doc_number": "K-100",
            "doc_date": "2026-04-01",
            "corr_account": "40.10",
            "subconto1": "Real Client",
            "amount_local": 50000.0,
        },
    ]
    result = parse_cash_xls(_build_cash_xlsx(fixture), filename_hint="kassa.xlsx")
    assert result["ok"] is True
    assert len(result["payments"]) == 1
    assert result["payments"][0]["doc_number_1c"] == "K-100"
    assert result["payments"][0]["client_name_1c"] == "Real Client"


def test_parse_cash_unknown_corr_account_falls_back_to_uzs():
    """corr_account '40.10' → UZS, '40.11' → USD, anything else → UZS fallback.
    Pin the fallback so a future "smart" rewrite can't silently drop USD
    payments by mis-detecting a malformed account code.
    """
    fixture = [
        {
            "marker": "V",
            "doc_number": "K-X",
            "doc_date": "2026-04-01",
            "corr_account": "99.99",  # unknown
            "subconto1": "Fallback Client",
            "amount_local": 1000.0,
            "amount_currency": 0,
        },
    ]
    result = parse_cash_xls(_build_cash_xlsx(fixture), filename_hint="kassa.xlsx")
    assert result["ok"] is True
    assert result["payments"][0]["currency"] == "UZS"


def test_parse_cash_missing_header_returns_error():
    """A file without the expected Cyrillic headers should return ok=False
    with a diagnostic — must NOT raise, must NOT silently parse zero rows
    (which would look like 'no payments today' to the operator).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Long enough to bypass the "< 3 rows" early-exit, but no Cyrillic headers
    # anywhere — exercises the header-detection failure path specifically.
    for r in range(1, 20):
        ws.cell(row=r, column=1, value=f"junk row {r}")
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_cash_xls(buf.getvalue(), filename_hint="empty.xlsx")
    assert result["ok"] is False
    # Parser must give the operator something actionable, not a silent zero.
    assert result.get("error"), "ok=False with no error message would be silent failure"
    assert "header" in result["error"].lower() or "column" in result["error"].lower()
