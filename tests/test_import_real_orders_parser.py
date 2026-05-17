"""Smoke test for the Реализация товаров (1C sales/shipments) parser.

Pins the parsing rules that have caused recurring incidents:
  - Header detection scans rows 0..25 for 'Контрагент' + ('Номер'|'Дата')
  - Document header rows have a non-empty marker ('V') in column 0;
    line-item rows have a blank column 0
  - Line-level dual currency: UZS leg uses Цена/Сумма/Всего;
    USD leg uses ЦенаВалютная/СуммаВал/ВсегоВал
  - Post-process fallback chain derives missing item totals:
      total_local := total_local || (sum_local + vat) || (price * quantity)
      total_currency := total_currency || sum_currency || (price_currency * quantity)
  - Doc-header totals derived from SUM(items) when missing
  - Missing header column returns ok=False with diagnostic (not silent zero)

Per memory `no_data_loss` (zero data loss red line) and Error Log #49
(`CURRENCY_LEG_DROPPED_AT_UI`), this is the importer with the biggest
blast radius if a column rename slips through silently.

The single-invoice Счет-фактура variant of the parser is covered
separately by tests/test_invoice_parser.py — this file covers the
bulk tabular Реализация report (multi-document layout).

Synthetic .xlsx fixtures only — no PII committed.
"""
import io

import openpyxl

from backend.services.import_real_orders import parse_real_orders_xls


# Header captions exactly as 1C exports them (matched against _HEADER_FIELDS
# / _ITEM_FIELDS synonyms by lowercased exact-or-prefix match in the parser).
_HEADER_CAPTIONS = [
    (1, ""),                 # column 0: V-marker column (no caption)
    (2, "Номер"),
    (3, "Дата"),
    (4, "Время"),
    (5, "Контрагент"),       # <- header detection looks for this exact word
    (6, "Договор"),
    (7, "Курс"),
    (8, "Валюта"),
    # Item-level captions on the same header row (some shifted right
    # of the doc-level captions). The parser tolerates this layout.
    (9, "Товар"),
    (10, "Количество"),
    (11, "Цена"),
    (12, "Сумма"),
    (13, "НДС"),
    (14, "Всего"),
    (15, "СуммаВал"),
    (16, "Цена валютная"),
    (17, "ВсегоВал"),
]


def _build_real_orders_xlsx(documents: list[dict]) -> bytes:
    """Construct a synthetic bulk Реализация report as .xlsx bytes.

    Each `documents[i]` dict has:
        doc_number, doc_date, client_name, currency, exchange_rate,
        items: list of dicts with product, qty, price, sum_local, vat,
        total_local, sum_currency, price_currency, total_currency

    Row layout produced:
        Row 1:  title (ignored by parser)
        Row 2:  blank
        Row 3:  header row (parser finds 'Контрагент' here)
        Row 4+: alternating V-marker header rows + blank-marker item rows
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реализация"

    ws.cell(row=1, column=2, value="Отчёт по реализации товаров")

    for col, val in _HEADER_CAPTIONS:
        ws.cell(row=3, column=col, value=val)

    r = 4
    for doc in documents:
        # Document header row — 'V' marker in column 1 (openpyxl is 1-indexed
        # so col=1 == parser's column 0).
        ws.cell(row=r, column=1, value="V")
        ws.cell(row=r, column=2, value=doc.get("doc_number", ""))
        ws.cell(row=r, column=3, value=doc.get("doc_date", ""))
        ws.cell(row=r, column=5, value=doc.get("client_name", ""))
        ws.cell(row=r, column=6, value=doc.get("contract", ""))
        ws.cell(row=r, column=7, value=doc.get("exchange_rate", 0))
        ws.cell(row=r, column=8, value=doc.get("currency", ""))
        r += 1

        for item in doc.get("items", []):
            # Line-item row — column 1 left blank (no marker)
            ws.cell(row=r, column=9,  value=item.get("product", ""))
            ws.cell(row=r, column=10, value=item.get("qty", 0))
            ws.cell(row=r, column=11, value=item.get("price", 0))
            ws.cell(row=r, column=12, value=item.get("sum_local", 0))
            ws.cell(row=r, column=13, value=item.get("vat", 0))
            ws.cell(row=r, column=14, value=item.get("total_local", 0))
            ws.cell(row=r, column=15, value=item.get("sum_currency", 0))
            ws.cell(row=r, column=16, value=item.get("price_currency", 0))
            ws.cell(row=r, column=17, value=item.get("total_currency", 0))
            r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_real_orders_multi_doc_dual_currency():
    """Two documents: one UZS-only, one USD-only. Pins the currency-leg
    parsing that has caused CURRENCY_LEG_DROPPED_AT_UI #49 family bugs."""
    documents = [
        {
            # UZS document — Цена/Сумма/Всего populated, currency cols zero
            "doc_number": "R-001",
            "doc_date": "2026-04-01",
            "client_name": "ООО АЛЬФА",
            "currency": "UZS",
            "exchange_rate": 1,
            "items": [
                {
                    "product": "Цемент М400",
                    "qty": 10,
                    "price": 50000,
                    "sum_local": 500000,
                    "vat": 60000,
                    "total_local": 560000,
                },
                {
                    "product": "Кирпич красный",
                    "qty": 1000,
                    "price": 800,
                    "sum_local": 800000,
                    "vat": 96000,
                    "total_local": 896000,
                },
            ],
        },
        {
            # USD document — currency cols populated, UZS cols zero
            "doc_number": "R-002",
            "doc_date": "2026-04-02",
            "client_name": "ИП БЕТА",
            "currency": "USD",
            "exchange_rate": 12800,
            "items": [
                {
                    "product": "Профиль 50х50",
                    "qty": 5,
                    "price_currency": 120,
                    "sum_currency": 600,
                    "total_currency": 600,
                },
            ],
        },
    ]

    result = parse_real_orders_xls(
        _build_real_orders_xlsx(documents),
        filename_hint="realorders_2026-04.xlsx",
    )

    assert result["ok"] is True, f"parse failed: {result.get('error')}"
    docs = result["documents"]
    assert len(docs) == 2, f"expected 2 docs, got {len(docs)}: {[d['doc_number_1c'] for d in docs]}"

    # Doc 1 — UZS
    d1 = docs[0]
    assert d1["doc_number_1c"] == "R-001"
    assert d1["client_name_1c"] == "ООО АЛЬФА"
    assert d1["currency"] == "UZS"
    assert len(d1["items"]) == 2
    i1 = d1["items"][0]
    assert i1["product_name_1c"] == "Цемент М400"
    assert i1["quantity"] == 10
    assert i1["price"] == 50000
    assert i1["sum_local"] == 500000
    assert i1["total_local"] == 560000

    # Doc 2 — USD leg must be parsed from СуммаВал/ВсегоВал, not Сумма/Всего
    d2 = docs[1]
    assert d2["doc_number_1c"] == "R-002"
    assert d2["currency"] == "USD"
    assert d2["exchange_rate"] == 12800
    assert len(d2["items"]) == 1
    i2 = d2["items"][0]
    assert i2["product_name_1c"] == "Профиль 50х50"
    assert i2["quantity"] == 5
    # CRITICAL: USD leg landed in the right fields. If the parser ever
    # silently drops the currency columns, these three lines fail loudly.
    assert i2["price_currency"] == 120
    assert i2["sum_currency"] == 600
    assert i2["total_currency"] == 600
    # And the UZS leg stayed at 0 (no cross-contamination)
    assert i2["sum_local"] == 0
    assert i2["total_local"] == 0

    # Stats — pins the aggregate shape the bot summary depends on
    stats = result["stats"]
    assert stats["doc_count"] == 2
    assert stats["item_count"] == 3
    assert stats["client_count"] == 2
    assert stats["product_count"] == 3
    assert stats["total_local"] == 560000 + 896000  # both UZS items
    assert stats["total_currency"] == 600           # one USD item


def test_parse_real_orders_total_local_fallback():
    """Items missing total_local must derive it from sum_local+vat or price*qty.
    Pins the post-process logic at parse_real_orders_xls:702+; a silent
    regression here would render blank prices in the Cabinet (Error Log #49
    sibling — both share the 'derived total dropped silently' shape).
    """
    documents = [
        {
            "doc_number": "R-FB",
            "doc_date": "2026-04-03",
            "client_name": "Fallback Test",
            "currency": "UZS",
            "exchange_rate": 1,
            "items": [
                {
                    # Only price+qty filled; sum_local, total_local must be derived
                    "product": "Песок",
                    "qty": 4,
                    "price": 25000,
                    # vat / sum_local / total_local left 0 → must be derived
                },
            ],
        },
    ]
    result = parse_real_orders_xls(
        _build_real_orders_xlsx(documents),
        filename_hint="realorders_fallback.xlsx",
    )
    assert result["ok"] is True, f"parse failed: {result.get('error')}"
    item = result["documents"][0]["items"][0]
    # sum_local derived from price * qty
    assert item["sum_local"] == 100000.0, f"expected 100000.0, got {item['sum_local']}"
    # total_local derived from sum_local + vat (vat=0 here)
    assert item["total_local"] == 100000.0, f"expected 100000.0, got {item['total_local']}"
    # Doc-header total derived from SUM(items.total_local)
    assert result["documents"][0]["total_sum"] == 100000.0


def test_parse_real_orders_skips_unmarked_noise():
    """A row with the 'V' marker but no client_name and no doc_number is
    noise — must reset `current` to None so subsequent item rows don't
    attach to a stale document.
    """
    # Build manually to inject a noise row that the helper doesn't emit
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реализация"
    ws.cell(row=1, column=2, value="Отчёт")
    for col, val in _HEADER_CAPTIONS:
        ws.cell(row=3, column=col, value=val)
    # Noise row — has marker but no client / no doc number
    ws.cell(row=4, column=1, value="V")
    # Then a real document
    ws.cell(row=5, column=1, value="V")
    ws.cell(row=5, column=2, value="R-X")
    ws.cell(row=5, column=3, value="2026-04-01")
    ws.cell(row=5, column=5, value="Real Client")
    ws.cell(row=5, column=8, value="UZS")
    ws.cell(row=6, column=9, value="Product")
    ws.cell(row=6, column=10, value=1)
    ws.cell(row=6, column=11, value=100)
    ws.cell(row=6, column=12, value=100)
    ws.cell(row=6, column=14, value=100)
    buf = io.BytesIO()
    wb.save(buf)

    result = parse_real_orders_xls(buf.getvalue(), filename_hint="noise.xlsx")
    assert result["ok"] is True
    # Only the real document survives — the noise row didn't seed a doc
    assert len(result["documents"]) == 1
    assert result["documents"][0]["doc_number_1c"] == "R-X"
    assert len(result["documents"][0]["items"]) == 1


def test_parse_real_orders_missing_kontragent_returns_error():
    """A file with no 'Контрагент' anywhere in rows 0..25 cannot be a
    Реализация report — must return ok=False with a diagnostic, never a
    silent ok=True with zero documents (which would look like 'no sales today'
    to the operator).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Long enough to bypass the "< 5 rows" early-exit
    for r in range(1, 30):
        ws.cell(row=r, column=1, value=f"junk row {r}")
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_real_orders_xls(buf.getvalue(), filename_hint="empty.xlsx")
    assert result["ok"] is False
    assert result.get("error"), "ok=False with no error message is silent failure"
    assert "контрагент" in result["error"].lower() or "header" in result["error"].lower()
