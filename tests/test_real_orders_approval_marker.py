"""Tests for v18 col-0 approval marker capture in real_orders.

Pins the V/X parsing + INSERT/UPDATE semantics introduced 2026-05-27:
  - 'V' in col 0 -> is_approved=1, first_pending_at=NULL
  - 'X' in col 0 -> is_approved=0, first_pending_at=now
  - X→V transition preserves first_pending_at (COALESCE)
  - V→X transition sets first_pending_at on the row that previously had NULL

Synthetic XLSX fixtures — no PII committed.
"""
import io

import openpyxl

from backend.services.import_real_orders import (
    apply_real_orders_import,
    parse_real_orders_xls,
)


_HEADER_CAPTIONS = [
    (1, ""),
    (2, "Номер"),
    (3, "Дата"),
    (5, "Контрагент"),
    (7, "Курс"),
    (8, "Валюта"),
    (9, "Товар"),
    (10, "Количество"),
    (11, "Цена"),
    (12, "Сумма"),
    (13, "НДС"),
    (14, "Всего"),
    (15, "СуммаВал"),
    (16, "Цена валютная"),
    (17, "ВсегоВал"),
]


def _xlsx(documents):
    """`documents` is a list of (marker, doc_number, client, items) tuples
    where `items` is a list of (product, qty, total_local, total_currency).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="Отчёт по реализации товаров")
    for col, val in _HEADER_CAPTIONS:
        ws.cell(row=3, column=col, value=val)
    r = 4
    for marker, doc_no, client, items in documents:
        ws.cell(row=r, column=1, value=marker)
        ws.cell(row=r, column=2, value=doc_no)
        ws.cell(row=r, column=3, value="2026-05-27")
        ws.cell(row=r, column=5, value=client)
        ws.cell(row=r, column=7, value=1)
        ws.cell(row=r, column=8, value="UZS")
        r += 1
        for product, qty, total_l, total_c in items:
            ws.cell(row=r, column=9, value=product)
            ws.cell(row=r, column=10, value=qty)
            ws.cell(row=r, column=11, value=total_l)
            ws.cell(row=r, column=12, value=total_l)
            ws.cell(row=r, column=14, value=total_l)
            ws.cell(row=r, column=15, value=total_c)
            ws.cell(row=r, column=17, value=total_c)
            r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parser_captures_v_marker():
    docs = parse_real_orders_xls(_xlsx([("V", "T-100", "TestClient", [("Product A", 5, 50000, 5)])]))
    assert docs["ok"] is True
    assert len(docs["documents"]) == 1
    assert docs["documents"][0]["is_approved"] == 1


def test_parser_captures_x_marker():
    docs = parse_real_orders_xls(_xlsx([("X", "T-200", "TestClient", [("Product B", 3, 30000, 3)])]))
    assert docs["ok"] is True
    assert docs["documents"][0]["is_approved"] == 0


def test_parser_treats_unknown_marker_as_null():
    docs = parse_real_orders_xls(_xlsx([("Z", "T-300", "TestClient", [("Product C", 1, 10000, 1)])]))
    assert docs["ok"] is True
    assert docs["documents"][0]["is_approved"] is None


def test_v_insert_then_query(db):
    res = apply_real_orders_import(_xlsx([("V", "I-100", "Client A", [("Prod", 1, 5000, 0.5)])]))
    assert res["ok"] is True
    row = db.execute(
        "SELECT is_approved, first_pending_at FROM real_orders WHERE doc_number_1c='I-100'"
    ).fetchone()
    assert row["is_approved"] == 1
    assert row["first_pending_at"] is None


def test_x_insert_records_first_pending_at(db):
    res = apply_real_orders_import(_xlsx([("X", "I-200", "Client B", [("Prod", 1, 5000, 0.5)])]))
    assert res["ok"] is True
    row = db.execute(
        "SELECT is_approved, first_pending_at FROM real_orders WHERE doc_number_1c='I-200'"
    ).fetchone()
    assert row["is_approved"] == 0
    assert row["first_pending_at"] is not None
    assert len(row["first_pending_at"]) >= 10  # 'YYYY-MM-DD ...'


def test_x_to_v_transition_preserves_first_pending_at(db):
    # First import as X
    apply_real_orders_import(_xlsx([("X", "I-300", "Client C", [("Prod", 1, 5000, 0.5)])]))
    original = db.execute(
        "SELECT is_approved, first_pending_at FROM real_orders WHERE doc_number_1c='I-300'"
    ).fetchone()
    assert original["is_approved"] == 0
    assert original["first_pending_at"] is not None
    original_pending_at = original["first_pending_at"]

    # Re-import same doc as V (X→V transition — approved + shipped)
    apply_real_orders_import(_xlsx([("V", "I-300", "Client C", [("Prod", 1, 5000, 0.5)])]))
    updated = db.execute(
        "SELECT is_approved, first_pending_at FROM real_orders WHERE doc_number_1c='I-300'"
    ).fetchone()
    assert updated["is_approved"] == 1
    # COALESCE preserves the original first-pending timestamp.
    assert updated["first_pending_at"] == original_pending_at


def test_repull_diff_flags_edited_order(db):
    # /realordersweek: a re-import whose total dropped = items removed/returned
    # in 1C → must surface in repull_report.edited (the daily feed silently
    # overwrites this; the re-pull report makes it visible).
    apply_real_orders_import(_xlsx([("V", "I-500", "Client E", [("Prod", 2, 50000, 5)])]))
    res = apply_real_orders_import(_xlsx([("V", "I-500", "Client E", [("Prod", 1, 30000, 3)])]))
    edited = res["repull_report"]["edited"]
    hit = next((e for e in edited if e["doc"] == "I-500"), None)
    assert hit is not None
    assert hit["old_uzs"] == 50000 and hit["new_uzs"] == 30000
    assert hit["old_usd"] == 5 and hit["new_usd"] == 3


def test_repull_diff_flags_swept_order(db):
    # An order present last import but absent now (same date window) = deleted in
    # 1C → swept + reported.
    apply_real_orders_import(_xlsx([
        ("V", "I-600", "C", [("P", 1, 10000, 1)]),
        ("V", "I-601", "C", [("P", 1, 20000, 2)]),
    ]))
    res = apply_real_orders_import(_xlsx([("V", "I-600", "C", [("P", 1, 10000, 1)])]))
    swept = res["repull_report"]["swept"]
    assert any(s["doc"] == "I-601" for s in swept)
    assert db.execute("SELECT COUNT(*) FROM real_orders WHERE doc_number_1c='I-601'").fetchone()[0] == 0


def test_repull_diff_clean_when_unchanged(db):
    apply_real_orders_import(_xlsx([("V", "I-700", "C", [("P", 1, 10000, 1)])]))
    res = apply_real_orders_import(_xlsx([("V", "I-700", "C", [("P", 1, 10000, 1)])]))
    assert res["repull_report"]["edited"] == []
    assert res["repull_report"]["swept"] == []


def test_v_to_x_transition_sets_first_pending_at(db):
    # First import as V (no pending history yet)
    apply_real_orders_import(_xlsx([("V", "I-400", "Client D", [("Prod", 1, 5000, 0.5)])]))
    original = db.execute(
        "SELECT is_approved, first_pending_at FROM real_orders WHERE doc_number_1c='I-400'"
    ).fetchone()
    assert original["is_approved"] == 1
    assert original["first_pending_at"] is None

    # Re-import same doc as X (V→X — e.g., approval was rolled back in 1C)
    apply_real_orders_import(_xlsx([("X", "I-400", "Client D", [("Prod", 1, 5000, 0.5)])]))
    updated = db.execute(
        "SELECT is_approved, first_pending_at FROM real_orders WHERE doc_number_1c='I-400'"
    ).fetchone()
    assert updated["is_approved"] == 0
    assert updated["first_pending_at"] is not None
