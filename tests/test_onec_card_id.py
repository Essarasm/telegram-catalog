"""Client Identity Anchoring — Phase 0a (onec_card_id stable card anchor).

The daily importer now resolves an existing allowed_clients row by the 1C card
anchor ``onec_card_id = "{folder}:{Код}"`` BEFORE phone/name. A client whose
phone changed or got corrupted is still recognised → no silent duplicate INSERT
(ends the #74/#75/#81 recurring family).

Two surfaces under test:
  - ``_apply_folder_anchor`` — folder-tracking parse that stamps the anchor and
    drops folder-header rows (and is a strict no-op for non-1C uploads).
  - ``_upsert_client_from_row`` — resolve-by-card precedence, rename tolerance,
    fill-only anchor capture, and the anchor-conflict review flag.
"""
import io

import pytest

from backend.services.import_clients import (
    _upsert_client_from_row,
    _apply_folder_anchor,
    _normalize_card_code,
    _iter_rows_from_xlsx,
)


# ── helpers ──────────────────────────────────────────────────────────────────

def _seed(db, cid, name, phone, card=None, gps=None):
    db.execute(
        "INSERT INTO allowed_clients (id, client_id_1c, name, phone_normalized, "
        "onec_card_id, gps_latitude, status) VALUES (?, ?, ?, ?, ?, ?, 'active')",
        (cid, name, name, phone, card, gps),
    )


def _upsert(db, phone, cid_1c=None, name=None, card=None):
    return _upsert_client_from_row(
        db, raw_phone_str=phone, client_name=(name or cid_1c or "X"),
        location="", source="clients_upload", cid_1c=(cid_1c or ""),
        company="", changed_by_tag="test", onec_card_id=card,
    )


def _active_count(db):
    return db.execute(
        "SELECT COUNT(*) FROM allowed_clients "
        "WHERE COALESCE(status,'active') NOT LIKE 'merged%'"
    ).fetchone()[0]


def _row(db, cid):
    return db.execute(
        "SELECT client_id_1c, phone_normalized, onec_card_id, needs_review "
        "FROM allowed_clients WHERE id = ?", (cid,)
    ).fetchone()


def _drift_queue(db):
    return db.execute(
        "SELECT * FROM client_identity_drift_queue WHERE resolved = 0"
    ).fetchall()


# ── _normalize_card_code ──────────────────────────────────────────────────────

@pytest.mark.parametrize("raw, expected", [
    (1056, "1056"),
    (1056.0, "1056"),       # xlrd yields floats
    ("1056", "1056"),
    ("1056.0", "1056"),     # stringified float
    ("  638 ", "638"),
    (None, ""),
    ("", ""),
])
def test_normalize_card_code(raw, expected):
    assert _normalize_card_code(raw) == expected


# ── resolver precedence ────────────────────────────────────────────────────────

def test_resolve_by_card_id_before_phone(db):
    # Existing row anchored to a card; 1C export now shows a DIFFERENT phone for
    # the same card. Card-id resolves first → same row updated, no duplicate.
    _seed(db, 1, "АБДУЛЛО ЯНГИ", "915490095", card="Прочие:1056")
    out, rid = _upsert(db, "915194019", name="АБДУЛЛО ЯНГИ", card="Прочие:1056")
    assert out == "updated" and rid == 1
    assert _active_count(db) == 1                       # NO duplicate spawned
    assert _row(db, 1)["phone_normalized"] == "915194019"


def test_card_id_match_allows_rename_no_drift(db):
    # Curated row (pin) whose 1C card was renamed. A phone match would HOLD
    # (#74 drift guard), but a card-id match is definitively the same client →
    # the rename goes through, nothing is held.
    _seed(db, 5, "OLD NAME", "971112233", card="Прочие:200", gps=39.6)
    out, rid = _upsert(db, "971112233", cid_1c="NEW NAME", card="Прочие:200")
    assert out == "updated" and rid == 5
    assert _row(db, 5)["client_id_1c"] == "NEW NAME"    # name updated, not frozen
    assert _drift_queue(db) == []                        # not held


def test_card_id_filled_on_phone_match(db):
    # Legacy row has no anchor yet; import matches by phone and back-fills the
    # card id (the natural backfill that happens on every import).
    _seed(db, 7, "СОМ", "900000001", card=None)
    out, rid = _upsert(db, "900000001", name="СОМ", card="Прочие:42")
    assert out == "updated" and rid == 7
    assert _row(db, 7)["onec_card_id"] == "Прочие:42"


def test_card_id_conflict_flags_review_no_churn(db):
    # Phone match lands on a row already anchored to a DIFFERENT card → the
    # incoming card wasn't found by the card lookup, so this is a genuine
    # conflict. Don't churn the stored anchor; flag for review.
    _seed(db, 9, "X", "933334444", card="Прочие:111")
    out, rid = _upsert(db, "933334444", name="X", card="Прочие:999")
    assert out == "updated" and rid == 9
    assert _row(db, 9)["onec_card_id"] == "Прочие:111"   # unchanged
    assert _row(db, 9)["needs_review"] == 1


def test_new_client_insert_stamps_card_id(db):
    out, rid = _upsert(db, "918887766", name="ЯНГИ", card="Прочие:1500")
    assert out == "inserted"
    assert _row(db, rid)["onec_card_id"] == "Прочие:1500"


def test_unique_index_blocks_two_active_same_card(db):
    # The partial UNIQUE index is the structural tripwire: two ACTIVE rows can
    # never share a card id.
    import sqlite3
    _seed(db, 11, "A", "900000011", card="Прочие:777")
    with pytest.raises(sqlite3.IntegrityError):
        _seed(db, 12, "B", "900000012", card="Прочие:777")


def test_merged_row_does_not_collide_on_card(db):
    # A merged (retired) row keeps its card id without blocking the survivor.
    db.execute(
        "INSERT INTO allowed_clients (id, name, phone_normalized, onec_card_id, status) "
        "VALUES (13, 'OLD', '900000013', 'Прочие:888', 'merged_into_14')"
    )
    _seed(db, 14, "NEW", "900000014", card="Прочие:888")  # must NOT raise
    assert _active_count(db) == 1


# ── _apply_folder_anchor (parse) ───────────────────────────────────────────────

def _xlsx_bytes(rows):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_folder_walk_stamps_anchor_and_drops_headers():
    rows = [
        ['Справочник "Контрагенты"', "", "", ""],            # title (row 0)
        ["Код", "Наименование", "Вид контрагента", "Телефоны контрагента"],
        [20, "Покупатели", "", ""],                          # folder header
        [15, "АКРАМ Челак", "Организация", "97 111 11 11"],
        [19, "Поставщики", "", ""],                          # folder header
        [638, "COLOREX", "Организация", "97 222 22 22"],
        [1, "Прочие", "", ""],                               # folder header
        [1056, "АБДУЛЛО ЯНГИ", "Организация", "91 549 00 95"],
    ]
    _hdr, out = _iter_rows_from_xlsx(_xlsx_bytes(rows))
    # 3 data rows survive; 3 folder headers + title dropped.
    assert len(out) == 3
    cards = {r["name"]: r.get("onec_card_id") for r in out}
    assert cards["АКРАМ Челак"] == "Покупатели:15"
    assert cards["COLOREX"] == "Поставщики:638"
    assert cards["АБДУЛЛО ЯНГИ"] == "Прочие:1056"          # de-collided by folder


def test_no_card_columns_is_strict_noop():
    # A non-1C upload (Client Master / manual phone-fix sheet) without Код / Вид
    # columns must be returned untouched — no row dropped, no anchor stamped.
    rows = [
        ["Наименование", "Телефоны контрагента"],
        ["ШОП А", "97 111 11 11"],
        ["ШОП Б", "97 222 22 22"],
    ]
    _hdr, out = _iter_rows_from_xlsx(_xlsx_bytes(rows))
    assert len(out) == 2
    assert all("onec_card_id" not in r for r in out)


def test_importer_adopts_pending_registration_row(db):
    # Phase 3 foundation: a user-first registration creates a pending row (no
    # card id, no 1C name). When the daily 1C import later sees that phone under
    # a real card, it must ADOPT the same row (stamp card id + name) — not INSERT
    # a competing duplicate.
    from backend.services.client_resolver import resolve_for_registration
    from backend.services.import_clients import _upsert_client_from_row

    res = resolve_for_registration(db, telegram_id=777, phone="998 90 111 22 33",
                                   name="Yangi Dukon", source="bot_new_client")
    pid = res["client_id"]
    assert res["action"] == "created"
    pre = db.execute("SELECT onec_card_id, client_id_1c FROM allowed_clients WHERE id=?",
                     (pid,)).fetchone()
    assert pre[0] is None  # pending: no card id yet
    before = db.execute("SELECT COUNT(*) FROM allowed_clients").fetchone()[0]

    out, rid = _upsert_client_from_row(
        db, raw_phone_str="90 111 22 33", client_name="ДУКОН 1С", location="",
        source="clients_upload", cid_1c="ДУКОН 1С", company="",
        changed_by_tag="import", onec_card_id="Прочие:5000",
    )
    assert rid == pid and out == "updated"  # adopted the SAME row
    assert db.execute("SELECT COUNT(*) FROM allowed_clients").fetchone()[0] == before
    row = db.execute("SELECT onec_card_id, client_id_1c FROM allowed_clients WHERE id=?",
                     (pid,)).fetchone()
    assert row[0] == "Прочие:5000" and row[1] == "ДУКОН 1С"


def test_apply_folder_anchor_noop_without_both_columns():
    # Only onec_code present, no onec_vid → no-op (can't track folders safely).
    rows = [{"onec_code": 1, "name": "X", "phone": "900000001"}]
    assert _apply_folder_anchor(["onec_code", "name", "phone"], rows) == rows
