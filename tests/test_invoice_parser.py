"""Pin the Счет-фактура (single invoice) layout parser.

Builds synthetic .xlsx files in memory to mirror the two real-world shapes
seen from 1C — bare 1C card name (real client) and parens-with-phone
(walk-in / one-off). No PII fixtures committed.
"""
import io

import openpyxl

from backend.services.import_real_orders import parse_real_orders_xls


def _build_invoice_xlsx(*, doc_no: str, date_words: str, client_line: str, items: list) -> bytes:
    """Construct a Счет-фактура printable invoice as .xlsx bytes.

    Layout mirrors what 1C exports: row 0 title, row 1 date, row 3 client,
    row 6 column headers (with the duplicated 'Цена'/'Стоимость' for UZS
    and USD pairs), then item rows, then 'Итого:' totals row.

    `items` is a list of dicts with keys: product, qty,
    price_uzs, total_uzs, price_usd, total_usd. Use 0 for the inactive
    pair (matches 1C's blank-cell convention; '-' is also tolerated by
    the parser but float-zero is closer to real openpyxl output).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 1-indexed: openpyxl row/col are 1-based. Parser uses 0-based.
    # Row 1 (index 0) — title. Cell B1 (col 2) holds the title text.
    ws.cell(row=1, column=2, value=f"СЧЕТ-ФАКТУРА  №  {doc_no}  ")
    ws.cell(row=2, column=2, value=date_words)
    ws.cell(row=4, column=2, value=client_line)

    # Row 7 (index 6) — column headers
    headers = [
        (2, "№"),
        (3, "Наименование товаров \n(работ, услуг)"),
        (5, "Склад"),
        (6, "Ед.\nизм"),
        (7, "Кол-\nво"),
        (8, "Упаковка"),
        (9, "Цена"),
        (10, "Стоимость\nпоставки"),
        (11, "Цена"),
        (12, "Стоимость\nпоставки"),
        (13, "Вес"),
        (14, "Общий \nвес"),
    ]
    for col, val in headers:
        ws.cell(row=7, column=col, value=val)

    # Item rows starting at row 9 (index 8)
    item_start_row = 9
    for i, it in enumerate(items):
        r = item_start_row + i
        ws.cell(row=r, column=2, value=i + 1)
        ws.cell(row=r, column=3, value=it["product"])
        ws.cell(row=r, column=5, value="Основной склад")
        ws.cell(row=r, column=6, value="шт")
        ws.cell(row=r, column=7, value=it["qty"])
        ws.cell(row=r, column=8, value=" ")
        ws.cell(row=r, column=9, value=it["price_uzs"] or " ")
        ws.cell(row=r, column=10, value=it["total_uzs"] or " ")
        ws.cell(row=r, column=11, value=it["price_usd"] or " ")
        ws.cell(row=r, column=12, value=it["total_usd"] or " ")

    # Итого row
    totals_row = item_start_row + len(items)
    ws.cell(row=totals_row, column=8, value="Итого:")
    ws.cell(row=totals_row, column=10, value=sum(it["total_uzs"] for it in items) or "-")
    ws.cell(row=totals_row, column=12, value=sum(it["total_usd"] for it in items))

    # Footer line — total in words (not parsed, just shape parity)
    ws.cell(row=totals_row + 2, column=2, value="Всего к оплате в сумах: ...")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_invoice_parser_real_client_usd_only():
    """Bare 1C card name + USD-only items (matches sample_single_order_01)."""
    items = [
        {"product": "Проволока оцинк. /Тош.-0,8/", "qty": 100, "price_uzs": 0, "total_uzs": 0, "price_usd": 1.25, "total_usd": 125.0},
        {"product": "ХАЯТ грунт /1 кг/",          "qty": 10,  "price_uzs": 0, "total_uzs": 0, "price_usd": 2.2,  "total_usd": 22.0},
    ]
    data = _build_invoice_xlsx(
        doc_no="3096",
        date_words="от 5 Мая 2026 г.",
        client_line="Клиент : ФАХРИДДДИН /ЧЕЛЕК/",
        items=items,
    )
    result = parse_real_orders_xls(data, filename_hint="invoice.xlsx")
    assert result["ok"], f"parse failed: {result.get('error')}"
    assert len(result["documents"]) == 1
    doc = result["documents"][0]
    assert doc["doc_number_1c"] == "3096"
    assert doc["doc_date"] == "2026-05-05"
    assert doc["client_name_1c"] == "ФАХРИДДДИН /ЧЕЛЕК/"
    assert doc["client_phone"] is None
    assert doc["currency"] == "USD"
    assert len(doc["items"]) == 2
    assert doc["items"][0]["product_name_1c"] == "Проволока оцинк. /Тош.-0,8/"
    assert doc["items"][0]["quantity"] == 100
    assert doc["items"][0]["total_currency"] == 125.0
    assert doc["items"][0]["total_local"] == 0
    assert result["stats"]["total_usd"] == 147.0
    assert result["stats"]["total_uzs"] == 0


def test_invoice_parser_walkin_with_phone_mixed_currency():
    """Walk-in with phone in nested parens + mixed UZS/USD items."""
    items = [
        {"product": "ГУДФИКС Клей /310Ml/", "qty": 5, "price_uzs": 0, "total_uzs": 0, "price_usd": 1.5, "total_usd": 7.5},
        {"product": "Разбавитель /0,9/",    "qty": 15, "price_uzs": 5300, "total_uzs": 79500, "price_usd": 0, "total_usd": 0},
    ]
    data = _build_invoice_xlsx(
        doc_no="3118",
        date_words="от 5 Мая 2026 г.",
        client_line="Клиент :  Ивонов И. (Ивонов Иван (+998902277176))",
        items=items,
    )
    result = parse_real_orders_xls(data, filename_hint="invoice.xlsx")
    assert result["ok"], f"parse failed: {result.get('error')}"
    doc = result["documents"][0]
    assert doc["doc_number_1c"] == "3118"
    assert doc["client_name_1c"] == "Ивонов И."
    assert doc["client_phone"] == "+998902277176"
    assert doc["currency"] == "MIXED"
    assert len(doc["items"]) == 2
    assert doc["items"][0]["total_currency"] == 7.5
    assert doc["items"][0]["total_local"] == 0
    assert doc["items"][1]["total_local"] == 79500
    assert doc["items"][1]["total_currency"] == 0
    assert result["stats"]["total_uzs"] == 79500
    assert result["stats"]["total_usd"] == 7.5


def test_invoice_parser_dash_total_coerced_to_zero():
    """1C uses literal '-' in the inactive Стоимость column for single-
    currency rows; parser must coerce to 0 (not raise / not stringify)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=2, value="СЧЕТ-ФАКТУРА  №  9999  ")
    ws.cell(row=2, column=2, value="от 1 Января 2026 г.")
    ws.cell(row=4, column=2, value="Клиент : Test Client")
    headers = [(3, "Наименование"), (7, "Кол-во"), (9, "Цена"), (10, "Стоимость поставки"), (11, "Цена"), (12, "Стоимость поставки")]
    for col, val in headers:
        ws.cell(row=7, column=col, value=val)
    ws.cell(row=9, column=3, value="Test Item")
    ws.cell(row=9, column=7, value=10)
    ws.cell(row=9, column=9, value="-")
    ws.cell(row=9, column=10, value="-")
    ws.cell(row=9, column=11, value=5.0)
    ws.cell(row=9, column=12, value=50.0)
    ws.cell(row=10, column=8, value="Итого:")

    buf = io.BytesIO()
    wb.save(buf)
    result = parse_real_orders_xls(buf.getvalue(), filename_hint="invoice.xlsx")
    assert result["ok"]
    item = result["documents"][0]["items"][0]
    assert item["total_local"] == 0.0  # '-' coerced
    assert item["total_currency"] == 50.0


def test_invoice_dispatcher_falls_through_on_bulk_format():
    """Files NOT starting with СЧЕТ-ФАКТУРА fall through to the bulk parser
    and fail with the bulk-format error (no Контрагент header) — confirms
    the dispatch isn't accidentally claiming non-invoice files."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="some other header")
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_real_orders_xls(buf.getvalue(), filename_hint="other.xlsx")
    assert not result["ok"]
    # Should NOT hit the invoice-parser error path (which mentions "Invoice")
    err = result.get("error", "")
    assert "Invoice" not in err, f"Dispatcher wrongly claimed non-invoice file: {err}"
