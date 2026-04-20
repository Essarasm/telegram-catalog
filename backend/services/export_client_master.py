"""Full-mirror Client Master xlsx export.

Produces a single xlsx file covering every row in `allowed_clients` plus
every known mirror field (1C, Mini App live, financials, credit score,
lifecycle, cashback). Editable columns carry an `✏️` marker header so
the operator knows what they may safely edit; all other columns are
read-only (auto-refreshed on the next export).

See obsidian-vault/Client Data Workflow — Design v0.1.md v0.1.3 for full
column spec.

Design invariants:
  * Row count = COUNT(*) from allowed_clients (no filtering — user sees
    the whole universe)
  * ✏️ columns preserved verbatim from DB (they came from the user)
  * 🔒 columns pulled live from authoritative source tables at export
  * `active_status` column reflects `COALESCE(status, 'active')`
"""
from __future__ import annotations

import io
import os
from datetime import datetime, timezone, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.database import get_db

TASHKENT = timezone(timedelta(hours=5))

# Column spec — (display_header, editable, DB-source function or column name)
# The `source` indicates where the value comes from when building the export.
# Editable columns bubble up directly from `allowed_clients` fields.
COLUMNS = [
    # — identifiers & 1C —
    ("№",                              False, "seq"),
    ("1C Наименование",               False, "client_id_1c"),
    ("Полное наименование",           False, "company_name"),

    # — geography (editable; mirrored from Mini App when empty) —
    ("✏️ Viloyat",                    True,  "viloyat"),
    ("✏️ Shahar/Tuman",               True,  "tuman"),
    ("✏️ Mo'ljal",                    True,  "moljal"),
    ("✏️ Izoh",                       True,  "location"),

    # — contact (editable) —
    ("✏️ Ism 01 / Familiya",          True,  "name"),
    ("✏️ Raqam 01",                   True,  "phone_formatted"),  # computed from phone_normalized
    ("✏️ Ism 02",                     True,  "ism_02"),
    ("✏️ Raqam 02",                   True,  "raqam_02"),
    ("✏️ Ism 03",                     True,  "ism_03"),
    ("✏️ Raqam 03",                   True,  "raqam_03"),

    # — operator notes (editable) —
    ("✏️ Mijoz holati",               True,  "mijoz_holati"),
    ("✏️ Hajm",                       True,  "hajm"),
    ("✏️ Segment",                    True,  "segment"),
    ("✏️ Active status",              True,  "active_status"),
    ("✏️ Eslatmalar",                 True,  "eslatmalar"),

    # — Mini App live (read-only) —
    ("Telegram ID",                   False, "matched_telegram_id"),
    ("Telegram username",             False, "tg_username"),
    ("Telegram first_name",           False, "tg_first_name"),
    ("Registered at",                 False, "tg_registered_at"),
    ("Last active",                   False, "tg_last_active"),
    ("Latest GPS lat",                False, "gps_lat"),
    ("Latest GPS lng",                False, "gps_lng"),
    ("Reverse-geo address",           False, "gps_address"),
    ("GPS shared at",                 False, "gps_shared_at"),

    # — activity / financial (read-only) —
    ("Orders 12mo",                   False, "orders_12mo"),
    ("Last order date",               False, "last_order_date"),
    ("Outstanding debt UZS",          False, "debt_uzs"),
    ("Outstanding debt USD",          False, "debt_usd"),
    ("Last payment date",             False, "last_payment_date"),

    # — scoring & lifecycle (read-only) —
    ("Credit score",                  False, "credit_score"),
    ("Score tier",                    False, "score_tier"),
    ("Volume bucket",                 False, "volume_bucket"),

    # — sync-pipeline metadata (read-only) —
    ("needs_review",                  False, "needs_review"),
    ("needs_verification",            False, "needs_verification"),
    ("master_row_id",                 False, "master_row_id"),
    ("allowed_clients.id",            False, "id"),
    ("Last synced",                   False, "last_master_synced_at"),
]


_FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_FILL_EDITABLE = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_FILL_REVIEW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_BORDER = Border(
    left=Side(border_style="thin", color="DDDDDD"),
    right=Side(border_style="thin", color="DDDDDD"),
    top=Side(border_style="thin", color="DDDDDD"),
    bottom=Side(border_style="thin", color="DDDDDD"),
)


def _format_phone(phone_normalized: Optional[str]) -> str:
    """Render 9-digit Uzbek local number as 'NN NNN-NN-NN'."""
    if not phone_normalized or len(phone_normalized) < 9:
        return phone_normalized or ""
    d = phone_normalized[-9:]
    return f"{d[0:2]} {d[2:5]}-{d[5:7]}-{d[7:9]}"


def _fetch_rows(conn) -> list[dict]:
    """Join allowed_clients with all mirror sources in one query for speed.
    Falls back gracefully when a source table is missing / empty."""
    rows = conn.execute(
        """
        SELECT
            ac.id,
            ac.phone_normalized,
            ac.name,
            ac.location,
            ac.client_id_1c,
            ac.company_name,
            COALESCE(ac.status, 'active')       AS active_status,
            ac.matched_telegram_id,
            ac.credit_score,
            ac.segment,
            ac.hajm,
            ac.mijoz_holati,
            ac.eslatmalar,
            ac.ism_02, ac.raqam_02, ac.ism_03, ac.raqam_03,
            ac.viloyat, ac.tuman, ac.moljal,
            ac.needs_review,
            ac.needs_verification,
            ac.master_row_id,
            ac.last_master_synced_at,
            u.username                          AS tg_username,
            u.first_name                        AS tg_first_name,
            u.registered_at                     AS tg_registered_at,
            u.location_updated                  AS tg_last_active,
            u.latitude                          AS gps_lat,
            u.longitude                         AS gps_lng,
            u.location_address                  AS gps_address,
            u.location_updated                  AS gps_shared_at,
            cs.tier                             AS score_tier,
            cs.volume_bucket                    AS volume_bucket
        FROM allowed_clients ac
        LEFT JOIN users u  ON u.telegram_id = ac.matched_telegram_id
        LEFT JOIN (
            SELECT client_id, tier, volume_bucket,
                   ROW_NUMBER() OVER (PARTITION BY client_id ORDER BY recalc_date DESC) AS rn
            FROM client_scores
        ) cs ON cs.client_id = ac.id AND cs.rn = 1
        ORDER BY ac.id
        """
    ).fetchall()
    return [dict(r) for r in rows]


def _enrich_financials(conn, data: list[dict]) -> None:
    """Populate debt/last-order/last-payment per client_id_1c (name-keyed)."""
    # Orders 12mo + last order
    cutoff_12mo = (datetime.now(TASHKENT).date() - timedelta(days=365)).isoformat()
    orders_map = {}
    for r in conn.execute(
        """SELECT ro.client_id,
                  COUNT(*) FILTER (WHERE ro.doc_date >= ?) AS orders_12mo,
                  MAX(ro.doc_date)                         AS last_order_date
           FROM real_orders ro
           WHERE ro.client_id IS NOT NULL
           GROUP BY ro.client_id""",
        (cutoff_12mo,),
    ):
        orders_map[r["client_id"]] = (r["orders_12mo"] or 0, r["last_order_date"])

    # Outstanding debt from latest client_debts snapshot per client
    debt_map = {}
    try:
        for r in conn.execute(
            """SELECT client_id, currency, balance, period_end,
                      ROW_NUMBER() OVER (PARTITION BY client_id, currency ORDER BY period_end DESC) AS rn
               FROM client_debts
               WHERE client_id IS NOT NULL"""
        ):
            if r["rn"] == 1:
                d = debt_map.setdefault(r["client_id"], {"UZS": 0, "USD": 0})
                d[r["currency"]] = r["balance"] or 0
    except Exception:
        pass  # table may not exist on older DB snapshots

    # Last payment from client_payments
    pay_map = {}
    try:
        for r in conn.execute(
            """SELECT client_id, MAX(doc_date) AS last_payment_date
               FROM client_payments WHERE client_id IS NOT NULL GROUP BY client_id"""
        ):
            pay_map[r["client_id"]] = r["last_payment_date"]
    except Exception:
        pass

    for row in data:
        cid = row["id"]
        orders_12mo, last_order = orders_map.get(cid, (0, None))
        row["orders_12mo"] = orders_12mo
        row["last_order_date"] = last_order
        d = debt_map.get(cid, {"UZS": 0, "USD": 0})
        row["debt_uzs"] = d.get("UZS") or 0
        row["debt_usd"] = d.get("USD") or 0
        row["last_payment_date"] = pay_map.get(cid)


def _enrich_location(conn, data: list[dict]) -> None:
    """GPS/address now comes directly from users table via the LEFT JOIN
    in _fetch_rows (users.latitude/longitude/location_address/location_updated).
    This function is a no-op kept for clarity / future extension (e.g. if
    a separate client_locations audit table lands)."""
    for row in data:
        row.setdefault("gps_lat", None)
        row.setdefault("gps_lng", None)
        row.setdefault("gps_address", None)
        row.setdefault("gps_shared_at", None)


def build_xlsx_bytes() -> bytes:
    """Render the full-mirror Client Master xlsx. Returns the bytes blob."""
    conn = get_db()
    try:
        data = _fetch_rows(conn)
        _enrich_financials(conn, data)
        _enrich_location(conn, data)
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Master"

    # Header row
    for col_idx, (header, editable, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _FILL_HEADER
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, d in enumerate(data, start=2):
        for col_idx, (_, editable, source) in enumerate(COLUMNS, start=1):
            if source == "seq":
                value = row_idx - 1
            elif source == "phone_formatted":
                value = _format_phone(d.get("phone_normalized"))
            else:
                value = d.get(source)
            if isinstance(value, bool):
                value = int(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _BORDER
            if editable:
                cell.fill = _FILL_EDITABLE
            if d.get("needs_review"):
                if col_idx in (1, 2):
                    cell.fill = _FILL_REVIEW

    # Column widths — heuristic
    widths = [6, 30, 28, 14, 18, 18, 22, 22, 14, 18, 14, 18, 14,
              16, 10, 12, 14, 28,
              12, 18, 22, 20, 20, 12, 12, 30, 20,
              10, 14, 16, 16, 16,
              10, 12, 14,
              12, 16, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    summary = wb.create_sheet("Info")
    now = datetime.now(TASHKENT).strftime("%Y-%m-%d %H:%M (Toshkent)")
    summary.append(["Client Master — to'liq snapshot", ""])
    summary.append(["Generated", now])
    summary.append(["Total rows", len(data)])
    summary.append(["", ""])
    summary.append(["Rangli ustunlar:", ""])
    summary.append(["  ✏️ sariq", "tahrirlash mumkin (siz o'zgartirishingiz uchun)"])
    summary.append(["  🔒 oq",    "faqat o'qish uchun (har eksportda DB dan yangilanadi)"])
    summary.append(["  🔴 pushti", "needs_review — inson ko'rib chiqishi kerak"])
    summary.append(["", ""])
    summary.append(["Foydalanish:", ""])
    summary.append(["  1. ✏️ ustunlarni tahrirlang", ""])
    summary.append(["  2. Faylni saqlang", ""])
    summary.append(["  3. Telegramda /clientmaster caption bilan yuboring", ""])
    summary.append(["", ""])
    summary.append(["Muhim:", "Upload = DELTA. Bo'sh qoldirilgan qatorlar o'chirilmaydi."])

    for col in ("A", "B"):
        summary.column_dimensions[col].width = 40

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_xlsx_to_archive(archive_dir: str = "/data/master_archive") -> str:
    """Save a fresh export snapshot to the archive directory and return the path.
    Prunes to the last 12 files."""
    os.makedirs(archive_dir, exist_ok=True)
    ts = datetime.now(TASHKENT).strftime("%Y%m%d_%H%M%S")
    path = os.path.join(archive_dir, f"Client_Master_{ts}.xlsx")
    with open(path, "wb") as f:
        f.write(build_xlsx_bytes())
    # Retain last 12 files
    files = sorted([p for p in os.listdir(archive_dir) if p.startswith("Client_Master_")])
    while len(files) > 12:
        old = files.pop(0)
        try:
            os.remove(os.path.join(archive_dir, old))
        except OSError:
            pass
    return path
