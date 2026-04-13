"""Import real orders from 1C "Реализация товаров" (sales / shipments) export.

This is the source of truth for what was actually shipped/sold to clients,
as opposed to the wish-list orders submitted via the Mini App. Both live
side-by-side in the Personal Cabinet so clients can compare what they
requested vs. what they actually received.

File format (XLS, cp1251 or XLSX):
    - Two-row structure within each Реализация document:
        Header row: marked with "V" in column 0, contains document metadata
        Line item rows: follow the header, contain products on the document
    - First few rows of the sheet contain a title and the column header row
    - The very first column ("V" marker) distinguishes header from line item
      rows. Header rows have "V" (or similar single-char marker), item rows
      are blank in column 0.

Header columns (after the V marker):
    Номер, Дата, Время, Автор, Контрагент, Договор, Место хранения,
    Расчетный счет, Аванс покупателя, Зачет аванса покупателя, Счет покупателя,
    ФИО доверенности, Дата оплаты, Ответственное лицо, Торговый агент,
    Примечание, Курс, Валюта

Line item columns:
    № стр., Товар, Количество, Цена, Сумма, НДС, Всего, Счет учета,
    Остаток, Себестоимость, Итого себестоимость, СуммаВал, Цена валютная,
    ВсегоВал, Место хранения, Вес, Сумма веса

Idempotency: import keys on `doc_number_1c` (UNIQUE in the table). Re-uploading
the same month replaces existing records via INSERT OR REPLACE on doc_number,
which keeps the same `doc_number_1c` row but refreshes its data and items.
"""
import io
import re
import logging
from datetime import datetime, time
from typing import Dict, List, Optional, Tuple

from backend.database import get_db
from backend.services.import_balances import _try_match_client

logger = logging.getLogger(__name__)


# Header column synonyms used to find columns by name in the header row.
# Maps internal field name → list of possible 1C column captions (lowercased).
_HEADER_FIELDS = {
    "doc_number_1c":     ["номер", "№"],
    "doc_date":          ["дата"],
    "doc_time":          ["время"],
    "author":            ["автор"],
    "client_name_1c":    ["контрагент"],
    "contract":          ["договор"],
    "storage_location":  ["место хранения"],
    "payment_account":   ["расчетный счет", "счет покупателя"],
    "responsible_person":["ответственное лицо"],
    "sale_agent":        ["торговый агент"],
    "comment":           ["примечание"],
    "exchange_rate":     ["курс"],
    "currency":          ["валюта"],
}

_ITEM_FIELDS = {
    "line_no":          ["№ стр.", "№ стр", "стр.", "№"],
    "product_name_1c":  [
        "товар", "товары", "номенклатура",
        "наименование", "наименование товара",
        "наименование номенклатуры", "товар (услуга)",
    ],
    "quantity":         ["количество", "кол-во", "кол."],
    "price":            ["цена"],
    "sum_local":        ["сумма"],
    "vat":              ["ндс"],
    "total_local":      ["всего"],
    "account":          ["счет учета"],
    "stock_remainder":  ["остаток"],
    "cost":             ["себестоимость"],
    "total_cost":       ["итого себестоимость"],
    "sum_currency":     ["суммавал", "сумма вал"],
    "price_currency":   ["цена валютная", "цена вал"],
    "total_currency":   ["всеговал", "всего вал"],
    "item_storage":     ["место хранения"],
    "weight_per_unit":  ["вес"],
    "total_weight":     ["сумма веса"],
}


def _norm(s) -> str:
    """Normalize a header cell value for matching."""
    if s is None:
        return ""
    return str(s).strip().lower().replace("ё", "е")


def _parse_number(val) -> float:
    """Convert a cell value to float (0 if empty / unparseable)."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _parse_int(val) -> Optional[int]:
    n = _parse_number(val)
    if n == 0 and (val is None or str(val).strip() == ""):
        return None
    return int(n)


def _parse_doc_date(val) -> Optional[str]:
    """Parse a date cell from 1C. Accepts:
    - Excel float dates (e.g. 45292.0 → 2024-01-01)
    - String dates like '15.01.26', '15.01.2026', '15.01.2026 12:34:00'
    - datetime objects
    Returns ISO 'YYYY-MM-DD' or None.
    """
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        # Excel serial date
        try:
            import xlrd
            tup = xlrd.xldate_as_tuple(float(val), 0)
            return f"{tup[0]:04d}-{tup[1]:02d}-{tup[2]:02d}"
        except Exception:
            return None
    s = str(val).strip()
    if not s:
        return None
    # Try DD.MM.YY or DD.MM.YYYY (with optional time after a space)
    m = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return f"{y:04d}-{mo:02d}-{d:02d}"
        except Exception:
            return None
    # Try ISO already
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s[:10]
    return None


def _parse_doc_time(val) -> Optional[str]:
    """Parse a time-of-day cell. Returns 'HH:MM:SS' or None."""
    if val is None or val == "":
        return None
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")
    if isinstance(val, datetime):
        return val.strftime("%H:%M:%S")
    if isinstance(val, (int, float)):
        # Excel time fraction
        try:
            f = float(val)
            f = f - int(f)
            total_sec = int(round(f * 86400))
            h = total_sec // 3600
            m = (total_sec % 3600) // 60
            s = total_sec % 60
            return f"{h:02d}:{m:02d}:{s:02d}"
        except Exception:
            return None
    s = str(val).strip()
    if not s:
        return None
    m = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        sec = int(m.group(3) or 0)
        return f"{h:02d}:{mi:02d}:{sec:02d}"
    return None


# ── Workbook abstraction (xlrd for .xls, openpyxl for .xlsx) ─────────────

class _Sheet:
    """Tiny wrapper to give xlrd/openpyxl sheets a unified API."""
    def __init__(self, rows: List[List]):
        self.rows = rows
        self.nrows = len(rows)
        self.ncols = max((len(r) for r in rows), default=0)

    def cell(self, r: int, c: int):
        if r < 0 or r >= self.nrows:
            return None
        row = self.rows[r]
        if c < 0 or c >= len(row):
            return None
        return row[c]


def _load_workbook(file_bytes: bytes, filename_hint: str = "") -> Tuple[Optional[_Sheet], Optional[str]]:
    """Load a 1C export — try .xls (xlrd) first, fall back to .xlsx (openpyxl).

    Returns (sheet, error_message).
    """
    is_xlsx = filename_hint.lower().endswith(".xlsx")

    if not is_xlsx:
        # Try .xls first
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes, encoding_override="cp1251")
            sh = wb.sheet_by_index(0)
            rows = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    v = sh.cell_value(r, c)
                    ct = sh.cell_type(r, c)
                    # xlrd cell types: 0 empty, 1 text, 2 number, 3 date, 4 bool, 5 error
                    if ct == 3:  # date
                        try:
                            tup = xlrd.xldate_as_tuple(v, wb.datemode)
                            v = datetime(*tup) if tup[0] else time(tup[3], tup[4], tup[5])
                        except Exception:
                            pass
                    row.append(v)
                rows.append(row)
            return _Sheet(rows), None
        except Exception as e:
            xls_err = str(e)
            if not is_xlsx:
                # Don't give up — try openpyxl in case the file is actually xlsx
                pass

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sh = wb[wb.sheetnames[0]]
        rows = []
        for row in sh.iter_rows(values_only=True):
            rows.append(list(row))
        return _Sheet(rows), None
    except Exception as e:
        return None, f"Failed to open file as .xls or .xlsx: {e}"


# ── Header detection ─────────────────────────────────────────────────────

def _find_header_row(sh: _Sheet) -> Optional[int]:
    """Find the row index that contains the column headers.

    Heuristic: scan the first ~25 rows looking for a row that contains the
    word 'Контрагент' (key column for documents). The row may also contain
    'Номер' or 'Дата'. Return the row index, or None if not found.
    """
    max_scan = min(25, sh.nrows)
    for r in range(max_scan):
        norm_cells = [_norm(sh.cell(r, c)) for c in range(sh.ncols)]
        joined = " | ".join(norm_cells)
        if "контрагент" in joined and ("номер" in joined or "дата" in joined):
            return r
    return None


def _build_column_map(sh: _Sheet, header_row: int, fields: Dict[str, List[str]]) -> Dict[str, int]:
    """Map field names → column indices using the header row.

    For each (field, [synonyms]) pair, find the first column whose header
    text exactly matches one of the synonyms (case-insensitive, ё→е normalized).
    """
    col_map: Dict[str, int] = {}
    cells = [_norm(sh.cell(header_row, c)) for c in range(sh.ncols)]
    for field, synonyms in fields.items():
        for c, cell in enumerate(cells):
            if not cell:
                continue
            for syn in synonyms:
                if cell == syn or cell.startswith(syn + " ") or cell == syn.replace(" ", ""):
                    if field not in col_map:
                        col_map[field] = c
                        break
            if field in col_map:
                break
    return col_map


def _dump_first_rows(sh: _Sheet, n: int = 20) -> List[List[str]]:
    """Return first N rows as stringified cell grids (for error diagnostics)."""
    out = []
    for r in range(min(n, sh.nrows)):
        row = []
        for c in range(min(sh.ncols, 20)):
            v = sh.cell(r, c)
            row.append("" if v is None else str(v).strip()[:40])
        out.append(row)
    return out


def _is_header_row(sh: _Sheet, r: int) -> bool:
    """A document header row has a non-empty marker in column 0
    (typically 'V' or similar). Item rows are blank in column 0.
    """
    v = sh.cell(r, 0)
    if v is None:
        return False
    s = str(v).strip()
    return len(s) > 0


# ── Main parser ──────────────────────────────────────────────────────────

def parse_real_orders_xls(file_bytes: bytes, filename_hint: str = "") -> dict:
    """Parse a 1C "Реализация товаров" export into structured documents.

    Returns dict with keys:
        ok (bool)
        error (str) — only on failure
        documents (list of dicts) — each with header fields + items list
        stats (dict) — totals for the file
    """
    sh, err = _load_workbook(file_bytes, filename_hint)
    if err:
        return {"ok": False, "error": err}

    if sh.nrows < 5:
        return {"ok": False, "error": "File too short (< 5 rows)"}

    header_row = _find_header_row(sh)
    if header_row is None:
        return {"ok": False, "error": "Could not find header row (looked for 'Контрагент' in first 25 rows)"}

    header_cols = _build_column_map(sh, header_row, _HEADER_FIELDS)
    item_cols = _build_column_map(sh, header_row, _ITEM_FIELDS)

    # Some 1C exports put the item-level header on a SECOND row shortly
    # after the document-level header. If we didn't find the product column
    # on the primary header row, scan up to 10 rows below for one that
    # contains the item columns.
    item_header_row = header_row
    if "product_name_1c" not in item_cols:
        for rr in range(header_row + 1, min(header_row + 11, sh.nrows)):
            trial = _build_column_map(sh, rr, _ITEM_FIELDS)
            if "product_name_1c" in trial and "quantity" in trial:
                item_cols = trial
                item_header_row = rr
                break

    # The two row types share the same column header row in 1C exports —
    # header columns occupy some columns, item columns occupy others.
    if "client_name_1c" not in header_cols:
        return {
            "ok": False,
            "error": f"'Контрагент' column not found in header row {header_row}",
            "diagnostics": _dump_first_rows(sh, 20),
        }
    if "product_name_1c" not in item_cols:
        return {
            "ok": False,
            "error": (
                f"'Товар' / 'Номенклатура' column not found near header row {header_row}. "
                f"Agar Реализация товаров (actual sales) faylini yubormoqchi edingiz, "
                f"uning 1C'dagi to'g'ri variantini tanlab export qiling. "
                f"Agar оборотка / дебиторка yuborgan bo'lsangiz — /balances yoki /debtors "
                f"buyrug'idan foydalaning."
            ),
            "diagnostics": _dump_first_rows(sh, 20),
        }

    documents: List[dict] = []
    current: Optional[dict] = None

    # Walk data rows — start after whichever header row is lower
    product_col = item_cols.get("product_name_1c", -1)
    client_col = header_cols.get("client_name_1c", -1)
    start_row = max(header_row, item_header_row) + 1
    for r in range(start_row, sh.nrows):
        # Skip rows that are empty across the entire width
        if all(sh.cell(r, c) in (None, "") for c in range(sh.ncols)):
            continue

        if _is_header_row(sh, r):
            # New document
            if current is not None and (current["items"] or current.get("client_name_1c")):
                documents.append(current)

            client_name = str(sh.cell(r, header_cols.get("client_name_1c", -1)) or "").strip()
            doc_number = str(sh.cell(r, header_cols.get("doc_number_1c", -1)) or "").strip()
            if not client_name and not doc_number:
                # Skip noise rows that have a marker but no real content
                current = None
                continue

            doc_date_raw = sh.cell(r, header_cols["doc_date"]) if "doc_date" in header_cols else None
            doc_time_raw = sh.cell(r, header_cols["doc_time"]) if "doc_time" in header_cols else None

            current = {
                "doc_number_1c": doc_number,
                "doc_date": _parse_doc_date(doc_date_raw),
                "doc_time": _parse_doc_time(doc_time_raw),
                "client_name_1c": client_name,
                "contract": str(sh.cell(r, header_cols.get("contract", -1)) or "").strip() or None,
                "storage_location": str(sh.cell(r, header_cols.get("storage_location", -1)) or "").strip() or None,
                "payment_account": str(sh.cell(r, header_cols.get("payment_account", -1)) or "").strip() or None,
                "sale_agent": str(sh.cell(r, header_cols.get("sale_agent", -1)) or "").strip() or None,
                "responsible_person": str(sh.cell(r, header_cols.get("responsible_person", -1)) or "").strip() or None,
                "comment": str(sh.cell(r, header_cols.get("comment", -1)) or "").strip() or None,
                "currency": (str(sh.cell(r, header_cols.get("currency", -1)) or "").strip().upper() or "UZS"),
                "exchange_rate": _parse_number(sh.cell(r, header_cols.get("exchange_rate", -1))),
                "items": [],
            }
            if not current["exchange_rate"]:
                current["exchange_rate"] = 1.0
        else:
            # Line item — must belong to a current document
            if current is None:
                continue
            product_name = str(sh.cell(r, item_cols.get("product_name_1c", -1)) or "").strip()
            if not product_name:
                continue

            item = {
                "line_no": _parse_int(sh.cell(r, item_cols.get("line_no", -1))),
                "product_name_1c": product_name,
                "quantity": _parse_number(sh.cell(r, item_cols.get("quantity", -1))),
                "price": _parse_number(sh.cell(r, item_cols.get("price", -1))),
                "sum_local": _parse_number(sh.cell(r, item_cols.get("sum_local", -1))),
                "vat": _parse_number(sh.cell(r, item_cols.get("vat", -1))),
                "total_local": _parse_number(sh.cell(r, item_cols.get("total_local", -1))),
                "stock_remainder": _parse_number(sh.cell(r, item_cols.get("stock_remainder", -1))),
                "cost": _parse_number(sh.cell(r, item_cols.get("cost", -1))),
                "total_cost": _parse_number(sh.cell(r, item_cols.get("total_cost", -1))),
                "sum_currency": _parse_number(sh.cell(r, item_cols.get("sum_currency", -1))),
                "price_currency": _parse_number(sh.cell(r, item_cols.get("price_currency", -1))),
                "total_currency": _parse_number(sh.cell(r, item_cols.get("total_currency", -1))),
                "storage_location": str(sh.cell(r, item_cols.get("item_storage", -1)) or "").strip() or None,
                "weight_per_unit": _parse_number(sh.cell(r, item_cols.get("weight_per_unit", -1))),
                "total_weight": _parse_number(sh.cell(r, item_cols.get("total_weight", -1))),
            }
            current["items"].append(item)

    # Flush last document
    if current is not None and (current["items"] or current.get("client_name_1c")):
        documents.append(current)

    # ── Post-process: derive missing totals ─────────────────────────────
    # Some 1C "Реализация" exports omit the "Всего" column entirely (no VAT,
    # so total = sum), or omit the foreign-currency block entirely (when the
    # doc is recorded in UZS even though the contract is denominated in USD).
    # Without this fallback, total_local / total_currency / order header
    # totals all stay 0 and the Cabinet renders blank prices.
    #
    # Item-level fallback chain (UZS):
    #   total_local := total_local || (sum_local + vat) || sum_local || (price * quantity)
    #   sum_local   := sum_local || (price * quantity)
    # Item-level fallback chain (currency):
    #   total_currency := total_currency || sum_currency || (price_currency * quantity)
    #   sum_currency   := sum_currency || (price_currency * quantity)
    # Doc-header fallback:
    #   total_sum          := total_sum || SUM(items.total_local)
    #   total_sum_currency := total_sum_currency || SUM(items.total_currency)
    for d in documents:
        for it in d["items"]:
            if not it.get("sum_local"):
                if it.get("price") and it.get("quantity"):
                    it["sum_local"] = it["price"] * it["quantity"]
            if not it.get("total_local"):
                if it.get("sum_local"):
                    it["total_local"] = it["sum_local"] + (it.get("vat") or 0)
                elif it.get("price") and it.get("quantity"):
                    it["total_local"] = it["price"] * it["quantity"]
            if not it.get("sum_currency"):
                if it.get("price_currency") and it.get("quantity"):
                    it["sum_currency"] = it["price_currency"] * it["quantity"]
            if not it.get("total_currency"):
                if it.get("sum_currency"):
                    it["total_currency"] = it["sum_currency"]
                elif it.get("price_currency") and it.get("quantity"):
                    it["total_currency"] = it["price_currency"] * it["quantity"]
        # Doc-header totals
        if not d.get("total_sum"):
            d["total_sum"] = sum(i.get("total_local") or 0 for i in d["items"])
        if not d.get("total_sum_currency"):
            d["total_sum_currency"] = sum(i.get("total_currency") or 0 for i in d["items"])

    # Stats
    total_items = sum(len(d["items"]) for d in documents)
    total_sum_local = sum(sum(i["total_local"] for i in d["items"]) for d in documents)
    total_sum_currency = sum(sum(i["total_currency"] for i in d["items"]) for d in documents)
    unique_clients = {d["client_name_1c"] for d in documents if d["client_name_1c"]}
    unique_products = {i["product_name_1c"] for d in documents for i in d["items"]}
    dates = sorted({d["doc_date"] for d in documents if d.get("doc_date")})

    return {
        "ok": True,
        "documents": documents,
        "stats": {
            "doc_count": len(documents),
            "item_count": total_items,
            "client_count": len(unique_clients),
            "product_count": len(unique_products),
            "date_min": dates[0] if dates else None,
            "date_max": dates[-1] if dates else None,
            "date_count": len(dates),
            "total_local": total_sum_local,
            "total_currency": total_sum_currency,
        },
    }


# ── Product matching ─────────────────────────────────────────────────────

_PRODUCT_CACHE: Dict[str, Optional[int]] = {}


def _try_match_product(product_name_1c: str, conn) -> Optional[int]:
    """Match a 1C product name to products.id. Cached for the import.

    Strategy:
    1. Exact match on products.name (the original 1C name stored in DB)
    2. Normalized exact match (lowercase, stripped)
    """
    if product_name_1c in _PRODUCT_CACHE:
        return _PRODUCT_CACHE[product_name_1c]

    row = conn.execute(
        "SELECT id FROM products WHERE name = ? LIMIT 1",
        (product_name_1c,),
    ).fetchone()
    if row:
        _PRODUCT_CACHE[product_name_1c] = row[0]
        return row[0]

    normalized = product_name_1c.strip().lower()
    row = conn.execute(
        "SELECT id FROM products WHERE LOWER(TRIM(name)) = ? LIMIT 1",
        (normalized,),
    ).fetchone()
    pid = row[0] if row else None
    _PRODUCT_CACHE[product_name_1c] = pid
    return pid


# ── Apply (write to DB) ──────────────────────────────────────────────────

def apply_real_orders_import(file_bytes: bytes, filename_hint: str = "") -> dict:
    """Parse a Реализация file and upsert all documents + items.

    Idempotent on `doc_number_1c`: re-uploading the same file (or overlapping
    period) refreshes existing documents instead of duplicating them.
    """
    parsed = parse_real_orders_xls(file_bytes, filename_hint)
    if not parsed.get("ok"):
        return parsed

    documents = parsed["documents"]
    stats = parsed["stats"]

    if not documents:
        return {"ok": False, "error": "No documents found in file"}

    conn = get_db()

    # Reset product cache for this import (DB may have changed since last call)
    _PRODUCT_CACHE.clear()

    # Build a name → client_id cache for the unique clients in this file
    client_cache: Dict[str, Optional[int]] = {}

    inserted_docs = 0
    updated_docs = 0
    inserted_items = 0
    matched_clients = 0
    matched_products = 0
    unmatched_clients: List[str] = []
    unmatched_products: List[str] = []

    for d in documents:
        if not d["doc_number_1c"]:
            # Skip documents with no number (cannot dedupe)
            continue

        client_name = d["client_name_1c"]
        if client_name not in client_cache:
            client_cache[client_name] = _try_match_client(client_name, conn)
        client_id = client_cache[client_name]

        if client_id is not None:
            matched_clients += 1
        else:
            unmatched_clients.append(client_name)

        # Compute totals from items
        total_local = sum(i["total_local"] for i in d["items"])
        total_currency = sum(i["total_currency"] for i in d["items"])
        total_weight = sum(i["total_weight"] for i in d["items"])
        item_count = len(d["items"])

        # Upsert by doc_number_1c
        existing = conn.execute(
            "SELECT id FROM real_orders WHERE doc_number_1c = ?",
            (d["doc_number_1c"],),
        ).fetchone()

        if existing:
            real_order_id = existing[0]
            conn.execute(
                """UPDATE real_orders SET
                    doc_date=?, doc_time=?, client_name_1c=?, client_id=?,
                    contract=?, storage_location=?, payment_account=?,
                    sale_agent=?, responsible_person=?, comment=?,
                    currency=?, exchange_rate=?,
                    total_sum=?, total_sum_currency=?, total_weight=?, item_count=?,
                    imported_at=datetime('now')
                   WHERE id=?""",
                (
                    d["doc_date"], d["doc_time"], client_name, client_id,
                    d.get("contract"), d.get("storage_location"), d.get("payment_account"),
                    d.get("sale_agent"), d.get("responsible_person"), d.get("comment"),
                    d.get("currency"), d.get("exchange_rate") or 1,
                    total_local, total_currency, total_weight, item_count,
                    real_order_id,
                ),
            )
            # Wipe and re-insert items (simpler than line-level diff)
            conn.execute("DELETE FROM real_order_items WHERE real_order_id = ?", (real_order_id,))
            updated_docs += 1
        else:
            cur = conn.execute(
                """INSERT INTO real_orders
                   (doc_number_1c, doc_date, doc_time, client_name_1c, client_id,
                    contract, storage_location, payment_account,
                    sale_agent, responsible_person, comment,
                    currency, exchange_rate,
                    total_sum, total_sum_currency, total_weight, item_count)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    d["doc_number_1c"], d["doc_date"], d["doc_time"], client_name, client_id,
                    d.get("contract"), d.get("storage_location"), d.get("payment_account"),
                    d.get("sale_agent"), d.get("responsible_person"), d.get("comment"),
                    d.get("currency"), d.get("exchange_rate") or 1,
                    total_local, total_currency, total_weight, item_count,
                ),
            )
            real_order_id = cur.lastrowid
            inserted_docs += 1

        for it in d["items"]:
            product_id = _try_match_product(it["product_name_1c"], conn)
            if product_id is not None:
                matched_products += 1
            else:
                unmatched_products.append(it["product_name_1c"])

            conn.execute(
                """INSERT INTO real_order_items
                   (real_order_id, line_no, product_name_1c, product_id,
                    quantity, price, sum_local, vat, total_local,
                    price_currency, sum_currency, total_currency,
                    cost, total_cost, stock_remainder,
                    storage_location, weight_per_unit, total_weight)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    real_order_id, it.get("line_no"), it["product_name_1c"], product_id,
                    it["quantity"], it["price"], it["sum_local"], it["vat"], it["total_local"],
                    it["price_currency"], it["sum_currency"], it["total_currency"],
                    it["cost"], it["total_cost"], it["stock_remainder"],
                    it.get("storage_location"), it["weight_per_unit"], it["total_weight"],
                ),
            )
            inserted_items += 1

    conn.commit()

    db_total_docs = conn.execute("SELECT COUNT(*) FROM real_orders").fetchone()[0]
    db_total_items = conn.execute("SELECT COUNT(*) FROM real_order_items").fetchone()[0]
    conn.close()

    # Deduplicate samples
    unique_unmatched_clients = sorted(set(unmatched_clients))
    unique_unmatched_products = sorted(set(unmatched_products))

    return {
        "ok": True,
        "stats": stats,
        "inserted_docs": inserted_docs,
        "updated_docs": updated_docs,
        "inserted_items": inserted_items,
        "matched_clients": matched_clients,
        "matched_products": matched_products,
        "unmatched_clients_count": len(unique_unmatched_clients),
        "unmatched_products_count": len(unique_unmatched_products),
        "unmatched_clients_sample": unique_unmatched_clients[:15],
        "unmatched_products_sample": unique_unmatched_products[:15],
        "db_total_docs": db_total_docs,
        "db_total_items": db_total_items,
    }


# ── Read API for Cabinet ─────────────────────────────────────────────────

def list_real_orders_for_client(client_id, limit: int = 50) -> List[dict]:
    """List real orders for a client (newest first).

    client_id can be a single int or a list of ints (sibling IDs for
    multi-phone clients sharing the same client_id_1c).
    """
    conn = get_db()
    if isinstance(client_id, (list, tuple)):
        ids = client_id
    else:
        ids = [client_id]
    placeholders = ",".join("?" * len(ids))
    rows = conn.execute(
        f"""SELECT id, doc_number_1c, doc_date, doc_time, client_name_1c,
                  currency, exchange_rate, total_sum, total_sum_currency,
                  total_weight, item_count, sale_agent, comment, imported_at
           FROM real_orders
           WHERE client_id IN ({placeholders})
           ORDER BY doc_date DESC, doc_time DESC, id DESC
           LIMIT ?""",
        (*ids, limit),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_real_order_detail(real_order_id: int) -> Optional[dict]:
    """Fetch a single real order with all its line items."""
    conn = get_db()
    order = conn.execute(
        """SELECT id, doc_number_1c, doc_date, doc_time, client_name_1c, client_id,
                  contract, storage_location, payment_account, sale_agent,
                  responsible_person, comment, currency, exchange_rate,
                  total_sum, total_sum_currency, total_weight, item_count, imported_at
           FROM real_orders WHERE id = ?""",
        (real_order_id,),
    ).fetchone()
    if not order:
        conn.close()
        return None

    items = conn.execute(
        """SELECT roi.line_no, roi.product_name_1c, roi.product_id,
                  roi.quantity, roi.price, roi.sum_local, roi.total_local,
                  roi.price_currency, roi.total_currency,
                  roi.weight_per_unit, roi.total_weight,
                  p.name_display, pr.name AS producer_name, p.unit
           FROM real_order_items roi
           LEFT JOIN products p ON p.id = roi.product_id
           LEFT JOIN producers pr ON pr.id = p.producer_id
           WHERE roi.real_order_id = ?
           ORDER BY roi.line_no, roi.id""",
        (real_order_id,),
    ).fetchall()
    conn.close()

    return {
        "order": dict(order),
        "items": [dict(i) for i in items],
    }


def find_nearby_wishlist(client_id: int, doc_date: str, days: int = 5) -> List[dict]:
    """Find wish-list orders (from `orders` table) near a real order date.

    Used by the Cabinet 'compare' feature: when a client expands a real order,
    show wish-list orders from the same client within ±N days so they can see
    what they originally requested vs. what was actually shipped.
    """
    conn = get_db()
    user = conn.execute(
        "SELECT telegram_id FROM users WHERE client_id = ? LIMIT 1",
        (client_id,),
    ).fetchone()
    if not user:
        conn.close()
        return []

    rows = conn.execute(
        """SELECT id, total_usd, total_uzs, item_count, status, created_at
           FROM orders
           WHERE telegram_id = ?
             AND date(created_at) BETWEEN date(?, ?) AND date(?, ?)
           ORDER BY created_at DESC""",
        (user["telegram_id"], doc_date, f"-{days} days", doc_date, f"+{days} days"),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def find_nearby_real_orders(telegram_id: int, wishlist_created_at: str, days: int = 5) -> List[dict]:
    """Find real orders near a wish-list order date for the same client.

    Mirror of find_nearby_wishlist — used when the client expands a wish-list
    order in the Cabinet. Returns the real shipments from ±N days around it.
    """
    conn = get_db()
    user = conn.execute(
        "SELECT client_id FROM users WHERE telegram_id = ? LIMIT 1",
        (telegram_id,),
    ).fetchone()
    if not user or not user["client_id"]:
        conn.close()
        return []

    rows = conn.execute(
        """SELECT id, doc_number_1c, doc_date, doc_time,
                  currency, total_sum, total_sum_currency,
                  item_count, sale_agent
           FROM real_orders
           WHERE client_id = ?
             AND date(doc_date) BETWEEN date(?, ?) AND date(?, ?)
           ORDER BY doc_date DESC, doc_time DESC""",
        (
            user["client_id"],
            wishlist_created_at, f"-{days} days",
            wishlist_created_at, f"+{days} days",
        ),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ── Admin helpers for unmatched real-order clients ────────────────────────
#
# Added 2026-04-07 after verifying Jan/Feb/Mar 2026 ingestion showed ~27% of
# real-order documents with client_id = NULL. Root cause: `_try_match_client`
# (reused from import_balances.py) compares Python-lowercased cyrillic to
# SQLite `LOWER(TRIM(name))` — but SQLite's built-in LOWER is ASCII-only, so
# the name-fallback branch never succeeds for cyrillic. It only links clients
# that already have `allowed_clients.client_id_1c` populated.
#
# These helpers give ops two tools without touching the existing matcher:
#   1. `list_unmatched_real_clients()` — report the damage, ranked by doc count
#   2. `relink_real_orders()` — do a Python-side cyrillic-aware re-match pass
#      against allowed_clients and fill in any client_id it can resolve
#
# System names (1C correction docs, not real clients) are skipped in BOTH
# the listing and the relink pass so they don't clutter the report or cause
# spurious false matches.

def _py_normalize_client_name(name: Optional[str]) -> str:
    """Python-side normalization for cyrillic-aware client-name comparison.

    Unlike SQLite's LOWER() (ASCII-only), Python `str.lower()` folds cyrillic
    case correctly. We also fold ё→е and collapse whitespace, matching the
    spirit of `_norm()` used elsewhere in this file.
    """
    if name is None:
        return ""
    s = str(name).strip().lower().replace("ё", "е")
    # Collapse internal whitespace runs
    s = re.sub(r"\s+", " ", s)
    return s


# 1C placeholder / aggregate docs that should NOT be treated as real clients.
# These are booked on non-client buckets (cash registers, generic category
# holders, legal-entity aggregate) or are adjustment markers. Names are stored
# py-normalized so `_is_system_non_client` can do a single set lookup.
#
# The first two are 1C correction/adjustment markers. The middle block
# (added 2026-04-07 after user confirmed) are walk-in cash/aggregate
# buckets — together they were 77% of the post-relink unmatched residue
# on the Q1 2026 dataset. "В О З В Р А Т ПОСТАВЩИКУ" was added 2026-04-07
# after the full 2025 historical ingest — it is the 1C return-to-supplier
# doc marker (letter-spaced in the export, analogous to ИСПРАВЛЕНИЕ) and
# appeared in 10 of 12 months of 2025 unmatched residue.
SYSTEM_NON_CLIENT_NAMES = frozenset(
    _py_normalize_client_name(s) for s in [
        "ИСПРАВЛЕНИЕ",
        "ИСПРАВЛЕНИЕ СКЛАД 2",
        "Наличка №1",
        "Наличка №2",
        "Наличка №3",
        "Наличка СКЛАД",
        "Наличка - Магазин",
        "Организации (переч.)",
        "СТРОЙКА",
        "СТЕКЛОПЛАСТИК",
        "В О З В Р А Т ПОСТАВЩИКУ",
    ]
)


def _is_system_non_client(name: Optional[str]) -> bool:
    """True if the name is a known 1C correction/adjustment marker, not a real client."""
    if not name:
        return False
    return _py_normalize_client_name(name) in SYSTEM_NON_CLIENT_NAMES


def list_unmatched_real_clients(limit: int = 200) -> dict:
    """Report real_orders rows with client_id IS NULL, grouped by client_name_1c.

    Returns per-name stats (doc_count, total_local, first_seen, last_seen) sorted
    by doc_count DESC so operators can prioritize the biggest offenders. System
    non-client markers (e.g. ИСПРАВЛЕНИЕ) are excluded from the report.
    """
    conn = get_db()
    rows = conn.execute(
        """SELECT client_name_1c,
                  COUNT(*) AS doc_count,
                  SUM(COALESCE(total_sum, 0)) AS total_local,
                  SUM(COALESCE(total_sum_currency, 0)) AS total_currency,
                  MIN(doc_date) AS first_seen,
                  MAX(doc_date) AS last_seen
           FROM real_orders
           WHERE client_id IS NULL
           GROUP BY client_name_1c
           ORDER BY doc_count DESC, total_local DESC"""
    ).fetchall()

    total_docs_unmatched = 0
    total_local_unmatched = 0.0
    skipped_system = 0
    items: List[dict] = []

    for r in rows:
        name = r["client_name_1c"] or ""
        if _is_system_non_client(name):
            skipped_system += int(r["doc_count"] or 0)
            continue
        doc_count = int(r["doc_count"] or 0)
        total_local = float(r["total_local"] or 0)
        total_docs_unmatched += doc_count
        total_local_unmatched += total_local
        if len(items) < limit:
            items.append({
                "client_name_1c": name,
                "doc_count": doc_count,
                "total_local": round(total_local),
                "total_currency": round(float(r["total_currency"] or 0), 2),
                "first_seen": r["first_seen"],
                "last_seen": r["last_seen"],
            })

    # Overall DB totals for context
    db_total_docs = conn.execute("SELECT COUNT(*) FROM real_orders").fetchone()[0]
    db_matched_docs = conn.execute(
        "SELECT COUNT(*) FROM real_orders WHERE client_id IS NOT NULL"
    ).fetchone()[0]
    conn.close()

    return {
        "ok": True,
        "db_total_docs": db_total_docs,
        "db_matched_docs": db_matched_docs,
        "db_unmatched_docs": db_total_docs - db_matched_docs,
        "unique_unmatched_names": len(items),  # after skip list
        "total_unmatched_docs_after_skip": total_docs_unmatched,
        "total_unmatched_local_after_skip": round(total_local_unmatched),
        "skipped_system_docs": skipped_system,
        "items": items,
    }


def list_unmatched_real_products(limit: int = 100) -> dict:
    """Report real_order_items rows with product_id IS NULL, grouped by product_name_1c.

    Returns per-name stats (line_count, total_quantity, doc_count, client_count,
    total_local, total_currency, first_seen, last_seen) sorted by line_count DESC
    so ops/Session A can prioritize SKUs that hurt the most. Unlike the clients
    report, there is no system skip-list for products — unmatched products are
    always genuine catalog gaps.
    """
    conn = get_db()
    rows = conn.execute(
        """SELECT ri.product_name_1c                           AS product_name_1c,
                  COUNT(*)                                     AS line_count,
                  SUM(COALESCE(ri.quantity, 0))                AS total_quantity,
                  COUNT(DISTINCT ri.real_order_id)             AS doc_count,
                  COUNT(DISTINCT ro.client_id)                 AS client_count,
                  SUM(COALESCE(ri.total_local, 0))             AS total_local,
                  SUM(COALESCE(ri.total_currency, 0))          AS total_currency,
                  MIN(ro.doc_date)                             AS first_seen,
                  MAX(ro.doc_date)                             AS last_seen
           FROM real_order_items ri
           JOIN real_orders ro ON ro.id = ri.real_order_id
           WHERE ri.product_id IS NULL
           GROUP BY ri.product_name_1c
           ORDER BY line_count DESC, total_local DESC"""
    ).fetchall()

    total_lines_unmatched = 0
    total_local_unmatched = 0.0
    total_currency_unmatched = 0.0
    items: List[dict] = []

    for r in rows:
        name = (r["product_name_1c"] or "").strip()
        if not name:
            continue
        line_count = int(r["line_count"] or 0)
        total_local = float(r["total_local"] or 0)
        total_currency = float(r["total_currency"] or 0)
        total_lines_unmatched += line_count
        total_local_unmatched += total_local
        total_currency_unmatched += total_currency
        if len(items) < limit:
            items.append({
                "product_name_1c": name,
                "line_count": line_count,
                "total_quantity": round(float(r["total_quantity"] or 0), 3),
                "doc_count": int(r["doc_count"] or 0),
                "client_count": int(r["client_count"] or 0),
                "total_local": round(total_local),
                "total_currency": round(total_currency, 2),
                "first_seen": r["first_seen"],
                "last_seen": r["last_seen"],
            })

    # Overall DB totals for context
    db_total_items = conn.execute("SELECT COUNT(*) FROM real_order_items").fetchone()[0]
    db_matched_items = conn.execute(
        "SELECT COUNT(*) FROM real_order_items WHERE product_id IS NOT NULL"
    ).fetchone()[0]
    conn.close()

    return {
        "ok": True,
        "db_total_items": db_total_items,
        "db_matched_items": db_matched_items,
        "db_unmatched_items": db_total_items - db_matched_items,
        "unique_unmatched_names": len(items),  # capped by limit
        "total_unique_unmatched_names_full": len(rows),  # full count before limit
        "total_unmatched_lines": total_lines_unmatched,
        "total_unmatched_local": round(total_local_unmatched),
        "total_unmatched_currency": round(total_currency_unmatched, 2),
        "items": items,
    }


def relink_real_orders() -> dict:
    """Re-run client matching for every real_orders row where client_id IS NULL.

    Uses a cyrillic-aware Python-side normalization (unlike the SQLite LOWER
    used in `_try_match_client`), which can resolve most name-only matches
    that the original import missed. This does NOT touch the existing matcher
    used by fresh imports — it's purely a remediation sweep.

    Matching strategy per unmatched row, in order:
      1. If `client_id_1c` in allowed_clients equals the raw 1C name (exact) → match
      2. Python-normalized (lower, ё→е, whitespace-collapsed) equality against
         allowed_clients.name
      3. Else leave unmatched.

    Returns a summary of the sweep. Safe to run multiple times; rows already
    matched are never touched.
    """
    conn = get_db()
    # Load every allowed_client once, then build two in-memory indexes:
    #   id_1c_index: raw client_id_1c → allowed_clients.id
    #   name_index:  py-normalized name → allowed_clients.id (first wins on collision)
    allowed = conn.execute(
        "SELECT id, name, client_id_1c FROM allowed_clients WHERE COALESCE(status, 'active') != 'merged' ORDER BY id"
    ).fetchall()

    id_1c_index: Dict[str, int] = {}
    name_index: Dict[str, int] = {}
    for a in allowed:
        if a["client_id_1c"]:
            id_1c_index.setdefault(str(a["client_id_1c"]), a["id"])
        if a["name"]:
            norm = _py_normalize_client_name(a["name"])
            if norm and norm not in name_index:
                name_index[norm] = a["id"]

    # Get all distinct unmatched (name, list of row ids) pairs
    unmatched_rows = conn.execute(
        """SELECT id, client_name_1c
           FROM real_orders
           WHERE client_id IS NULL"""
    ).fetchall()

    relinked_by_id_1c = 0
    relinked_by_name = 0
    still_unmatched = 0
    skipped_system = 0
    # Per-name decision cache (big speedup — many rows per name)
    decision_cache: Dict[str, Optional[int]] = {}

    for row in unmatched_rows:
        raw_name = row["client_name_1c"] or ""
        if _is_system_non_client(raw_name):
            skipped_system += 1
            continue

        if raw_name in decision_cache:
            resolved = decision_cache[raw_name]
        else:
            resolved = None
            # Strategy 1: raw name matches some client_id_1c literally
            if raw_name in id_1c_index:
                resolved = id_1c_index[raw_name]
                decision_cache[raw_name] = resolved
            else:
                norm = _py_normalize_client_name(raw_name)
                if norm and norm in name_index:
                    resolved = name_index[norm]
                decision_cache[raw_name] = resolved

        if resolved is None:
            still_unmatched += 1
            continue

        conn.execute(
            "UPDATE real_orders SET client_id = ? WHERE id = ?",
            (resolved, row["id"]),
        )
        # Bookkeeping: we don't know which strategy won in this branch after
        # the cache hit, so approximate by rechecking once. Cheap enough.
        if raw_name in id_1c_index and id_1c_index[raw_name] == resolved:
            relinked_by_id_1c += 1
        else:
            relinked_by_name += 1

    conn.commit()

    # Post-sweep totals
    db_total_docs = conn.execute("SELECT COUNT(*) FROM real_orders").fetchone()[0]
    db_matched_docs = conn.execute(
        "SELECT COUNT(*) FROM real_orders WHERE client_id IS NOT NULL"
    ).fetchone()[0]

    # "Real client" denominator: exclude docs booked on placeholder / cash /
    # aggregate buckets (SYSTEM_NON_CLIENT_NAMES). This is the denominator ops
    # actually cares about — unmatched placeholder docs are not a data quality
    # problem, just how 1C records walk-in cash and bulk transfers. Computed
    # in Python because SQLite LOWER() is ASCII-only and can't match cyrillic
    # placeholder names like "Наличка №3".
    all_rows = conn.execute(
        "SELECT client_id, client_name_1c FROM real_orders"
    ).fetchall()
    db_system_docs = sum(1 for r in all_rows if _is_system_non_client(r["client_name_1c"]))
    conn.close()

    db_real_client_docs = db_total_docs - db_system_docs
    # Matched docs that are NOT system names (a matched system doc would be
    # weird but not impossible if someone aliased it in allowed_clients).
    db_real_client_matched = sum(
        1 for r in all_rows
        if r["client_id"] is not None and not _is_system_non_client(r["client_name_1c"])
    )
    real_match_pct = (
        (db_real_client_matched / db_real_client_docs * 100.0)
        if db_real_client_docs else 0.0
    )

    relinked_total = relinked_by_id_1c + relinked_by_name
    return {
        "ok": True,
        "scanned": len(unmatched_rows),
        "relinked_total": relinked_total,
        "relinked_by_client_id_1c": relinked_by_id_1c,
        "relinked_by_name": relinked_by_name,
        "still_unmatched": still_unmatched,
        "skipped_system": skipped_system,
        "db_total_docs": db_total_docs,
        "db_matched_docs": db_matched_docs,
        "db_unmatched_docs": db_total_docs - db_matched_docs,
        # "Real client" view — the one that matters for data quality reporting
        "db_system_docs": db_system_docs,
        "db_real_client_docs": db_real_client_docs,
        "db_real_client_matched": db_real_client_matched,
        "db_real_client_unmatched": db_real_client_docs - db_real_client_matched,
        "db_real_client_match_pct": round(real_match_pct, 1),
    }


def backfill_real_order_totals() -> dict:
    """One-shot backfill: derive missing total_local / sum_local / total_currency
    / sum_currency on existing real_order_items rows, and missing total_sum /
    total_sum_currency on existing real_orders rows.

    Mirrors the import-time post-processing in `_parse_workbook` so existing
    DB data heals without requiring re-upload of all 1C exports. Idempotent —
    only touches rows where the target column is currently 0 / NULL.

    Returns counts of rows touched per phase.
    """
    conn = get_db()
    counts = {}

    # ── Item-level backfills ─────────────────────────────────────────────
    # Phase 1a: sum_local := price * quantity (where sum_local is 0 but price+qty present)
    cur = conn.execute(
        """UPDATE real_order_items
           SET sum_local = price * quantity
           WHERE COALESCE(sum_local, 0) = 0
             AND COALESCE(price, 0) > 0
             AND COALESCE(quantity, 0) > 0"""
    )
    counts["item_sum_local_from_price_qty"] = cur.rowcount

    # Phase 1b: total_local := sum_local + vat (where total_local is 0)
    cur = conn.execute(
        """UPDATE real_order_items
           SET total_local = COALESCE(sum_local, 0) + COALESCE(vat, 0)
           WHERE COALESCE(total_local, 0) = 0
             AND COALESCE(sum_local, 0) > 0"""
    )
    counts["item_total_local_from_sum"] = cur.rowcount

    # Phase 1c: total_local := price * quantity (rescue path if sum_local missing too)
    cur = conn.execute(
        """UPDATE real_order_items
           SET total_local = price * quantity
           WHERE COALESCE(total_local, 0) = 0
             AND COALESCE(price, 0) > 0
             AND COALESCE(quantity, 0) > 0"""
    )
    counts["item_total_local_from_price_qty"] = cur.rowcount

    # Phase 1d: sum_currency := price_currency * quantity
    cur = conn.execute(
        """UPDATE real_order_items
           SET sum_currency = price_currency * quantity
           WHERE COALESCE(sum_currency, 0) = 0
             AND COALESCE(price_currency, 0) > 0
             AND COALESCE(quantity, 0) > 0"""
    )
    counts["item_sum_currency_from_price_qty"] = cur.rowcount

    # Phase 1e: total_currency := sum_currency
    cur = conn.execute(
        """UPDATE real_order_items
           SET total_currency = sum_currency
           WHERE COALESCE(total_currency, 0) = 0
             AND COALESCE(sum_currency, 0) > 0"""
    )
    counts["item_total_currency_from_sum"] = cur.rowcount

    # Phase 1f: total_currency := price_currency * quantity (rescue)
    cur = conn.execute(
        """UPDATE real_order_items
           SET total_currency = price_currency * quantity
           WHERE COALESCE(total_currency, 0) = 0
             AND COALESCE(price_currency, 0) > 0
             AND COALESCE(quantity, 0) > 0"""
    )
    counts["item_total_currency_from_price_qty"] = cur.rowcount

    # ── Order-header backfills ───────────────────────────────────────────
    # Phase 2a: total_sum := SUM(items.total_local)
    cur = conn.execute(
        """UPDATE real_orders
           SET total_sum = (
               SELECT COALESCE(SUM(total_local), 0)
               FROM real_order_items
               WHERE real_order_id = real_orders.id
           )
           WHERE COALESCE(total_sum, 0) = 0"""
    )
    counts["order_total_sum_from_items"] = cur.rowcount

    # Phase 2b: total_sum_currency := SUM(items.total_currency)
    cur = conn.execute(
        """UPDATE real_orders
           SET total_sum_currency = (
               SELECT COALESCE(SUM(total_currency), 0)
               FROM real_order_items
               WHERE real_order_id = real_orders.id
           )
           WHERE COALESCE(total_sum_currency, 0) = 0"""
    )
    counts["order_total_sum_currency_from_items"] = cur.rowcount

    conn.commit()

    # Sanity totals after backfill
    db_total_orders = conn.execute("SELECT COUNT(*) FROM real_orders").fetchone()[0]
    db_orders_with_total = conn.execute(
        "SELECT COUNT(*) FROM real_orders WHERE COALESCE(total_sum, 0) > 0"
    ).fetchone()[0]
    db_total_items = conn.execute("SELECT COUNT(*) FROM real_order_items").fetchone()[0]
    db_items_with_total = conn.execute(
        "SELECT COUNT(*) FROM real_order_items WHERE COALESCE(total_local, 0) > 0"
    ).fetchone()[0]
    conn.close()

    return {
        "ok": True,
        "phases": counts,
        "rows_touched_total": sum(counts.values()),
        "db_total_orders": db_total_orders,
        "db_orders_with_total": db_orders_with_total,
        "db_total_items": db_total_items,
        "db_items_with_total": db_items_with_total,
    }


def get_real_order_sample_for_client(client_substring: str) -> dict:
    """Diagnostic: dump the most recent real order for any client whose name
    matches the given substring (cyrillic-aware, case-insensitive). Returns
    raw DB price columns so ops can spot whether a "no price" Cabinet display
    is a parser bug (zeros in DB) or a render bug (data present, UI hiding it).

    Searches both `allowed_clients.name` and `real_orders.client_name_1c`.
    """
    if not client_substring or not client_substring.strip():
        return {"ok": False, "error": "client substring required"}

    needle = _py_normalize_client_name(client_substring)
    conn = get_db()

    # 1) Try matching via allowed_clients first (typical case: linked client)
    candidate_client_ids: List[int] = []
    for row in conn.execute(
        "SELECT id, name, phone_normalized FROM allowed_clients WHERE name IS NOT NULL"
    ).fetchall():
        if needle in _py_normalize_client_name(row["name"]):
            candidate_client_ids.append(row["id"])

    matched_client_rows: List[dict] = []
    if candidate_client_ids:
        placeholders = ",".join("?" * len(candidate_client_ids))
        for ac in conn.execute(
            f"SELECT id, name, phone_normalized FROM allowed_clients WHERE id IN ({placeholders})",
            candidate_client_ids,
        ).fetchall():
            matched_client_rows.append(dict(ac))

    # 2) Find candidate real_orders: matches by client_id OR by raw cyrillic name substring
    real_orders_rows: List[dict] = []
    if candidate_client_ids:
        placeholders = ",".join("?" * len(candidate_client_ids))
        for r in conn.execute(
            f"""SELECT id, doc_number_1c, doc_date, client_name_1c, client_id,
                       currency, exchange_rate, total_sum, total_sum_currency, item_count
                FROM real_orders
                WHERE client_id IN ({placeholders})
                ORDER BY doc_date DESC, id DESC
                LIMIT 5""",
            candidate_client_ids,
        ).fetchall():
            real_orders_rows.append(dict(r))

    # Always also try a name substring scan (fallback if not linked)
    if not real_orders_rows:
        for r in conn.execute(
            """SELECT id, doc_number_1c, doc_date, client_name_1c, client_id,
                      currency, exchange_rate, total_sum, total_sum_currency, item_count
               FROM real_orders
               ORDER BY doc_date DESC, id DESC
               LIMIT 500"""
        ).fetchall():
            if needle in _py_normalize_client_name(r["client_name_1c"]):
                real_orders_rows.append(dict(r))
                if len(real_orders_rows) >= 5:
                    break

    if not real_orders_rows:
        conn.close()
        return {
            "ok": True,
            "needle": client_substring,
            "matched_allowed_clients": matched_client_rows,
            "real_orders": [],
            "items_dump": [],
            "note": "no real_orders found for this substring",
        }

    # 3) Dump line items for the MOST RECENT order — that's the diagnostic payload
    sample = real_orders_rows[0]
    items = conn.execute(
        """SELECT line_no, product_name_1c, product_id,
                  quantity, price, sum_local, vat, total_local,
                  price_currency, sum_currency, total_currency
           FROM real_order_items
           WHERE real_order_id = ?
           ORDER BY line_no, id
           LIMIT 30""",
        (sample["id"],),
    ).fetchall()
    conn.close()

    items_dump = [dict(i) for i in items]

    # Sanity counts: how many items have any non-zero value in each price column?
    items_with_price = sum(1 for i in items_dump if (i["price"] or 0) > 0)
    items_with_sum = sum(1 for i in items_dump if (i["sum_local"] or 0) > 0)
    items_with_total = sum(1 for i in items_dump if (i["total_local"] or 0) > 0)
    items_with_vat = sum(1 for i in items_dump if (i["vat"] or 0) > 0)
    items_with_price_currency = sum(1 for i in items_dump if (i["price_currency"] or 0) > 0)
    items_with_sum_currency = sum(1 for i in items_dump if (i["sum_currency"] or 0) > 0)
    items_with_total_currency = sum(1 for i in items_dump if (i["total_currency"] or 0) > 0)

    return {
        "ok": True,
        "needle": client_substring,
        "matched_allowed_clients": matched_client_rows,
        "real_orders": real_orders_rows,
        "sample_doc": sample,
        "items_dump": items_dump,
        "items_count": len(items_dump),
        "items_with_price": items_with_price,
        "items_with_sum_local": items_with_sum,
        "items_with_total_local": items_with_total,
        "items_with_vat": items_with_vat,
        "items_with_price_currency": items_with_price_currency,
        "items_with_sum_currency": items_with_sum_currency,
        "items_with_total_currency": items_with_total_currency,
    }


# ── Ingest unmatched SKUs into products table + relink ─────────────

# Brand family → (category_id, producer_id) mapping.
# Determined by inspecting existing products for each brand in prod DB.
_BRAND_FAMILY_MAP = {
    # OSCAR products: Лак, Жидкое стекло, Олиф
    "БТ Лак OSCAR":    (10, 21),   # Laklar & Oliflar, Oscar
    "Жидкое стекло OSCAR": (10, 21),
    "Олиф OSCAR":      (10, 21),
    # ДЕЛЮКС water emulsions
    "ДЕЛЮКС в/э":      (17, 4),    # Suv Emulsiya & Gruntovka, De Luxe
    "ДЕЛЮКС ЧЕРНЫЙ":   (17, 4),
    # ДекоАРТ Эмаль variants
    "ДекоАРТ Эмаль":   (1, 5),     # Bo'yoq & Emal, Dekoart
    "Декор Лак":        (10, 5),    # Laklar & Oliflar, Dekoart
    "Декор DEKOCENTO":  (14, 5),    # Qorishma & Suvoq, Dekoart (decorative coating)
    # ДЕКОАРТ Универсал Эмульсия
    "ДЕКОАРТ Универсал": (17, 5),   # Suv Emulsiya & Gruntovka, Dekoart
    "ДЕКОАРТ СТРОНГ ФАСАД": (17, 5),
    "ДЕКОАРТ KF":       (17, 5),
    # ДЕКАСТАР
    "ДЕКАСТАР":         (17, 5),    # Suv Emulsiya & Gruntovka, Dekoart
    # Скоч Травертин
    "Скоч Травертин":   (15, 21),   # Qurilish Mollari, Oscar
    # АНТИМОРОЗ
    "АНТИМОРОЗ":        (3, 12),    # Boshqa Mahsulot, Gogle
    # Электрод MONOLIT
    "Электрод  MONOLIT": (7, 36),   # Elektrodlar, Xitoy
    # Кафель
    "Кафель":           (3, 36),    # Boshqa Mahsulot, Xitoy
    # ЛЕСКА-ЖИЛКА
    "ЛЕСКА-ЖИЛКА":      (3, 36),    # Boshqa Mahsulot, Xitoy
    # СИЛКОАТ
    "СИЛКОАТ":          (14, 30),   # Qorishma & Suvoq, Silkcoat
}


def _classify_product(name_1c: str) -> Tuple[int, int]:
    """Return (category_id, producer_id) for an unmatched 1C product name.

    Tries longest-prefix match against _BRAND_FAMILY_MAP keys.
    Falls back to category 20 (Yangi mahsulotlar) + producer 21 (Oscar)
    if no match found — the fallback is safe because these will be
    manually reviewed anyway.
    """
    # Sort by key length descending for longest-prefix-first matching
    for prefix in sorted(_BRAND_FAMILY_MAP.keys(), key=len, reverse=True):
        if name_1c.upper().startswith(prefix.upper()):
            return _BRAND_FAMILY_MAP[prefix]
    # Fallback: "Yangi mahsulotlar" (new products) category
    return (20, 21)


def ingest_unmatched_skus() -> dict:
    """Find all product names in real_order_items with product_id IS NULL,
    add them to the products table, and relink the items.

    For each unmatched product_name_1c:
    1. Check it doesn't already exist in products.name (case-insensitive)
    2. Classify into category/producer based on brand family patterns
    3. Generate a display name using the import_products pipeline
    4. INSERT into products
    5. UPDATE real_order_items SET product_id = new_id WHERE product_name_1c matches

    Returns summary stats.
    """
    from backend.services.import_products import generate_display_name
    from backend.database import build_search_text, rebuild_all_search_text

    conn = get_db()

    # Step 1: Get all distinct unmatched product names
    unmatched = conn.execute(
        """SELECT DISTINCT product_name_1c
           FROM real_order_items
           WHERE product_id IS NULL
             AND product_name_1c IS NOT NULL
             AND TRIM(product_name_1c) != ''"""
    ).fetchall()

    added = 0
    already_existed = 0
    relinked_items = 0
    skipped = 0
    details = []

    for row in unmatched:
        name_1c = row["product_name_1c"].strip()
        if not name_1c:
            skipped += 1
            continue

        # Check if product already exists (exact or case-insensitive)
        existing = conn.execute(
            "SELECT id FROM products WHERE LOWER(TRIM(name)) = LOWER(TRIM(?)) LIMIT 1",
            (name_1c,)
        ).fetchone()

        if existing:
            # Product exists but items weren't linked — relink them
            pid = existing[0]
            already_existed += 1
            updated = conn.execute(
                """UPDATE real_order_items
                   SET product_id = ?
                   WHERE product_name_1c = ? AND product_id IS NULL""",
                (pid, row["product_name_1c"])
            ).rowcount
            relinked_items += updated
            details.append({
                "name_1c": name_1c,
                "action": "relinked_existing",
                "product_id": pid,
                "items_relinked": updated,
            })
            continue

        # Step 2: Classify
        cat_id, prod_id = _classify_product(name_1c)

        # Step 3: Generate display name
        # Get the producer's Cyrillic name for stripping
        prod_row = conn.execute("SELECT name FROM producers WHERE id = ?", (prod_id,)).fetchone()
        producer_cyrillic = prod_row["name"] if prod_row else ""
        display_name = generate_display_name(name_1c, producer_cyrillic)

        # Step 4: Parse weight from name
        from backend.services.parse_weight import parse_weight_from_name
        weight = parse_weight_from_name(name_1c)

        # Build search text
        cat_row = conn.execute("SELECT name FROM categories WHERE id = ?", (cat_id,)).fetchone()
        cat_name = cat_row["name"] if cat_row else ""
        search_text = build_search_text(
            name_1c, display_name, producer_cyrillic, None, cat_name
        )

        # INSERT
        cursor = conn.execute(
            """INSERT INTO products
               (name, name_display, category_id, producer_id, unit,
                price_usd, price_uzs, weight, is_active, search_text)
               VALUES (?, ?, ?, ?, 'sht', 0, 0, ?, 1, ?)""",
            (name_1c, display_name, cat_id, prod_id, weight, search_text)
        )
        new_pid = cursor.lastrowid
        added += 1

        # Step 5: Relink items
        updated = conn.execute(
            """UPDATE real_order_items
               SET product_id = ?
               WHERE product_name_1c = ? AND product_id IS NULL""",
            (new_pid, row["product_name_1c"])
        ).rowcount
        relinked_items += updated

        details.append({
            "name_1c": name_1c,
            "action": "added",
            "product_id": new_pid,
            "display_name": display_name,
            "category_id": cat_id,
            "producer_id": prod_id,
            "items_relinked": updated,
        })

    # Update denormalized counts
    conn.execute("""
        UPDATE categories SET product_count = (
            SELECT COUNT(*) FROM products WHERE products.category_id = categories.id AND is_active = 1
        )
    """)
    conn.execute("""
        UPDATE producers SET product_count = (
            SELECT COUNT(*) FROM products WHERE products.producer_id = producers.id AND is_active = 1
        )
    """)

    conn.commit()

    # Get new match stats
    db_total = conn.execute("SELECT COUNT(*) FROM real_order_items").fetchone()[0]
    db_matched = conn.execute(
        "SELECT COUNT(*) FROM real_order_items WHERE product_id IS NOT NULL"
    ).fetchone()[0]
    db_unmatched = db_total - db_matched

    conn.close()

    return {
        "ok": True,
        "products_added": added,
        "products_already_existed": already_existed,
        "items_relinked": relinked_items,
        "skipped": skipped,
        "new_match_rate": f"{db_matched}/{db_total} ({100*db_matched/db_total:.1f}%)" if db_total else "N/A",
        "remaining_unmatched": db_unmatched,
        "details": details,
    }
