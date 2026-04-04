"""Update product display names, weights, units, and categories from Rassvet_Master.xlsx.

Reads the 'Katalog' sheet and syncs name_display, weight, unit, and category_id
in the products table from the master spreadsheet.

IMPORTANT: Excel IDs are 4881-7320 but DB IDs are 1-2440 (AUTOINCREMENT).
The offset is 4880: DB_ID = EXCEL_ID - 4880.

This runs AFTER import_products to override auto-transliterated names
with manually corrected Uzbek display names.
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

from openpyxl import load_workbook
from backend.database import get_db, init_db, rebuild_all_search_text

ID_OFFSET = 4880  # Excel ID 4881 = DB ID 1

MASTER_CANDIDATES = [
    os.path.join(os.path.dirname(__file__), '..', '..', 'Rassvet_Master.xlsx'),
    '/data/Rassvet_Master.xlsx',
]


def find_master():
    for p in MASTER_CANDIDATES:
        if os.path.exists(p):
            return p
    return None


def update_display_names():
    path = find_master()
    if not path:
        print("update_display_names: Rassvet_Master.xlsx not found — skipping.")
        return

    init_db()
    conn = get_db()

    existing = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    if existing == 0:
        print("update_display_names: No products in DB — skipping.")
        conn.close()
        return

    # Build category id → name map from DB
    db_cat_id_to_name = {}
    for row in conn.execute("SELECT id, name FROM categories").fetchall():
        db_cat_id_to_name[row[0]] = row[1]
    db_cat_name_to_id = {v: k for k, v in db_cat_id_to_name.items()}

    print(f"update_display_names: Loading {path} (DB has {existing} products)...")
    wb = load_workbook(path, read_only=True, data_only=True)
    if 'Katalog' not in wb.sheetnames:
        print("update_display_names: 'Katalog' sheet not found — skipping.")
        wb.close()
        conn.close()
        return

    ws = wb['Katalog']

    # ── Phase 1: Sync category names from master ────────────────────
    # Match master→DB categories by product count (all counts are unique).
    # This avoids fragile ID-offset matching that breaks when rows are
    # added or deleted from the master.
    master_cat_counts = {}  # master category name → product count
    for row in ws.iter_rows(min_row=2, values_only=True):
        category = row[1]
        if category is None:
            continue
        cat_name = str(category).strip()
        master_cat_counts[cat_name] = master_cat_counts.get(cat_name, 0) + 1

    # DB category counts
    db_cat_counts = {}  # db category name → (cat_id, count)
    for row in conn.execute(
        "SELECT c.id, c.name, COUNT(p.id) FROM categories c "
        "LEFT JOIN products p ON p.category_id = c.id "
        "GROUP BY c.id"
    ).fetchall():
        db_cat_counts[row[1]] = (row[0], row[2])

    # Find categories in master that don't exist in DB (new names)
    new_names = {n: c for n, c in master_cat_counts.items() if n not in db_cat_counts}
    # Find categories in DB that don't exist in master (old names)
    old_names = {n: db_cat_counts[n] for n in db_cat_counts if n not in master_cat_counts}

    # Match by product count
    rename_plan = {}  # db_category_id → new_name
    for new_name, new_count in new_names.items():
        for old_name, (old_cat_id, old_count) in old_names.items():
            if old_count == new_count and old_cat_id not in rename_plan:
                rename_plan[old_cat_id] = new_name
                break

    # Execute all renames
    renamed_cats = 0
    for old_cat_id, new_name in rename_plan.items():
        old_name = db_cat_id_to_name.get(old_cat_id, '???')
        conn.execute(
            "UPDATE categories SET name = ? WHERE id = ?",
            (new_name, old_cat_id)
        )
        db_cat_id_to_name[old_cat_id] = new_name
        if old_name in db_cat_name_to_id:
            del db_cat_name_to_id[old_name]
        db_cat_name_to_id[new_name] = old_cat_id
        renamed_cats += 1
        print(f"  Category renamed: '{old_name}' → '{new_name}'")

    # ── Phase 2: Sync product display names, weights, units, categories ────
    updated_names = 0
    updated_weights = 0
    updated_units = 0
    updated_cats = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        excel_id = row[0]       # Column A = Excel ID (4881-7320)
        category = row[1]       # Column B = Category name
        name = row[4]           # Column E = Ilovadagi nomi (Latin)
        unit = row[7]           # Column H = Birlik (unit type)
        weight = row[10]        # Column K = Weight
        if excel_id is None or name is None:
            skipped += 1
            continue
        try:
            excel_id = int(excel_id)
        except (ValueError, TypeError):
            skipped += 1
            continue

        db_id = excel_id - ID_OFFSET
        if db_id < 1:
            skipped += 1
            continue

        name = str(name).strip()
        if not name:
            skipped += 1
            continue

        # Update name_display
        result = conn.execute(
            "UPDATE products SET name_display = ? WHERE id = ?",
            (name, db_id)
        )
        if result.rowcount > 0:
            updated_names += 1

        # Update weight if Excel has a value; fallback to parsing from DB name
        if weight is not None:
            try:
                w = float(weight)
                conn.execute(
                    "UPDATE products SET weight = ? WHERE id = ?",
                    (w, db_id)
                )
                updated_weights += 1
            except (ValueError, TypeError):
                pass
        else:
            # No weight in Excel — try parsing from original Cyrillic name
            from backend.services.parse_weight import parse_weight_from_name
            row_db = conn.execute("SELECT name, weight FROM products WHERE id = ?", (db_id,)).fetchone()
            if row_db and (row_db["weight"] is None or row_db["weight"] == 0):
                parsed_w = parse_weight_from_name(row_db["name"] or "")
                if parsed_w is not None:
                    conn.execute("UPDATE products SET weight = ? WHERE id = ?", (parsed_w, db_id))
                    updated_weights += 1

        # Update unit if Excel has a value
        if unit is not None:
            unit_str = str(unit).strip()
            if unit_str:
                conn.execute(
                    "UPDATE products SET unit = ? WHERE id = ?",
                    (unit_str, db_id)
                )
                updated_units += 1

        # Update category if it differs
        if category and str(category).strip() in db_cat_name_to_id:
            new_cat_id = db_cat_name_to_id[str(category).strip()]
            conn.execute(
                "UPDATE products SET category_id = ? WHERE id = ?",
                (new_cat_id, db_id)
            )
            updated_cats += 1

    conn.commit()

    # Rebuild search_text index to reflect updated names, units, categories
    print("update_display_names: Rebuilding search_text index...")
    count = rebuild_all_search_text(conn)
    print(f"update_display_names: Rebuilt search_text for {count} products.")

    conn.close()
    wb.close()
    print(f"update_display_names: Renamed {renamed_cats} categories, "
          f"updated {updated_names} names, {updated_weights} weights, "
          f"{updated_units} units, {updated_cats} category assignments "
          f"({skipped} skipped).")


if __name__ == '__main__':
    update_display_names()
