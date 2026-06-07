"""
Import allowed clients from the CSV file into the allowed_clients table.
Normalizes phone numbers for matching against Telegram contacts.
Supports client_id_1c and company_name columns for 1C integration.
"""
import io
import sqlite3
import os
import re

DATABASE_PATH = os.getenv("DATABASE_PATH", "/data/catalog.db")
CLIENTS_FILE = os.path.join(os.path.dirname(__file__), "..", "..", "clients_data.csv")


# 1C's "Телефоны контрагента" cell frequently packs 2-3 phones into one string,
# sometimes with Cyrillic relationship markers (эри / ўғли / дадаси / укаси) or
# contact names between them. Examples from clients 28.05.26.xls:
#   "93 356 12 12, эри 97 918 33 33"                       (comma + relationship)
#   "+998 90 605 79 36, 90 194 34 74 - Гульноза"           (comma + contact name)
#   "90 199 51 51 - дадаси, 97 892 87 77 - укаси"          (annotation per phone)
#   "+998 97 396 26 00, 77031 1116+  88 392 01 01"         (3 phones, mixed sep)
#   "93 333 8070- 97 931 04 04"                            (dash+space, no comma)
#   "97 927 17 77 (99 548 47 67)"                          (second phone in parens)
#   "+998 99 455 00 57 994550057 Шарофиддин"               (same number glued twice)
#   "99 165 73 80 952657380"                               (two phones, 1-space sep)
#   "+99897-776-22-26 . 91 532 33 23"                      (period as separator)
# The pre-2026-05-29 normalize_phone() concatenated all digits and took the last 9,
# silently dropping the primary and keeping the LAST phone listed (e.g., Гулноза's
# husband, not Гулноза). See Error Log MULTI_PHONE_CELL_TRUNCATION.
#
# Parse strategy (two stages):
#  1. Split the cell on multi-phone separators (comma/semicolon/pipe, 2+ spaces,
#     dash-followed-by-digits). The split's third alternative — ``-\s+(?=\d)`` —
#     is what distinguishes "93 333 8070- 97 931 04 04" (separator) from
#     "90-194-34-74" (intra-phone) and "90 194 34 74 - Гульноза" (dash before
#     annotation, not another phone).
#  2. Within each piece, find phone-shaped digit runs and walk their digits
#     9-at-a-time (or 12 if the run starts with the 998 country code). Walking
#     handles the parens/glued/period cases uniformly without needing more
#     separators — every Uzbek mobile is exactly 9 local digits.
_PHONE_CELL_SPLIT_RE = re.compile(r"[,;|]|\s{2,}|-\s+(?=\d)")
_PHONE_RUN_RE = re.compile(r"\+?\(?\d[\d\s\-\.\(\)]{7,}\d")

# Known Uzbek mobile operator prefixes (the first two local digits). A valid
# 9-digit mobile always starts with one of these. Centralised here so the
# parser, the consistency audit, and the phone-repair backfill all agree on
# what "valid" means — see .claude/rules/12-dual-source-columns.md (one helper,
# no inline column/format checks scattered across readers).
# Derived empirically from the prod allowed_clients corpus (2026-06): these are
# the prefixes with real operator support; the scattered 1–2-count tail (54, 35,
# 66, …) is corruption/landline noise, not operators. "50" is Ucell — omitting
# it would make the parser drop valid 50… numbers as stray digits.
UZ_MOBILE_OPERATOR_CODES = frozenset({
    "20", "33", "50", "77", "88", "90", "91", "93", "94", "95", "97", "98", "99",
})


def is_valid_uz_mobile(phone) -> bool:
    """True when `phone` is a 9-digit string starting with a known operator code."""
    return (isinstance(phone, str) and len(phone) == 9
            and phone[:2] in UZ_MOBILE_OPERATOR_CODES)


def _walk_digits_into_phones(digits):
    """Slice a digit string into 9-digit phones, anchored on operator prefixes.

    A 1C phone cell's concatenated digits are NOT guaranteed to be a clean
    multiple of 9: stray leading digits (an ``8`` trunk prefix, a typo, a
    glued partial number) shift every fixed-width window so each slice is a
    rotation of garbage with an invalid operator prefix — the 2026-06
    MULTI_PHONE_CELL_MISALIGNMENT corruption (Error Log) that overwrote ~88
    primaries with malformed numbers like ``549009591``.

    The fix: only emit a 9-digit slice when it *starts on a valid operator
    code*. When the current position doesn't, advance one digit and retry —
    realigning past the stray digit instead of emitting corruption. The 998
    country code is stripped only when a valid operator code follows it.
    """
    out = []
    pos = 0
    n = len(digits)
    while pos + 9 <= n:
        if (n - pos >= 12 and digits[pos:pos + 3] == "998"
                and digits[pos + 3:pos + 5] in UZ_MOBILE_OPERATOR_CODES):
            out.append(digits[pos + 3:pos + 12])
            pos += 12
        elif digits[pos:pos + 2] in UZ_MOBILE_OPERATOR_CODES:
            out.append(digits[pos:pos + 9])
            pos += 9
        else:
            # Misaligned — skip the stray digit and re-anchor on the next
            # operator-code boundary rather than slicing invalid digits.
            pos += 1
    return out


def parse_phone_cell(raw):
    """Extract ordered phones with optional annotations from a 1C phone cell.

    Returns: list of dicts ``[{"digits": "9-digit-str", "annotation": "text"}, ...]``.
    Primary phone first. Empty list if no phone-shaped run (>=9 digits) found.
    The annotation is the non-digit text accompanying the phone within its piece
    (relationship markers like "эри" or names like "Гульноза"); empty when the
    piece contained nothing but the phone itself.
    """
    if not raw or not isinstance(raw, str):
        return []
    out = []
    seen = set()
    for piece in _PHONE_CELL_SPLIT_RE.split(raw):
        if not piece:
            continue
        for m in _PHONE_RUN_RE.finditer(piece):
            digits = re.sub(r"\D", "", m.group(0))
            if len(digits) < 9:
                continue
            anno = re.sub(r"[\d\+\-\(\)\.,;:]", " ", piece)
            anno = re.sub(r"\s+", " ", anno).strip()
            for d9 in _walk_digits_into_phones(digits):
                if d9 in seen:
                    continue
                seen.add(d9)
                out.append({"digits": d9, "annotation": anno})
    return out


def normalize_phone(raw: str) -> str:
    """Backward-compat: return just the PRIMARY phone's 9-digit form.

    Callers that need all phones from a multi-phone 1C cell should use
    ``parse_phone_cell()`` instead.
    """
    if not raw or not isinstance(raw, str):
        return ""
    cells = parse_phone_cell(raw)
    if cells:
        return cells[0]["digits"]
    # Last-resort fallback for non-cell single-phone strings the regex missed.
    digits = re.sub(r"\D", "", raw)
    return digits[-9:] if len(digits) >= 9 else digits


def _curated_state(conn, ac_id, gps_lat, credit_score, credit_limit):
    """Return the list of curated-state markers present on an allowed_clients
    row (pin / credit / linked user). Non-empty → the row is a #74 drift target
    and its client_id_1c must not be silently rewritten on a phone-match upsert.
    """
    state = []
    if gps_lat is not None:
        state.append("gps")
    if credit_score is not None:
        state.append("credit_score")
    if credit_limit is not None:
        state.append("credit_limit")
    if conn.execute(
        "SELECT 1 FROM users WHERE client_id = ? LIMIT 1", (ac_id,)
    ).fetchone():
        state.append("linked_user")
    return state


def _sync_client_phones_safe(conn, client_id):
    """Mirror the row's phone slots into client_phones (Phase 1). Best-effort:
    client_phones is not yet a production read source, so a sync hiccup must
    never break the load-bearing import — log and continue."""
    try:
        from backend.services.phone_slots import sync_client_phones
        sync_client_phones(conn, client_id, source="import")
    except Exception as e:  # pragma: no cover — defensive on the import hot path
        print(f"[import_clients] client_phones sync skipped for id={client_id}: {e}")


def _upsert_client_from_row(conn, raw_phone_str, client_name, location, source,
                            cid_1c, company, changed_by_tag, onec_card_id=None):
    """Shared upsert used by both the bot path (apply_clients_upload) and the
    CSV CLI path (import_clients). Returns ``("inserted"|"updated"|"skipped"|"drift_held", existing_id_or_new)``.

    ``drift_held`` (Error Log #74): a phone/raqam match would have rewritten
    ``client_id_1c`` on a curated-state row → held in
    ``client_identity_drift_queue`` instead, row left untouched + needs_review=1.

    Handles 1C multi-phone cells: primary goes to ``phone_normalized``, extras
    fill ``raqam_02/03`` (fill-only — never overwrites a non-null existing slot).
    Relationship markers (e.g. "эри") land in ``ism_02/03`` (also fill-only).

    Lookup precedence (Client Identity Anchoring Phase 0, 2026-06-03):
      0. ``onec_card_id`` — the STABLE 1C card anchor ("{folder}:{Код}"). When
         present, resolves first and definitively: a client whose phone changed
         or got corrupted is still recognised, so no duplicate row is spawned
         (ends the #74/#75/#81 family). A card-id match is "same client" by
         definition → it bypasses the #74 drift guard and the cid-collision
         freeze (a differing name is a legitimate 1C rename, not drift).
      1. ``phone_normalized = primary``
      2. ``client_id_1c = cid_1c`` — catches cases where the importer's primary
         interpretation changed across runs (e.g. parser bugfix moved the
         primary), preventing a duplicate row.
      3. ``raqam_02 = primary OR raqam_03 = primary`` — catches cases where a
         number we previously stored as secondary is now the 1C primary.
    """
    phones = parse_phone_cell(raw_phone_str)
    if not phones:
        return ("skipped", None)
    primary = phones[0]["digits"]
    extras = phones[1:]

    onec_card_id = (onec_card_id or "").strip() or None

    # client_id_1c sanity: never accept a purely-numeric value (1C "Код" leakage).
    if cid_1c and cid_1c.isdigit():
        cid_1c = ""
    if not cid_1c and client_name and not client_name.isdigit():
        cid_1c = client_name

    select_cols = ("id, phone_normalized, raqam_02, raqam_03, ism_02, ism_03, "
                   "client_id_1c, gps_latitude, credit_score, credit_limit, "
                   "onec_card_id")
    not_merged = "COALESCE(status, 'active') NOT LIKE 'merged%'"

    # Track which key matched the existing row — the #74 drift guard only fires
    # on a phone/raqam match (a cid_1c or onec_card_id match can't drift the
    # name by definition).
    match_via = None
    existing = None

    # Step 0 — resolve by the stable card anchor first.
    if onec_card_id:
        existing = conn.execute(
            f"SELECT {select_cols} FROM allowed_clients "
            f"WHERE onec_card_id = ? AND {not_merged} "
            f"ORDER BY id LIMIT 1",
            (onec_card_id,),
        ).fetchone()
        if existing is not None:
            match_via = "onec_card_id"

    if existing is None:
        existing = conn.execute(
            f"SELECT {select_cols} FROM allowed_clients "
            f"WHERE phone_normalized = ? AND {not_merged} "
            f"ORDER BY id LIMIT 1",
            (primary,),
        ).fetchone()
        if existing is not None:
            match_via = "phone"

    if existing is None and cid_1c:
        existing = conn.execute(
            f"SELECT {select_cols} FROM allowed_clients "
            f"WHERE client_id_1c = ? AND {not_merged} "
            f"ORDER BY id LIMIT 1",
            (cid_1c,),
        ).fetchone()
        if existing is not None:
            match_via = "cid_1c"

    if existing is None:
        existing = conn.execute(
            f"SELECT {select_cols} FROM allowed_clients "
            f"WHERE (raqam_02 = ? OR raqam_03 = ?) AND {not_merged} "
            f"ORDER BY id LIMIT 1",
            (primary, primary),
        ).fetchone()
        if existing is not None:
            match_via = "raqam"

    if existing is not None:
        (existing_id, existing_phone, existing_r02, existing_r03,
         existing_i02, existing_i03, existing_cid,
         existing_gps, existing_cs, existing_cl, existing_card) = existing

        # ── #74 IDENTITY-DRIFT GUARD ──────────────────────────────────────
        # A phone/raqam match whose incoming client_id_1c differs from the
        # existing row's, on a row carrying CURATED state (pin / credit /
        # linked user), is identity drift — 1C reassigned a phone between
        # exports and the old upsert would silently rewrite client_id_1c,
        # hijacking that row's curated state (САРДОР Пищевой → Мурод ака
        # Вокзал, Error Log #74). Hold it in client_identity_drift_queue for
        # manual resolution instead of mutating the curated row. The queue
        # row is the audit-first record (zero-data-loss rule).
        if (match_via in ("phone", "raqam") and existing_cid and cid_1c
                and existing_cid != cid_1c):
            curated = _curated_state(conn, existing_id,
                                     existing_gps, existing_cs, existing_cl)
            if curated:
                conn.execute(
                    """INSERT INTO client_identity_drift_queue
                         (allowed_client_id, phone_normalized,
                          existing_client_id_1c, incoming_client_id_1c,
                          incoming_name, curated_state, matched_via)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (existing_id, primary, existing_cid, cid_1c,
                     client_name or "", ",".join(curated), match_via),
                )
                conn.execute(
                    "UPDATE allowed_clients SET needs_review = 1 WHERE id = ?",
                    (existing_id,),
                )
                return ("drift_held", existing_id)
        # ──────────────────────────────────────────────────────────────────
        updates, params = [], []
        flag_needs_review = False

        # Detect cross-client phone collision: phone-match brought up a row
        # whose client_id_1c is a different non-empty label than the incoming.
        # When this happens, the two clients legitimately share a phone digit
        # pattern (or the historical data got tangled). Without protection,
        # the second row's name/location/company would clobber the first's
        # identity even though it's a different real client. The cid_1c
        # tiebreaker decides the label; everything else freezes.
        #
        # EXCEPTION: a card-id match is definitively the SAME client (the card
        # is the stable anchor), so a differing name is a legitimate 1C rename,
        # not a collision — don't freeze; let the name update through.
        cid_collision = bool(
            existing_cid and cid_1c and existing_cid != cid_1c
            and match_via != "onec_card_id"
        )

        if existing_phone != primary and not cid_collision:
            # Cross-client phone-collision guard. The onec_card_id match (step 0)
            # bypasses the phone lookup (step 1), so `primary` may already belong
            # to a DIFFERENT active row. Overwriting it would violate
            # idx_allowed_phone_unique and — because the upload loop had no
            # per-row guard — abort the ENTIRE import (Error Log #85). 1C moving a
            # phone between two active clients is a genuine identity contention,
            # not churn: mirror import_client_master_v2's policy — freeze the
            # phone, flag BOTH rows for review, never auto-move/merge.
            phone_clash = conn.execute(
                "SELECT id FROM allowed_clients "
                "WHERE phone_normalized = ? AND id != ? "
                f"AND {not_merged}",
                (primary, existing_id),
            ).fetchone()
            if phone_clash:
                conn.execute(
                    "INSERT INTO phone_history (client_id, old_phone, new_phone, reason, changed_by) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (existing_id, existing_phone, primary,
                     f"collision_with_{phone_clash[0]}", changed_by_tag),
                )
                conn.execute(
                    "UPDATE allowed_clients SET needs_review = 1 WHERE id IN (?, ?)",
                    (existing_id, phone_clash[0]),
                )
                flag_needs_review = True
            else:
                # Primary phone changed — log to phone_history before overwriting.
                conn.execute(
                    "INSERT INTO phone_history (client_id, old_phone, new_phone, reason, changed_by) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (existing_id, existing_phone, primary,
                     "clients_upload multi-phone-cell parse", changed_by_tag),
                )
                updates.append("phone_normalized = ?"); params.append(primary)

        # Identity fields (name, location, company) — only overwrite when this
        # is unambiguously the same client. Cross-client collision freezes them.
        if not cid_collision:
            if client_name:
                updates.append("name = ?"); params.append(client_name)
            if location:
                updates.append("location = ?"); params.append(location)
            if company:
                updates.append("company_name = ?"); params.append(company)
        else:
            flag_needs_review = True

        if source:
            updates.append("source_sheet = ?"); params.append(source)
        if cid_1c:
            resolved_cid, tb_flag = _resolve_cid_1c_tiebreaker(conn, existing_id, cid_1c)
            if tb_flag:
                flag_needs_review = True
            # Name-as-attribute (Phase 4): a real rename (existing non-empty name
            # → different new name) is logged to client_name_history — the audit
            # trail + old→new map for relinking historical finance rows still
            # keyed by the old client_name_1c. First-time name set (empty → X) is
            # not a rename, so it's not logged.
            if existing_cid and resolved_cid and existing_cid != resolved_cid:
                conn.execute(
                    "INSERT INTO client_name_history "
                    "(client_id, old_name, new_name, reason, changed_by) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (existing_id, existing_cid, resolved_cid,
                     f"1C rename via {match_via}", changed_by_tag),
                )
            updates.append("client_id_1c = ?"); params.append(resolved_cid)

        # Anchor capture (Phase 0) — fill-only. When this row has no card id yet,
        # stamp it (the natural backfill on every import). Never churn an existing
        # anchor: if a phone/cid/raqam match landed on a row already anchored to a
        # DIFFERENT card, the incoming card wasn't found by the step-0 card lookup
        # → genuine anchor conflict (two 1C cards entangled on one row). Leave the
        # stored anchor intact and flag for review rather than silently reassign.
        if onec_card_id:
            if not existing_card:
                updates.append("onec_card_id = ?"); params.append(onec_card_id)
            elif existing_card != onec_card_id and match_via != "onec_card_id":
                flag_needs_review = True

        # Fill-only writes for raqam_02/03 + ism_02/03 from the extra phones.
        slots = [(existing_r02, existing_i02, "raqam_02", "ism_02"),
                 (existing_r03, existing_i03, "raqam_03", "ism_03")]
        for (cur_phone, cur_name, slot_phone, slot_name), extra in zip(slots, extras[:2]):
            if not cur_phone:
                updates.append(f"{slot_phone} = ?"); params.append(extra["digits"])
            if extra["annotation"] and not cur_name:
                updates.append(f"{slot_name} = ?"); params.append(extra["annotation"])

        if flag_needs_review:
            updates.append("needs_review = 1")
        if updates:
            params.append(existing_id)
            conn.execute(
                f"UPDATE allowed_clients SET {', '.join(updates)} WHERE id = ?",
                params,
            )
            _sync_client_phones_safe(conn, existing_id)
            return ("updated", existing_id)
        return ("skipped", existing_id)

    # INSERT path — fresh client.
    r02 = extras[0]["digits"] if len(extras) >= 1 else None
    i02 = (extras[0]["annotation"] or None) if len(extras) >= 1 else None
    r03 = extras[1]["digits"] if len(extras) >= 2 else None
    i03 = (extras[1]["annotation"] or None) if len(extras) >= 2 else None
    cur = conn.execute(
        "INSERT INTO allowed_clients "
        "(phone_normalized, name, location, source_sheet, status, "
        " client_id_1c, company_name, raqam_02, ism_02, raqam_03, ism_03, "
        " onec_card_id) "
        "VALUES (?, ?, ?, ?, 'active', ?, ?, ?, ?, ?, ?, ?)",
        (primary, client_name, location, source or "clients_upload",
         cid_1c, company, r02, i02, r03, i03, onec_card_id),
    )
    _sync_client_phones_safe(conn, cur.lastrowid)
    return ("inserted", cur.lastrowid)


def _resolve_cid_1c_tiebreaker(conn, existing_id: int, new_cid_1c: str):
    """Activity-aware tiebreaker for client_id_1c overwrites.

    When a client_id_1c update would change an existing allowed_clients row,
    check `real_orders` for last-180-day activity under each name. Prefer
    whichever has activity. If both have activity, keep the more recent and
    flag `needs_review` so an operator can adjudicate.

    Used by both the bot path (apply_clients_upload) and the CSV CLI
    (import_clients) so a stale shorthand in clients_data.csv can't blindly
    overwrite a canonical name that's actively in use by 1C imports.

    Returns: (resolved_cid_1c, flag_needs_review)
    """
    existing_cid_row = conn.execute(
        "SELECT client_id_1c FROM allowed_clients WHERE id = ?",
        (existing_id,),
    ).fetchone()
    prev_cid = (existing_cid_row[0] if existing_cid_row else None) or ""

    if not prev_cid or prev_cid == new_cid_1c:
        return new_cid_1c, False

    try:
        prev_recent = conn.execute(
            "SELECT MAX(doc_date) FROM real_orders "
            "WHERE client_name_1c = ? "
            "AND doc_date >= date('now','-180 days')",
            (prev_cid,),
        ).fetchone()[0]
        new_recent = conn.execute(
            "SELECT MAX(doc_date) FROM real_orders "
            "WHERE client_name_1c = ? "
            "AND doc_date >= date('now','-180 days')",
            (new_cid_1c,),
        ).fetchone()[0]
        if prev_recent and new_recent:
            # Both have activity within 180d — flag ambiguity, keep more recent.
            if prev_recent >= new_recent:
                return prev_cid, True
            return new_cid_1c, True
        elif prev_recent and not new_recent:
            # Existing wins — has activity, incoming doesn't.
            return prev_cid, False
        # else: no prev activity → new overwrites normally
    except Exception:
        pass  # activity check is best-effort

    return new_cid_1c, False


def import_clients():
    if not os.path.exists(CLIENTS_FILE):
        print("[import_clients] No clients_data.csv found, skipping.")
        return

    conn = sqlite3.connect(DATABASE_PATH)
    conn.execute("PRAGMA journal_mode=WAL")

    # Read CSV
    import csv
    rows_inserted = 0
    rows_updated = 0
    rows_drift_held = 0
    rows_errored = 0
    seen_phones = set()

    with open(CLIENTS_FILE, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            raw_phone = row.get("phone", "")
            primary = normalize_phone(raw_phone)
            if not primary or primary in seen_phones:
                continue
            seen_phones.add(primary)

            # Per-row resilience: one bad row must never abort the startup import
            # (this runs in the railway boot chain). Error Log #85.
            try:
                outcome, _ = _upsert_client_from_row(
                    conn,
                    raw_phone_str=raw_phone,
                    client_name=row.get("name", "").strip(),
                    location=row.get("location", "").strip(),
                    source=row.get("source", "").strip(),
                    cid_1c=row.get("client_id_1c", "").strip(),
                    company=row.get("company_name", "").strip(),
                    changed_by_tag="import_clients_csv",
                )
            except sqlite3.IntegrityError as e:
                rows_errored += 1
                print(f"[import_clients] row skipped on IntegrityError "
                      f"(phone={primary}): {e}")
                continue
            if outcome == "inserted":
                rows_inserted += 1
            elif outcome == "updated":
                rows_updated += 1
            elif outcome == "drift_held":
                rows_drift_held += 1

    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM allowed_clients").fetchone()[0]

    # Retroactively approve existing registered users whose phone matches whitelist
    existing_users = conn.execute(
        "SELECT telegram_id, phone FROM users WHERE phone IS NOT NULL"
    ).fetchall()
    approved_count = 0
    for u in existing_users:
        phone_norm = normalize_phone(u[1])
        match = conn.execute(
            "SELECT id FROM allowed_clients WHERE phone_normalized = ? AND COALESCE(status, 'active') NOT LIKE 'merged%' LIMIT 1",
            (phone_norm,),
        ).fetchone()
        if match:
            conn.execute(
                "UPDATE users SET is_approved = 1, client_id = ? WHERE telegram_id = ?",
                (match[0], u[0]),
            )
            approved_count += 1

    # Set is_approved to 0 for any NULL values (from migration)
    conn.execute("UPDATE users SET is_approved = 0 WHERE is_approved IS NULL")

    # Step 8 — mutator chokepoint (same rationale as apply_clients_upload).
    # Heal any orphan finance rows that the new/updated allowed_clients rows
    # now resolve. Idempotent; safe on every CLI run.
    from backend.services import client_identity
    orphans_healed = client_identity.heal_all_finance_tables(conn)

    conn.commit()
    conn.close()

    print(f"[import_clients] Inserted {rows_inserted}, updated {rows_updated}, "
          f"drift-held {rows_drift_held}, errored {rows_errored}. Total: {total}")
    if approved_count:
        print(f"[import_clients] Retroactively approved {approved_count} existing users.")
    if any(orphans_healed.values()):
        print(f"[import_clients] Orphan finance rows healed: {orphans_healed}")


_HEADER_ALIAS = {
    # phone variants
    "phone": "phone", "tel": "phone", "tel.": "phone",
    "telefon": "phone", "telefon raqam": "phone",
    "телефон": "phone", "телефоны": "phone",
    "телефон контрагента": "phone", "телефоны контрагента": "phone",
    "тел": "phone", "тел.": "phone",
    "тел. номер": "phone", "тел.номер": "phone", "тел номер": "phone",
    "контактный телефон": "phone", "контакт": "phone", "контакты": "phone",
    "phone number": "phone", "телефон номер": "phone",
    "mobile": "phone", "мобильный": "phone", "nomer": "phone", "номер": "phone",
    # name variants (1C "Наименование" is the short client name on Контрагенты)
    "name": "name", "ism": "name", "имя": "name", "nom": "name",
    "fish": "name", "fio": "name", "фио": "name",
    "klient": "name", "клиент": "name", "mijoz": "name",
    "ф.и.о": "name", "ф.и.о.": "name", "ф и о": "name",
    "наименование": "name", "название": "name", "наим.": "name",
    # location
    "location": "location", "manzil": "location", "адрес": "location",
    "address": "location", "город": "location",
    "юридический адрес": "location", "юр.адрес": "location", "юр адрес": "location",
    "почтовый адрес": "location", "фактический адрес": "location",
    # source
    "source": "source", "manba": "source", "источник": "source",
    # 1c id (Контрагент in 1C is the client row itself; keep it mapping to client_id_1c
    # because the human-readable 1C NAME is still the display label — the numeric
    # "Код" must NOT land in client_id_1c (it would overwrite the name).
    # Client Identity Anchoring Phase 0 (2026-06-03): "Код" IS captured now, but
    # to its own anchor field onec_card_id (folder-scoped → de-collided), NEVER to
    # client_id_1c. See _apply_folder_anchor + _upsert_client_from_row.
    "код": "onec_code", "kod": "onec_code",
    # "Вид контрагента" — empty cell marks a folder-header row (Покупатели /
    # Поставщики / Прочие); used only to track the current folder, never stored.
    "вид контрагента": "onec_vid", "вид": "onec_vid",
    "client_id_1c": "client_id_1c", "1c": "client_id_1c",
    "1c nomi": "client_id_1c", "1с nomi": "client_id_1c",
    "1c ismi": "client_id_1c", "1c name": "client_id_1c",
    "client 1c": "client_id_1c", "1c клиент": "client_id_1c",
    "1с клиент": "client_id_1c", "контрагент": "client_id_1c",
    "kontragent": "client_id_1c",
    # company (1C "Полное наименование" = legal entity form)
    "company": "company_name", "company_name": "company_name",
    "kompaniya": "company_name", "компания": "company_name",
    "firma": "company_name", "фирма": "company_name",
    "полное наименование": "company_name", "полн. наименование": "company_name",
    "юр.лицо": "company_name", "юрлицо": "company_name",
    "организация": "company_name",
}


def _normalize_headers(raw_headers: list) -> list:
    return [_HEADER_ALIAS.get(
        str(h or "").strip().lower().replace("  ", " "),
        str(h or "").strip().lower(),
    ) for h in raw_headers]


def _score_header_row(raw_row) -> int:
    """Return the number of canonical fields this row hits."""
    known = {"phone", "name", "client_id_1c", "company_name", "location", "source",
             "onec_code", "onec_vid"}
    return sum(1 for h in _normalize_headers(raw_row) if h in known)


def _find_header_row(table_rows, max_scan: int = 15) -> int:
    """1C exports often put the sheet title + blank rows + meta on rows 1-5 and
    the real header row below. Pick the earliest row with ≥ 2 alias hits;
    fall back to the single best-scoring row; if nothing scores return -1 so
    the caller knows no header was found (better than silently using row 0)."""
    best_idx, best_score = 0, -1
    for i, row in enumerate(table_rows[:max_scan]):
        s = _score_header_row(row)
        if s > best_score:
            best_score = s
            best_idx = i
        if s >= 2:
            return i
    if best_score <= 0:
        return -1
    return best_idx


def _normalize_card_code(v) -> str:
    """1C 'Код' cell → bare integer string ('1056'). xlrd yields floats, openpyxl
    yields ints/floats/strings; normalise all to the digits we anchor on."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _apply_folder_anchor(headers: list, rows: list) -> list:
    """Client Identity Anchoring Phase 0 — stamp ``onec_card_id`` on each data row.

    The 1C Контрагенты export groups rows under folder-header rows
    (Покупатели / Поставщики / Прочие) whose 'Вид контрагента' cell is empty and
    whose 'Наименование' cell holds the folder name. 'Код' is folder-scoped (it
    collides across folders), so the stable anchor is
    ``onec_card_id = "{folder}:{Код}"`` (e.g. ``"Прочие:1056"``). Walk the rows in
    order tracking the current folder, stamp each data row, and drop the
    folder-header rows themselves.

    **No-op unless the file has BOTH the Код and Вид контрагента columns** — any
    other upload (Client Master, manual phone-fix sheets) is returned untouched,
    so a missing-column file can never be misread as "all folder headers" and
    silently drop every data row.
    """
    if not ("onec_code" in headers and "onec_vid" in headers):
        return rows
    out, folder = [], None
    for r in rows:
        vid = str(r.get("onec_vid") or "").strip()
        if vid == "":
            # Folder-header row — its Наименование cell names the folder.
            fname = str(r.get("name") or "").strip()
            if fname:
                folder = fname
            continue  # never a real client row; drop it
        code = _normalize_card_code(r.get("onec_code"))
        if folder and code:
            r["onec_card_id"] = f"{folder}:{code}"
        out.append(r)
    return out


def _iter_rows_from_xlsx(file_bytes: bytes):
    """Return (headers_raw, list of dicts) from an xlsx, auto-detecting the
    real header row (1C exports often have a title row above it)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    hdr_idx = _find_header_row(rows)
    if hdr_idx < 0:
        # No alias-matching header row found — return raw preview so operator
        # can see what's in the file.
        preview = rows[0] if rows else tuple()
        header_raw = [("" if c is None else str(c)) for c in preview]
        return header_raw, []
    header_row = rows[hdr_idx]
    header_raw = [("" if c is None else str(c)) for c in header_row]
    headers = _normalize_headers(header_raw)
    out = []
    for row in rows[hdr_idx + 1:]:
        data = {headers[i]: (row[i] if i < len(row) else None)
                for i in range(len(headers))}
        out.append(data)
    out = _apply_folder_anchor(headers, out)
    return header_raw, out


def _iter_rows_from_xls(file_bytes: bytes):
    """Return (headers_raw, list of dicts) from a legacy .xls, auto-detecting
    the real header row."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=file_bytes, encoding_override="cp1251")
    sh = wb.sheet_by_index(0)
    if sh.nrows < 1:
        return [], []
    all_rows = [
        [sh.cell_value(r, c) for c in range(sh.ncols)]
        for r in range(min(sh.nrows, 15))
    ]
    hdr_idx = _find_header_row(all_rows)
    if hdr_idx < 0:
        # No alias-matching header row — surface raw row 0 to operator.
        preview = [str(sh.cell_value(0, c) or "") for c in range(sh.ncols)] if sh.nrows else []
        return preview, []
    header_raw = [str(sh.cell_value(hdr_idx, c) or "") for c in range(sh.ncols)]
    headers = _normalize_headers(header_raw)
    out = []
    for r in range(hdr_idx + 1, sh.nrows):
        row = {}
        for c in range(sh.ncols):
            v = sh.cell_value(r, c)
            if isinstance(v, float) and v.is_integer():
                v = str(int(v))
            row[headers[c]] = v
        out.append(row)
    out = _apply_folder_anchor(headers, out)
    return header_raw, out


def apply_clients_upload(file_bytes: bytes, filename_hint: str = "") -> dict:
    """Bot entry point: upsert allowed_clients from an uploaded xls/xlsx file.

    Required column: phone. Optional: name, location, source, client_id_1c,
    company_name. Matching behavior mirrors the CSV import.
    """
    name = (filename_hint or "").lower()
    try:
        if name.endswith(".xlsx"):
            header_raw, rows = _iter_rows_from_xlsx(file_bytes)
        else:
            header_raw, rows = _iter_rows_from_xls(file_bytes)
    except Exception as e:
        return {"ok": False, "error": f"Fayl o'qib bo'lmadi: {e}"}

    if not rows:
        return {"ok": False, "error": "Faylda ma'lumot topilmadi"}

    # If no row produced a valid phone, surface the header list so the operator
    # (and we) can see which columns the file actually has. Aliases can then
    # be added to _HEADER_ALIAS without guessing.
    any_phone = any(normalize_phone(str(r.get("phone") or "")) for r in rows)

    conn = sqlite3.connect(DATABASE_PATH)
    conn.execute("PRAGMA journal_mode=WAL")

    inserted = updated = skipped = drift_held = errored = 0
    seen = set()
    for raw in rows:
        raw_phone_str = str(raw.get("phone") or "")
        primary = normalize_phone(raw_phone_str)
        if not primary or primary in seen:
            skipped += 1
            continue
        seen.add(primary)

        # Per-row resilience: a single row's constraint violation must never
        # abort the whole upload (Error Log #85). The collision guard in
        # _upsert_client_from_row handles the known phone case; this catches any
        # other UNIQUE violation — log + count + continue (zero silent failure).
        try:
            outcome, _ = _upsert_client_from_row(
                conn,
                raw_phone_str=raw_phone_str,
                client_name=str(raw.get("name") or "").strip(),
                location=str(raw.get("location") or "").strip(),
                source=str(raw.get("source") or "clients_upload").strip(),
                cid_1c=str(raw.get("client_id_1c") or "").strip(),
                company=str(raw.get("company_name") or "").strip(),
                changed_by_tag="apply_clients_upload",
                onec_card_id=str(raw.get("onec_card_id") or "").strip() or None,
            )
        except sqlite3.IntegrityError as e:
            errored += 1
            print(f"[apply_clients_upload] row skipped on IntegrityError "
                  f"(phone={primary}): {e}")
            continue
        if outcome == "inserted":
            inserted += 1
        elif outcome == "updated":
            updated += 1
        elif outcome == "drift_held":
            drift_held += 1
        else:
            skipped += 1

    # Step 8 — mutator chokepoint. /clients runs after /debtors and /cash in the
    # daily upload order, so any allowed_clients row added/updated above may
    # unblock orphan finance rows that the per-row resolve couldn't link at
    # /debtors- or /cash-import time. Heal in same transaction → atomic at commit.
    from backend.services import client_identity
    orphans_healed = client_identity.heal_all_finance_tables(conn)

    conn.commit()
    total = conn.execute("SELECT COUNT(*) FROM allowed_clients").fetchone()[0]
    conn.close()

    return {
        "ok": True,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "drift_held": drift_held,
        "errored": errored,
        "total_clients": total,
        "orphans_healed": orphans_healed,
        "headers_seen": header_raw,
        "phone_column_detected": any_phone,
    }


if __name__ == "__main__":
    import_clients()
