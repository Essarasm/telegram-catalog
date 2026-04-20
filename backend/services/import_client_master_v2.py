"""Client Master v2 importer — reads the full-mirror xlsx produced by
`/exportmaster` and applies only the ✏️ editable columns back to
`allowed_clients`, with conflict detection, phone audit trail, upload
archiving, and chunked commits.

See obsidian-vault/Client Data Workflow — Design v0.1.md v0.1.3 for the
design invariants this implementation guarantees.
"""
from __future__ import annotations

import io
import os
import re
import sqlite3
import unicodedata
from datetime import datetime, timezone, timedelta
from typing import Optional

from openpyxl import load_workbook

TASHKENT = timezone(timedelta(hours=5))
ARCHIVE_DIR = "/data/master_archive"

# Column headers exactly as produced by `export_client_master.py`. If any
# column header changes there, this map must be updated. The ✏️ prefix
# marks editable columns — they come from the operator.
EDITABLE_HEADERS = {
    "✏️ Viloyat":         "viloyat",
    "✏️ Shahar/Tuman":    "tuman",
    "✏️ Mo'ljal":         "moljal",
    "✏️ Izoh":            "location",
    "✏️ Ism 01 / Familiya": "name",
    "✏️ Raqam 01":        "phone_normalized",  # special: re-normalized and may change
    "✏️ Ism 02":          "ism_02",
    "✏️ Raqam 02":        "raqam_02",
    "✏️ Ism 03":          "ism_03",
    "✏️ Raqam 03":        "raqam_03",
    "✏️ Mijoz holati":    "mijoz_holati",
    "✏️ Hajm":            "hajm",
    "✏️ Segment":         "segment",
    "✏️ Active status":   "status",     # 'active' | 'inactive' | 'merged'
    "✏️ Eslatmalar":      "eslatmalar",
}
ID_HEADER = "allowed_clients.id"
MASTER_ROW_HEADER = "master_row_id"

CHUNK_SIZE = 500


def _normalize_phone(raw: str) -> str:
    digits = re.sub(r"\D", "", (raw or ""))
    return digits[-9:] if len(digits) >= 9 else digits


def _cell_val(raw) -> Optional[str]:
    """Normalize a cell value to a stripped str or None."""
    if raw is None:
        return None
    s = str(raw).strip()
    return s if s else None


def _archive_upload(file_bytes: bytes) -> str:
    """Save the uploaded xlsx to master_archive with timestamp. Prune to last 12."""
    try:
        os.makedirs(ARCHIVE_DIR, exist_ok=True)
    except OSError:
        return ""
    ts = datetime.now(TASHKENT).strftime("%Y%m%d_%H%M%S")
    path = os.path.join(ARCHIVE_DIR, f"Client_Master_uploaded_{ts}.xlsx")
    try:
        with open(path, "wb") as f:
            f.write(file_bytes)
        # Retain last 12 uploaded files
        files = sorted([p for p in os.listdir(ARCHIVE_DIR) if p.startswith("Client_Master_uploaded_")])
        while len(files) > 12:
            old = files.pop(0)
            try:
                os.remove(os.path.join(ARCHIVE_DIR, old))
            except OSError:
                pass
        return path
    except OSError:
        return ""


def _record_phone_history(conn, client_id: int, old: Optional[str], new: Optional[str],
                          reason: str, changed_by: Optional[str]) -> None:
    conn.execute(
        """INSERT INTO phone_history (client_id, old_phone, new_phone, reason, changed_by)
           VALUES (?, ?, ?, ?, ?)""",
        (client_id, old, new, reason, changed_by),
    )


def apply_client_master_v2(
    file_bytes: bytes,
    uploaded_by_user_id: Optional[int] = None,
    uploaded_by_name: Optional[str] = None,
    db_path: Optional[str] = None,
) -> dict:
    """Apply the editable-column subset of an uploaded Client Master xlsx
    back to allowed_clients.

    Returns:
        {
          "ok": True/False,
          "error": "..." if not ok,
          "archive_path": "/data/master_archive/Client_Master_uploaded_YYYYMMDD_HHMMSS.xlsx",
          "totals": {
              "rows_seen": int, "updated": int, "unchanged": int,
              "conflicts": int, "phone_edits": int, "phone_collisions": int,
              "status_changes": int, "new_rows": int, "skipped": int,
              "commits": int,
          }
        }
    """
    # 1. Archive the upload up front
    archive_path = _archive_upload(file_bytes)

    # 2. Load workbook
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return {"ok": False, "error": f"Fayl o'qib bo'lmadi: {e}", "archive_path": archive_path}

    # Prefer "Client Master" sheet; fall back to first sheet if exported by
    # hand under a different name.
    if "Client Master" in wb.sheetnames:
        ws = wb["Client Master"]
    else:
        ws = wb[wb.sheetnames[0]]

    # 3. Read header row, build column → field map
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [(str(h) if h is not None else "").strip() for h in header_row]
    col_field: dict[int, str] = {}
    id_col = None
    master_row_col = None
    for idx, h in enumerate(headers):
        if h == ID_HEADER:
            id_col = idx
        elif h == MASTER_ROW_HEADER:
            master_row_col = idx
        elif h in EDITABLE_HEADERS:
            col_field[idx] = EDITABLE_HEADERS[h]

    if id_col is None:
        return {
            "ok": False,
            "error": f"'{ID_HEADER}' ustuni topilmadi — yangi /exportmaster formatidan foydalaning.",
            "archive_path": archive_path,
        }
    if not col_field:
        return {
            "ok": False,
            "error": "Hech qanday ✏️ ustun topilmadi.",
            "archive_path": archive_path,
        }

    # 4. Open DB
    db_path = db_path or os.environ.get("DATABASE_PATH", "/data/catalog.db")
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.row_factory = sqlite3.Row

    # Log the upload attempt
    log_row = conn.execute(
        """INSERT INTO master_upload_log
           (archived_file_path, uploaded_by_user_id, uploaded_by_name,
            row_count, inserted_count, updated_count, conflict_count)
           VALUES (?, ?, ?, 0, 0, 0, 0)""",
        (archive_path, uploaded_by_user_id, uploaded_by_name),
    )
    upload_log_id = log_row.lastrowid
    conn.commit()

    uploaded_source_tag = f"upload:{upload_log_id}"

    totals = {
        "rows_seen": 0, "updated": 0, "unchanged": 0, "conflicts": 0,
        "phone_edits": 0, "phone_collisions": 0,
        "status_changes": 0, "new_rows": 0, "skipped": 0, "commits": 0,
    }

    # 5. Iterate + chunk-commit
    since_commit = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        totals["rows_seen"] += 1

        # Resolve target row
        id_raw = row[id_col] if id_col < len(row) else None
        if id_raw is None:
            totals["skipped"] += 1
            continue
        try:
            ac_id = int(float(id_raw))
        except (TypeError, ValueError):
            totals["skipped"] += 1
            continue

        current = conn.execute(
            """SELECT id, phone_normalized, name, location, viloyat, tuman, moljal,
                      ism_02, raqam_02, ism_03, raqam_03,
                      mijoz_holati, hajm, segment, eslatmalar,
                      COALESCE(status, 'active') AS status, source_master
               FROM allowed_clients WHERE id = ?""",
            (ac_id,),
        ).fetchone()
        if not current:
            # Row with this id no longer exists (deleted or never existed)
            # — skip to avoid creating ghost rows.
            totals["skipped"] += 1
            continue

        updates: list[str] = []
        params: list = []
        row_conflict = False
        new_phone: Optional[str] = None
        had_phone_edit = False

        # Collect proposed edits
        for col_idx, field in col_field.items():
            incoming = _cell_val(row[col_idx]) if col_idx < len(row) else None
            existing = current[field] if field in current.keys() else None
            if field == "phone_normalized":
                # Incoming phone is formatted like "90 123-45-67"; re-normalize.
                norm = _normalize_phone(incoming or "")
                if norm == (existing or ""):
                    continue
                new_phone = norm
                had_phone_edit = True
                continue  # handled specially below (collision + audit)
            if field == "status":
                allowed = {"active", "inactive", "merged"}
                val = (incoming or "").lower() if incoming else "active"
                if val not in allowed:
                    val = "active"
                if val != (existing or "active"):
                    updates.append("status = ?")
                    params.append(val)
                    totals["status_changes"] += 1
                continue
            # Generic column — normalise empty → NULL
            val = incoming if incoming else None
            # No change
            if (val or None) == (existing or None):
                continue
            updates.append(f"{field} = ?")
            params.append(val)

        # Phone edit: audit + collision check
        if had_phone_edit and new_phone:
            if not new_phone:
                pass  # empty phone — skip (would break dedup key)
            else:
                # Collision check: does another row already have this phone?
                clash = conn.execute(
                    "SELECT id, client_id_1c FROM allowed_clients WHERE phone_normalized = ? AND id != ?",
                    (new_phone, ac_id),
                ).fetchone()
                if clash:
                    totals["phone_collisions"] += 1
                    # Flag BOTH rows for review; do NOT auto-merge.
                    conn.execute(
                        "UPDATE allowed_clients SET needs_review = 1 WHERE id IN (?, ?)",
                        (ac_id, clash["id"]),
                    )
                    _record_phone_history(
                        conn, ac_id, current["phone_normalized"], new_phone,
                        reason=f"collision_with_{clash['id']}",
                        changed_by=uploaded_by_name,
                    )
                    row_conflict = True
                else:
                    # Safe phone replacement
                    updates.append("phone_normalized = ?")
                    params.append(new_phone)
                    _record_phone_history(
                        conn, ac_id, current["phone_normalized"], new_phone,
                        reason="master_upload",
                        changed_by=uploaded_by_name,
                    )
                    totals["phone_edits"] += 1

        if not updates and not row_conflict:
            totals["unchanged"] += 1
            continue

        # Mark bookkeeping columns
        updates.append("source_master = ?")
        params.append(uploaded_source_tag)
        updates.append("last_master_synced_at = ?")
        params.append(datetime.now(TASHKENT).isoformat(timespec="seconds"))
        if row_conflict:
            updates.append("needs_review = 1")

        params.append(ac_id)
        conn.execute(
            f"UPDATE allowed_clients SET {', '.join(updates)} WHERE id = ?",
            params,
        )
        if row_conflict:
            totals["conflicts"] += 1
        totals["updated"] += 1
        since_commit += 1

        # Chunked commit
        if since_commit >= CHUNK_SIZE:
            conn.commit()
            totals["commits"] += 1
            since_commit = 0

    # Final commit
    if since_commit > 0:
        conn.commit()
        totals["commits"] += 1

    # Update log row with final counts
    conn.execute(
        """UPDATE master_upload_log
           SET row_count = ?, inserted_count = ?, updated_count = ?, conflict_count = ?
           WHERE id = ?""",
        (totals["rows_seen"], totals["new_rows"], totals["updated"],
         totals["conflicts"], upload_log_id),
    )
    conn.commit()
    conn.close()

    return {
        "ok": True,
        "archive_path": archive_path,
        "upload_log_id": upload_log_id,
        "totals": totals,
    }
