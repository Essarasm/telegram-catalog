"""Refresh product catalog from an Excel file without full redeploy.

Supports two formats:
1. 'Catalog Clean' (Rassvet_Master.xlsx) — full catalog refresh with categories/producers
2. '1C Номенклатура' export — price/weight sync + new product detection

For Catalog Clean format:
- Adds NEW products with proper category/producer linking
- Updates existing products (price, weight, unit changes)
- Marks products NOT in the new Excel as inactive (is_active=0)
- Reactivates previously deactivated products if they reappear

For 1C format:
- Updates prices and weights of matched products (with safety guards)
- Reports new products not yet in catalog (requires manual categorization)
- Does NOT deactivate products (1C export may be a subset)

Both modes are non-destructive — products are never deleted, only toggled.
"""
import io
import re
import logging
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from backend.database import get_db, init_db, build_search_text

logger = logging.getLogger(__name__)

# Safety thresholds (shared with update_prices.py)
PLACEHOLDER_PRICE = 0.10  # 1C placeholder is $0.09 — block only this, not real low prices
MAX_DROP_PCT = 80

# 1C column indices (0-based, from Номенклатура export)
C1_NAME = 1          # Наименование
C1_TYPE = 2          # Тип номенклатуры (== "Товар" for products)
C1_UNIT = 5          # Единица измерения
C1_UZS = 6           # Цена (UZS)
C1_USD = 15          # ЦенаВал (wholesale USD)
C1_WEIGHT = 18       # Вес


# ── Cyrillic → Latin transliteration (duplicated from import_products for independence)
CYRILLIC_MAP = {
    'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E',
    'Ё': 'Yo', 'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K',
    'Л': 'L', 'М': 'M', 'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R',
    'С': 'S', 'Т': 'T', 'У': 'U', 'Ф': 'F', 'Х': 'Kh', 'Ц': 'Ts',
    'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch', 'Ъ': '', 'Ы': 'Y', 'Ь': '',
    'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya',
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e',
    'ё': 'yo', 'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k',
    'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r',
    'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'kh', 'ц': 'ts',
    'ч': 'ch', 'ш': 'sh', 'щ': 'shch', 'ъ': '', 'ы': 'y', 'ь': '',
    'э': 'e', 'ю': 'yu', 'я': 'ya',
}


def transliterate(text):
    if not text:
        return text
    return ''.join(CYRILLIC_MAP.get(ch, ch) for ch in text)


def cyrillic_title_case(text):
    words = text.split()
    result = []
    for word in words:
        if not any('\u0400' <= c <= '\u04ff' for c in word):
            result.append(word)
        elif len(word) <= 3 and word.isupper():
            result.append(word)
        else:
            result.append(word[0].upper() + word[1:].lower() if len(word) > 1 else word.upper())
    return ' '.join(result)


def normalize_name(name: str) -> str:
    """Normalize a product name for matching."""
    if not name:
        return ""
    n = name.strip().lower()
    n = re.sub(r'\s+', ' ', n)
    return n


def detect_format(file_bytes: bytes) -> str:
    """Detect whether the file is 'catalog_clean' or '1c_nomenklatura'.

    Returns 'catalog_clean' if file has a 'Catalog Clean' or 'Katalog' sheet,
    or '1c_nomenklatura' if it looks like a 1C Номенклатура export.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()

        # Check for known Catalog Clean sheet names
        if 'Catalog Clean' in sheets or 'Katalog' in sheets:
            return 'catalog_clean'

        # Check for 1C signature: first cell usually contains "Справочник"
        df = pd.read_excel(io.BytesIO(file_bytes), header=None, nrows=5)
        first_cell = str(df.iloc[0, 0]) if len(df) > 0 else ""
        if 'Справочник' in first_cell or 'Номенклатура' in first_cell:
            return '1c_nomenklatura'

        # Check if col 2 has "Тип номенклатуры" in header row
        for i in range(min(3, len(df))):
            row_vals = [str(v) for v in df.iloc[i] if pd.notna(v)]
            if any('Тип номенклатуры' in v for v in row_vals):
                return '1c_nomenklatura'

        # Default: assume 1C format if no Catalog Clean sheet
        return '1c_nomenklatura'
    except Exception as e:
        logger.warning(f"Format detection failed: {e}, defaulting to 1c_nomenklatura")
        return '1c_nomenklatura'


def refresh_catalog_from_excel(file_bytes: bytes) -> dict:
    """Refresh catalog from uploaded Excel. Auto-detects format.

    Returns summary of changes.
    """
    fmt = detect_format(file_bytes)
    logger.info(f"Detected format: {fmt}")

    if fmt == 'catalog_clean':
        return _refresh_from_catalog_clean(file_bytes)
    else:
        return _refresh_from_1c(file_bytes)


# ────────────────────────────────────────────────────────────────────
# Mode 1: Catalog Clean (Rassvet_Master) — full catalog refresh
# ────────────────────────────────────────────────────────────────────

def _refresh_from_catalog_clean(file_bytes: bytes) -> dict:
    """Full catalog refresh from Catalog Clean / Katalog sheet."""
    init_db()
    conn = get_db()

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Try known sheet names
    ws = None
    for sheet_name in ['Catalog Clean', 'Katalog']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active
        if ws is None:
            wb.close()
            return {"ok": False, "error": "No sheets found in Excel file"}

    db_products_before = conn.execute(
        "SELECT COUNT(*) FROM products WHERE is_active = 1"
    ).fetchone()[0]

    # Build lookup of existing products by normalized Cyrillic name
    existing_products = conn.execute(
        "SELECT id, name, name_display, category_id, producer_id, price_usd, price_uzs, weight, unit, is_active FROM products"
    ).fetchall()
    db_by_name = {}
    for p in existing_products:
        norm = normalize_name(p["name"])
        if norm not in db_by_name:
            db_by_name[norm] = p

    # Existing categories and producers
    cat_rows = conn.execute("SELECT id, name FROM categories").fetchall()
    cat_map = {r["name"]: r["id"] for r in cat_rows}

    prod_rows = conn.execute("SELECT id, name FROM producers").fetchall()
    prod_map = {r["name"]: r["id"] for r in prod_rows}

    # Process Excel rows
    rows = list(ws.iter_rows(min_row=2, values_only=False))

    excel_names_seen = set()
    new_products = 0
    updated_products = 0
    reactivated_products = 0
    new_product_names = []

    for row in rows:
        category = row[0].value    # A
        producer = row[1].value    # B
        name = row[2].value        # C — curated Latin display name
        weight_val = row[3].value if len(row) > 3 else None
        unit = row[4].value if len(row) > 4 else 'sht'
        price_uzs = row[5].value if len(row) > 5 else None
        price_usd = row[6].value if len(row) > 6 else None
        original_cyrillic_col = row[11].value if len(row) > 11 else None

        if not category or not producer or not name:
            continue

        category = str(category).strip()
        producer = str(producer).strip()
        name = str(name).strip()
        unit = str(unit).strip() if unit else 'sht'

        producer_latin = transliterate(cyrillic_title_case(producer))
        original_cyrillic = str(original_cyrillic_col).strip() if original_cyrillic_col else name

        p_usd = 0
        p_uzs = 0
        try:
            if price_usd is not None and price_usd != '' and price_usd != 0:
                p_usd = float(price_usd)
        except (ValueError, TypeError):
            pass
        try:
            if price_uzs is not None and price_uzs != '' and price_uzs != 0:
                p_uzs = float(price_uzs)
        except (ValueError, TypeError):
            pass

        weight = None
        try:
            if weight_val is not None and weight_val != '':
                weight = float(weight_val)
        except (ValueError, TypeError):
            pass

        display_name = name
        if any('\u0400' <= c <= '\u04ff' for c in display_name):
            display_name = transliterate(cyrillic_title_case(display_name))

        # Ensure category exists
        if category not in cat_map:
            conn.execute(
                "INSERT OR IGNORE INTO categories (name, sort_order) VALUES (?, ?)",
                (category, len(cat_map) + 1)
            )
            cat_id = conn.execute("SELECT id FROM categories WHERE name = ?", (category,)).fetchone()[0]
            cat_map[category] = cat_id

        # Ensure producer exists
        if producer_latin not in prod_map:
            conn.execute(
                "INSERT OR IGNORE INTO producers (name) VALUES (?)", (producer_latin,)
            )
            prod_id = conn.execute("SELECT id FROM producers WHERE name = ?", (producer_latin,)).fetchone()[0]
            prod_map[producer_latin] = prod_id

        norm_cyrillic = normalize_name(original_cyrillic)
        excel_names_seen.add(norm_cyrillic)

        existing = db_by_name.get(norm_cyrillic)

        if existing:
            changes = []
            if p_usd > 0 and abs((existing["price_usd"] or 0) - p_usd) > 0.001:
                changes.append(("price_usd", p_usd))
            if p_uzs > 0 and abs((existing["price_uzs"] or 0) - p_uzs) > 0.5:
                changes.append(("price_uzs", p_uzs))
            if weight and abs((existing["weight"] or 0) - weight) > 0.001:
                changes.append(("weight", weight))
            if existing["category_id"] != cat_map[category]:
                changes.append(("category_id", cat_map[category]))
            if existing["producer_id"] != prod_map[producer_latin]:
                changes.append(("producer_id", prod_map[producer_latin]))

            if not existing["is_active"]:
                changes.append(("is_active", 1))
                reactivated_products += 1

            if changes:
                set_clauses = ", ".join(f"{col} = ?" for col, _ in changes)
                values = [v for _, v in changes] + [existing["id"]]
                conn.execute(f"UPDATE products SET {set_clauses} WHERE id = ?", values)
                updated_products += 1
        else:
            search_text = build_search_text(original_cyrillic, display_name, producer_latin)
            conn.execute(
                """INSERT INTO products
                   (name, name_display, category_id, producer_id, unit,
                    price_usd, price_uzs, weight, is_active, search_text)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)""",
                (original_cyrillic, display_name, cat_map[category], prod_map[producer_latin],
                 unit, p_usd, p_uzs, weight, search_text)
            )
            new_products += 1
            new_product_names.append(display_name[:50])

    # Deactivate products NOT in the new Excel
    deactivated_products = 0
    deactivated_names = []
    for p in existing_products:
        norm = normalize_name(p["name"])
        if norm not in excel_names_seen and p["is_active"]:
            conn.execute("UPDATE products SET is_active = 0 WHERE id = ?", (p["id"],))
            deactivated_products += 1
            deactivated_names.append((p["name_display"] or p["name"])[:50])

    # Update denormalized counts
    conn.execute("""
        UPDATE categories SET product_count = (
            SELECT COUNT(*) FROM products WHERE products.category_id = categories.id AND is_active = 1
        )
    """)
    conn.execute("""
        UPDATE producers SET product_count = (
            SELECT COUNT(*) FROM products WHERE products.producer_id = producers.id AND is_active = 1
        )
    """)

    conn.commit()
    conn.close()
    wb.close()

    return {
        "ok": True,
        "format": "catalog_clean",
        "excel_products": len(excel_names_seen),
        "db_products_before": db_products_before,
        "new_products": new_products,
        "updated_products": updated_products,
        "reactivated_products": reactivated_products,
        "deactivated_products": deactivated_products,
        "new_product_names": new_product_names[:20],
        "deactivated_names": deactivated_names[:20],
    }


# ────────────────────────────────────────────────────────────────────
# Mode 2: 1C Номенклатура — price/weight sync + new product detection
# ────────────────────────────────────────────────────────────────────

def _refresh_from_1c(file_bytes: bytes) -> dict:
    """Sync prices/weights from 1C Номенклатура export.

    Does NOT add new products (no category/producer in 1C).
    Does NOT deactivate products (1C may be a subset).
    Reports new product names for manual categorization.
    """
    init_db()
    conn = get_db()

    df = pd.read_excel(io.BytesIO(file_bytes), header=None)
    products = df[(df[C1_TYPE] == 'Товар') & df[C1_NAME].notna()]

    if products.empty:
        conn.close()
        return {"ok": False, "error": "No products found (no rows with Тип='Товар')"}

    # Build DB lookups
    db_products = conn.execute(
        "SELECT id, name, name_display, price_usd, price_uzs, weight FROM products WHERE is_active = 1"
    ).fetchall()

    db_by_exact = {}
    db_by_normalized = {}
    for p in db_products:
        db_name = p["name"].strip()
        db_by_exact[db_name] = p
        norm = normalize_name(db_name)
        if norm not in db_by_normalized:
            db_by_normalized[norm] = p

    matched = 0
    updated = 0
    price_changes = []
    weight_changes = []
    skipped_low_price = 0
    skipped_big_drop = 0
    placeholder_fixes = 0
    new_in_1c = []  # Products in 1C but not in our DB

    for _, row in products.iterrows():
        name_1c = str(row[C1_NAME]).strip()
        usd_1c = pd.to_numeric(row[C1_USD], errors='coerce')
        weight_1c = pd.to_numeric(row[C1_WEIGHT], errors='coerce')

        # Match against DB
        product = None
        if name_1c in db_by_exact:
            product = db_by_exact[name_1c]
        else:
            norm = normalize_name(name_1c)
            if norm in db_by_normalized:
                product = db_by_normalized[norm]

        if not product:
            if len(name_1c) > 3:  # Skip junk like "3"
                new_in_1c.append(name_1c[:60])
            continue

        matched += 1
        changes_made = False

        # ── Price update with safety guards ──
        if pd.notna(usd_1c) and usd_1c > 0:
            old_usd = product["price_usd"] or 0

            if usd_1c < PLACEHOLDER_PRICE:
                if old_usd >= PLACEHOLDER_PRICE:
                    skipped_low_price += 1
                # Skip: incoming price is a placeholder
            elif old_usd > PLACEHOLDER_PRICE and usd_1c < old_usd:
                drop_pct = (old_usd - usd_1c) / old_usd * 100
                if drop_pct > MAX_DROP_PCT:
                    skipped_big_drop += 1
                else:
                    # Legitimate price decrease
                    if abs(old_usd - usd_1c) > 0.001:
                        conn.execute("UPDATE products SET price_usd = ? WHERE id = ?",
                                     (float(usd_1c), product["id"]))
                        price_changes.append({
                            "name": (product["name_display"] or product["name"])[:50],
                            "old": old_usd, "new": float(usd_1c)
                        })
                        changes_made = True
            else:
                # Price increase or placeholder fix
                if abs(old_usd - usd_1c) > 0.001:
                    if old_usd < PLACEHOLDER_PRICE:
                        placeholder_fixes += 1
                    conn.execute("UPDATE products SET price_usd = ? WHERE id = ?",
                                 (float(usd_1c), product["id"]))
                    price_changes.append({
                        "name": (product["name_display"] or product["name"])[:50],
                        "old": old_usd, "new": float(usd_1c)
                    })
                    changes_made = True

        # ── Weight update ──
        if pd.notna(weight_1c) and weight_1c > 0:
            old_weight = product["weight"] or 0
            if abs(old_weight - weight_1c) > 0.001:
                conn.execute("UPDATE products SET weight = ? WHERE id = ?",
                             (float(weight_1c), product["id"]))
                weight_changes.append({
                    "name": (product["name_display"] or product["name"])[:50],
                    "old": old_weight, "new": float(weight_1c)
                })
                changes_made = True

        if changes_made:
            updated += 1

    conn.commit()
    conn.close()

    return {
        "ok": True,
        "format": "1c_nomenklatura",
        "excel_products": len(products),
        "db_products_before": len(db_products),
        "matched": matched,
        "updated_products": updated,
        "price_changes": len(price_changes),
        "weight_changes": len(weight_changes),
        "placeholder_fixes": placeholder_fixes,
        "skipped_low_price": skipped_low_price,
        "skipped_big_drop": skipped_big_drop,
        "new_in_1c": new_in_1c[:30],
        "new_in_1c_total": len(new_in_1c),
        "sample_price_changes": price_changes[:15],
        # No deactivation in 1C mode
        "new_products": 0,
        "deactivated_products": 0,
        "reactivated_products": 0,
    }
