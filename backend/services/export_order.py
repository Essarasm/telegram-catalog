"""Generate order exports as PDF or Excel."""
import io
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ---------------------------------------------------------------------------
# Register Cyrillic-capable font (DejaVu Sans) — fixes black squares
# ---------------------------------------------------------------------------
_FONTS_DIR = Path(__file__).resolve().parent.parent / "fonts"

# Try bundled font first, then system font paths
_DEJAVU_PATHS = [
    _FONTS_DIR / "DejaVuSans.ttf",
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    Path("/usr/share/fonts/dejavu/DejaVuSans.ttf"),
]
_DEJAVU_BOLD_PATHS = [
    _FONTS_DIR / "DejaVuSans-Bold.ttf",
    Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    Path("/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"),
]

_FONT_NAME = "Helvetica"       # fallback
_FONT_BOLD = "Helvetica-Bold"  # fallback

for p in _DEJAVU_PATHS:
    if p.exists():
        pdfmetrics.registerFont(TTFont("DejaVuSans", str(p)))
        _FONT_NAME = "DejaVuSans"
        break

for p in _DEJAVU_BOLD_PATHS:
    if p.exists():
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", str(p)))
        _FONT_BOLD = "DejaVuSans-Bold"
        break


def generate_pdf(items: List[Dict], client_name: str = "") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'Title_UZ', parent=styles['Title'], fontName=_FONT_BOLD,
        fontSize=16, spaceAfter=12
    )
    normal_style = ParagraphStyle(
        'Normal_UZ', parent=styles['Normal'], fontName=_FONT_NAME,
        fontSize=10
    )

    elements = []

    # Header
    elements.append(Paragraph("BUYURTMA / ЗАКАЗ", title_style))
    elements.append(Spacer(1, 5*mm))

    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    info_lines = [f"Sana: {now}"]
    if client_name:
        info_lines.append(f"Mijoz: {client_name}")
    for line in info_lines:
        elements.append(Paragraph(line, normal_style))
    elements.append(Spacer(1, 8*mm))

    # Split items by currency
    usd_items = [it for it in items if it.get("currency", "USD") == "USD"]
    uzs_items = [it for it in items if it.get("currency", "USD") == "UZS"]

    def build_table(item_list, currency_label):
        header = ["#", "Mahsulot nomi", "Birlik", "Miqdor",
                  f"Narx ({currency_label})", f"Jami ({currency_label})"]
        data = [header]
        grand_total = 0

        for i, item in enumerate(item_list, 1):
            qty = item.get("quantity", 1)
            price = item.get("price", 0)
            total = qty * price
            grand_total += total
            data.append([
                str(i),
                item.get("name", ""),
                item.get("unit", "шт"),
                str(qty),
                f"{price:,.2f}",
                f"{total:,.2f}",
            ])

        data.append(["", "", "", "", "JAMI:", f"{grand_total:,.2f}"])

        col_widths = [25, 200, 45, 45, 70, 70]
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, -2), _FONT_NAME),
            ('FONTNAME', (0, 0), (-1, 0), _FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F3F4F6')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E5E7EB')),
            ('FONTNAME', (0, -1), (-1, -1), _FONT_BOLD),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        return table

    if usd_items:
        elements.append(Paragraph("Mahsulotlar (USD)", normal_style))
        elements.append(Spacer(1, 3*mm))
        elements.append(build_table(usd_items, "USD"))
        elements.append(Spacer(1, 8*mm))

    if uzs_items:
        elements.append(Paragraph("Mahsulotlar (UZS)", normal_style))
        elements.append(Spacer(1, 3*mm))
        elements.append(build_table(uzs_items, "UZS"))

    doc.build(elements)
    return buffer.getvalue()


def generate_excel(items: List[Dict], client_name: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Buyurtma"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2563EB")
    total_fill = PatternFill("solid", fgColor="E5E7EB")
    total_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "BUYURTMA / ЗАКАЗ"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws['A2'] = f"Sana: {now}"
    if client_name:
        ws['A3'] = f"Mijoz: {client_name}"

    usd_items = [it for it in items if it.get("currency", "USD") == "USD"]
    uzs_items = [it for it in items if it.get("currency", "USD") == "UZS"]

    current_row = 5

    def write_section(item_list, currency_label, start_row):
        ws.cell(row=start_row, column=1, value=f"Mahsulotlar ({currency_label})")
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=11)
        start_row += 1

        headers = ["#", "Mahsulot nomi", "Birlik", "Miqdor",
                   f"Narx ({currency_label})", f"Jami ({currency_label})"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        data_start = start_row + 1
        for i, item in enumerate(item_list, 1):
            row = data_start + i - 1
            qty = item.get("quantity", 1)
            price = item.get("price", 0)

            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=item.get("name", "")).border = thin_border
            ws.cell(row=row, column=3, value=item.get("unit", "шт")).border = thin_border
            qty_cell = ws.cell(row=row, column=4, value=qty)
            qty_cell.border = thin_border
            qty_cell.alignment = Alignment(horizontal='right')
            price_cell = ws.cell(row=row, column=5, value=price)
            price_cell.border = thin_border
            price_cell.number_format = '#,##0.00'
            price_cell.alignment = Alignment(horizontal='right')
            # Computed line total (not formula — works in all viewers)
            line_total = qty * price
            total_cell = ws.cell(row=row, column=6, value=line_total)
            total_cell.border = thin_border
            total_cell.number_format = '#,##0.00'
            total_cell.alignment = Alignment(horizontal='right')

        # Computed grand total
        grand_total = sum(it.get("quantity", 1) * it.get("price", 0) for it in item_list)
        total_row = data_start + len(item_list)
        ws.cell(row=total_row, column=5, value="JAMI:").font = total_font
        ws.cell(row=total_row, column=5).fill = total_fill
        ws.cell(row=total_row, column=5).border = thin_border
        grand_cell = ws.cell(row=total_row, column=6, value=grand_total)
        grand_cell.font = total_font
        grand_cell.fill = total_fill
        grand_cell.border = thin_border
        grand_cell.number_format = '#,##0.00'

        return total_row + 2

    if usd_items:
        current_row = write_section(usd_items, "USD", current_row)

    if uzs_items:
        current_row = write_section(uzs_items, "UZS", current_row)

    # Auto-width
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
