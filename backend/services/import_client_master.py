"""
Import Client Master spreadsheet into allowed_clients.

Source: a curated XLSX (`Client Master 16.03.26.xlsx` and successors) maintained
by ops. The spreadsheet has multiple sheets; we read two:

  • `Contacts` (~1633 rows) — main directory. Critical column for /realorders
    linking is **`Original Ism (1C)`** (cyrillic name as it appears in 1C
    `Реализация товаров` exports). Latin name (`Ism 01` + `Familiya`) is stored
    in `company_name` for display.

  • `Usto` (~57 rows) — sub-clients/contractors. No `Original Ism (1C)` column;
    we synthesize a name from `Ism 01 [Familiya] (Izoh)` so the row is at least
    addressable from the admin side. These rows still help phone-based user
    auto-approval even though they will not match cyrillic real_orders names.

Per-row behavior:

  • Each row may have up to 3 phones (Raqam 01/02/03). We expand to ONE
    `allowed_clients` row PER PHONE — sharing the same `name`, `company_name`,
    `location`, `source_sheet`. This matches existing phone-keyed lookups in
    `_try_match_client` (import_balances.py) and `users.check`.

  • If a row has NO phones, we still insert ONE phoneless row (phone_normalized = '')
    so the cyrillic name is reachable for `relink_real_orders`. Phoneless dedupe is
    by python-normalized name (cyrillic-aware lower + ё→е + whitespace collapse).

  • Existing rows are matched (and updated, never overwritten with empty values)
    by phone first; phoneless updates fall back to name match.

  • After import we retroactively approve any registered users whose phone
    matches a freshly imported `allowed_clients` row, mirroring the existing
    `import_clients.import_clients()` logic.

This module deliberately reuses `_py_normalize_client_name` from
`import_real_orders` to keep the cyrillic-normalization pipeline single-sourced
with the relink pass.
"""
from __future__ import annotations

import io
import os
import re
import sqlite3
from typing import Any, Dict, List, Optional, Tuple

from backend.services.import_real_orders import _py_normalize_client_name

DATABASE_PATH = os.getenv("DATABASE_PATH", "/data/catalog.db")

CONTACTS_SHEET = "Contacts"
USTO_SHEET = "Usto"

# Spreadsheet column header names → internal field key.
# We resolve column INDEX from header text at parse time so a column re-order
# in the xlsx does not break the importer.
CONTACT_COLS = {
    "viloyat": "Viloyat",
    "tuman": "Shahar/Tuman",
    "moljal": "Mo'ljal",
    "izoh": "Izoh",
    "ism01": "Ism 01",
    "familiya": "Familiya",
    "raqam01": "Raqam 01",
    "raqam02": "Raqam 02",
    "raqam03": "Raqam 03",
    "original_1c": "Original Ism (1C)",
}

USTO_COLS = {
    "viloyat": "Viloyat",
    "tuman": "Shahar/Tuman",
    "moljal": "Mo'ljal",
    "izoh": "Izoh",
    "ism01": "Ism 01",
    "familiya": "Familiya",
    "raqam01": "Raqam 01",
    "raqam02": "Raqam 02",
    "raqam03": "Raqam 03",
}


def _normalize_phone(raw: Any) -> str:
    """Strip to last 9 digits (Uzbek local number without country code).

    Identical semantics to `import_clients.normalize_phone`. Re-implemented
    here to avoid a circular import (`import_clients` would otherwise pull in
    the migration path on first call in test runs).
    """
    if raw is None:
        return ""
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) >= 9:
        return digits[-9:]
    return ""  # too short to be a usable Uzbek mobile


def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _resolve_columns(header_row: List[Any], wanted: Dict[str, str]) -> Dict[str, int]:
    """Map field-key → column index using a header row.

    Header matching is case-sensitive on the trimmed text, exactly as the
    spreadsheet uses Mixed-case Latin labels. Missing optional columns are
    silently dropped (the caller checks `.get`).
    """
    by_text = {}
    for i, cell in enumerate(header_row):
        txt = _str(cell)
        if txt and txt not in by_text:
            by_text[txt] = i
    return {key: by_text[label] for key, label in wanted.items() if label in by_text}


def _build_latin_name(ism: str, familiya: str, izoh: str) -> str:
    """Compose a human-readable latin label from Ism, Familiya, Izoh.

    Examples:
      ('Akram', '', '')          -> 'Akram'
      ('Abdulaziz', '', 'Usto. Katt-N') -> 'Abdulaziz (Usto. Katt-N)'
      ('Aziz', 'Karimov', '')    -> 'Aziz Karimov'
    """
    parts = [p for p in (ism, familiya) if p]
    base = " ".join(parts).strip()
    if izoh:
        if base:
            return f"{base} ({izoh})"
        return izoh
    return base


def _build_location(viloyat: str, tuman: str, moljal: str) -> str:
    parts = [p for p in (viloyat, tuman, moljal) if p]
    return ", ".join(parts)


def apply_client_master_import(file_bytes: bytes, filename_hint: str = "") -> Dict[str, Any]:
    """Parse a Client Master xlsx blob and upsert into allowed_clients.

    Returns a JSON-friendly summary including per-sheet insert/update counts,
    skipped rows (no phone AND no name), retroactively approved users, and the
    new total of `allowed_clients`.

    Designed to be re-runnable: existing rows are updated in place and
    non-empty source values never overwrite existing non-empty target values
    when they would clobber an operator edit (we use COALESCE-on-update so
    blanks in the spreadsheet do not erase populated DB fields).
    """
    if not file_bytes:
        return {"ok": False, "error": "Empty file"}

    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl not installed"}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return {"ok": False, "error": f"Could not open workbook: {exc}"}

    conn = sqlite3.connect(DATABASE_PATH)
    conn.execute("PRAGMA journal_mode=WAL")

    # Build in-memory indexes of existing allowed_clients so the importer is
    # O(N) over the spreadsheet rather than O(N×rows in DB).
    phone_index: Dict[str, int] = {}     # phone_normalized -> id
    name_index: Dict[str, int] = {}      # py-normalized name -> id (used for
                                         # phoneless dedupe path only)
    for rid, ph, nm in conn.execute(
        "SELECT id, phone_normalized, name FROM allowed_clients"
    ):
        if ph:
            phone_index.setdefault(ph, rid)
        if nm:
            name_index.setdefault(_py_normalize_client_name(nm), rid)

    totals = {
        "inserted": 0,
        "updated": 0,
        "skipped_empty": 0,
        "phoneless_inserted": 0,
    }
    sheets_summary: List[Dict[str, Any]] = []

    def _upsert(
        *,
        phone: str,
        name: str,
        company_name: str,
        location: str,
        source: str,
    ) -> str:
        """Insert or update one allowed_clients row.

        Returns one of: 'inserted', 'updated', 'phoneless_inserted'.
        """
        if phone:
            existing = phone_index.get(phone)
            if existing is not None:
                conn.execute(
                    """UPDATE allowed_clients
                       SET name         = COALESCE(NULLIF(?, ''), name),
                           company_name = COALESCE(NULLIF(?, ''), company_name),
                           location     = COALESCE(NULLIF(?, ''), location),
                           source_sheet = ?
                       WHERE id = ?""",
                    (name, company_name, location, source, existing),
                )
                if name:
                    name_index.setdefault(_py_normalize_client_name(name), existing)
                return "updated"
            cur = conn.execute(
                """INSERT INTO allowed_clients
                   (phone_normalized, name, company_name, location, source_sheet)
                   VALUES (?, ?, ?, ?, ?)""",
                (phone, name, company_name, location, source),
            )
            new_id = cur.lastrowid
            phone_index[phone] = new_id
            if name:
                name_index.setdefault(_py_normalize_client_name(name), new_id)
            return "inserted"

        # Phoneless path: dedupe on cyrillic-aware name key.
        if not name:
            return "skipped"
        key = _py_normalize_client_name(name)
        existing = name_index.get(key)
        if existing is not None:
            conn.execute(
                """UPDATE allowed_clients
                   SET company_name = COALESCE(NULLIF(?, ''), company_name),
                       location     = COALESCE(NULLIF(?, ''), location),
                       source_sheet = ?
                   WHERE id = ?""",
                (company_name, location, source, existing),
            )
            return "updated"
        cur = conn.execute(
            """INSERT INTO allowed_clients
               (phone_normalized, name, company_name, location, source_sheet)
               VALUES ('', ?, ?, ?, ?)""",
            (name, company_name, location, source),
        )
        new_id = cur.lastrowid
        name_index[key] = new_id
        return "phoneless_inserted"

    def _process_sheet(sheet_name: str, col_spec: Dict[str, str], use_cyrillic_name: bool) -> Dict[str, Any]:
        if sheet_name not in wb.sheetnames:
            return {"sheet": sheet_name, "skipped": "sheet not found"}
        ws = wb[sheet_name]
        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        col = _resolve_columns(header, col_spec)
        if not col:
            return {"sheet": sheet_name, "skipped": "no recognized columns"}

        s_inserted = 0
        s_updated = 0
        s_skipped = 0
        s_phoneless_inserted = 0
        rows_seen = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or v == "" for v in row):
                continue
            rows_seen += 1

            def cell(key: str) -> str:
                idx = col.get(key)
                if idx is None or idx >= len(row):
                    return ""
                return _str(row[idx])

            viloyat = cell("viloyat")
            tuman = cell("tuman")
            moljal = cell("moljal")
            izoh = cell("izoh")
            ism01 = cell("ism01")
            familiya = cell("familiya")

            latin_name = _build_latin_name(ism01, familiya, izoh)
            cyrillic = cell("original_1c") if use_cyrillic_name else ""
            primary_name = cyrillic or latin_name
            company_name = latin_name if cyrillic else ""
            location = _build_location(viloyat, tuman, moljal)

            phones: List[str] = []
            for ph_key in ("raqam01", "raqam02", "raqam03"):
                ph = _normalize_phone(cell(ph_key))
                if ph and ph not in phones:
                    phones.append(ph)

            if not phones and not primary_name:
                s_skipped += 1
                continue

            source = f"client_master:{sheet_name.lower()}"

            if phones:
                for phone in phones:
                    result = _upsert(
                        phone=phone,
                        name=primary_name,
                        company_name=company_name,
                        location=location,
                        source=source,
                    )
                    if result == "inserted":
                        s_inserted += 1
                    elif result == "updated":
                        s_updated += 1
            else:
                result = _upsert(
                    phone="",
                    name=primary_name,
                    company_name=company_name,
                    location=location,
                    source=source,
                )
                if result == "phoneless_inserted":
                    s_phoneless_inserted += 1
                elif result == "updated":
                    s_updated += 1

        totals["inserted"] += s_inserted
        totals["updated"] += s_updated
        totals["skipped_empty"] += s_skipped
        totals["phoneless_inserted"] += s_phoneless_inserted
        return {
            "sheet": sheet_name,
            "rows_seen": rows_seen,
            "inserted_with_phone": s_inserted,
            "phoneless_inserted": s_phoneless_inserted,
            "updated": s_updated,
            "skipped_empty": s_skipped,
            "columns_recognized": sorted(col.keys()),
        }

    sheets_summary.append(_process_sheet(CONTACTS_SHEET, CONTACT_COLS, use_cyrillic_name=True))
    sheets_summary.append(_process_sheet(USTO_SHEET, USTO_COLS, use_cyrillic_name=False))

    conn.commit()

    # Retroactively approve registered users by phone.
    approved = 0
    rows = conn.execute(
        "SELECT telegram_id, phone FROM users "
        "WHERE phone IS NOT NULL AND (is_approved = 0 OR is_approved IS NULL)"
    ).fetchall()
    for tg_id, phone in rows:
        ph = _normalize_phone(phone)
        if not ph:
            continue
        match = conn.execute(
            "SELECT id FROM allowed_clients WHERE phone_normalized = ? LIMIT 1",
            (ph,),
        ).fetchone()
        if match:
            conn.execute(
                "UPDATE users SET is_approved = 1, client_id = ? WHERE telegram_id = ?",
                (match[0], tg_id),
            )
            approved += 1

    # Step 8 — mutator chokepoint. Master imports add/update allowed_clients
    # rows that may unblock orphan finance rows. Heal in same transaction.
    from backend.services import client_identity
    orphans_healed = client_identity.heal_all_finance_tables(conn)

    conn.commit()

    total_allowed = conn.execute(
        "SELECT COUNT(*) FROM allowed_clients"
    ).fetchone()[0]
    total_with_1c_name = conn.execute(
        "SELECT COUNT(DISTINCT name) FROM allowed_clients "
        "WHERE name IS NOT NULL AND name <> ''"
    ).fetchone()[0]

    conn.close()

    return {
        "ok": True,
        "filename": filename_hint,
        "totals": totals,
        "sheets": sheets_summary,
        "users_retroactively_approved": approved,
        "db_total_allowed_clients": total_allowed,
        "db_distinct_client_names": total_with_1c_name,
        "orphans_healed": orphans_healed,
    }
