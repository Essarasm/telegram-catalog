"""Import products from the Catalog Clean sheet into SQLite.

Reads the 'Catalog Clean' sheet from the FINAL xlsx file.
Columns: A=Kategoriya, B=Ishlab chiqaruvchi, C=Mahsulot nomi,
         D=Og'irligi, E=Birlik, F=Narx UZS, G=Narx USD

- name field: ORIGINAL Cyrillic (preserved for future Russian language support)
- name_display field: clean Latin transliteration for Uzbek display
"""
import sys
import os
import re
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

from openpyxl import load_workbook
from backend.database import get_db, init_db


# ── Cyrillic → Latin transliteration ────────────────────────────
# Applied AFTER converting to Title Case so multi-char mappings
# (Я→Ya, Ш→Sh, etc.) get natural casing automatically.
CYRILLIC_MAP = {
    'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E',
    'Ё': 'Yo', 'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K',
    'Л': 'L', 'М': 'M', 'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R',
    'С': 'S', 'Т': 'T', 'У': 'U', 'Ф': 'F', 'Х': 'Kh', 'Ц': 'Ts',
    'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch', 'Ъ': '', 'Ы': 'Y', 'Ь': '',
    'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya',
    'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e',
    'ё': 'yo', 'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k',
    'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r',
    'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'kh', 'ц': 'ts',
    'ч': 'ch', 'ш': 'sh', 'щ': 'shch', 'ъ': '', 'ы': 'y', 'ь': '',
    'э': 'e', 'ю': 'yu', 'я': 'ya',
}


def transliterate(text):
    """Convert Cyrillic characters to Latin equivalents."""
    if not text:
        return text
    result = []
    for ch in text:
        result.append(CYRILLIC_MAP.get(ch, ch))
    return ''.join(result)


def cyrillic_title_case(text):
    """Convert to Title Case, handling mixed Cyrillic/Latin.
    'ЗОЛОТИСТАЯ БЕЛАЯ' → 'Золотистая Белая'
    Preserves Latin segments and numbers as-is."""
    words = text.split()
    result = []
    for word in words:
        # If word is all-Latin or a number/code, keep as-is
        if not any('\u0400' <= c <= '\u04ff' for c in word):
            result.append(word)
        # If word is a short abbreviation (2-3 Cyrillic chars), keep uppercase
        elif len(word) <= 3 and word.isupper():
            result.append(word)
        else:
            result.append(word[0].upper() + word[1:].lower() if len(word) > 1 else word.upper())
    return ' '.join(result)


# ── Regex patterns for stripping weight/volume/size from names ──
WEIGHT_VOLUME_RE = re.compile(
    r'/\s*\d+[,.]?\d*\s*'          # opening slash + number
    r'(?:кг|kg|гр|gr|г|мл|ml|л|l)' # unit
    r'\.?\s*/',                     # optional dot + closing slash
    re.IGNORECASE
)

SIZE_RE = re.compile(
    r'/\s*\d+[,.]?\d*\s*'   # /6
    r'(?:см|cm|мм|mm)'      # см
    r'\.?\s*/',              # /
    re.IGNORECASE
)

# Standalone weight at end like "0,75л" without slashes
WEIGHT_END_RE = re.compile(
    r'\s+\d+[,.]?\d*\s*(?:кг|kg|гр|gr|г|мл|ml|л|l)\.?\s*$',
    re.IGNORECASE
)


def strip_weight_volume(name):
    """Remove weight/volume/size info from product name (e.g., /20кг/, /400мл./, /6см/)."""
    name = WEIGHT_VOLUME_RE.sub(' ', name)
    name = SIZE_RE.sub(' ', name)
    name = WEIGHT_END_RE.sub('', name)
    return name.strip()


def strip_producer(name, producer_cyrillic):
    """Remove producer name from the beginning of product name (case-insensitive)."""
    if not producer_cyrillic:
        return name
    # Try the Cyrillic producer name
    prod_upper = producer_cyrillic.strip().upper()
    name_upper = name.upper()
    if name_upper.startswith(prod_upper):
        name = name[len(prod_upper):].strip()
        name = re.sub(r'^[\s\-\u2013\u2014/\\:,.]+', '', name)
    return name


# Common brand name variants in Cyrillic that may differ from the producer field
BRAND_PREFIXES = [
    'ПОЛИСАН', 'СОБСАН', 'СИЛКОАТ', 'ПАЛИЖ', 'ХАЯТ', 'ВЕБЕР',
    'НЮМИКС', 'СОУДАЛ', 'ОСКАР', 'ГАММА', 'ДЕЛЮКС', 'ДЕ ЛЮКС',
    'АКФИКС', 'СОМО ФИКС', 'МАТТРОС', 'ДЕКОАРТ', 'ДАЙСОН',
    'МЕГАМИКС', 'ГУГЛЕ', 'ТИТАН', 'ЛЕОН', 'СЕМИКС', 'ДЕВИЛЮКС',
    'БУМЕРАНГ', 'ФОРМУЛА', 'НОВАМИКС', 'СВЕТОМИКС', 'ВЕРОРАКС',
    'ДУАФИКС', 'ИДЕАЛ', 'СЕНТА', 'ЭКОС', 'КОЛОРЕКС', 'ПОЛИМАКС',
    'СОБ', 'ЗИП',  # Partial brand abbreviations (Sobsan, Zip color)
]


def strip_brand_prefix(name):
    """Remove known brand prefixes that duplicate the producer."""
    name_upper = name.upper()
    for prefix in BRAND_PREFIXES:
        if name_upper.startswith(prefix):
            name = name[len(prefix):].strip()
            name = re.sub(r'^[\s\-\u2013\u2014/\\:,.]+', '', name)
            break
    return name


# Category-specific keywords that are redundant
# (the user already navigated to that category)
CATEGORY_STRIP_WORDS = {
    'ПЛИНТУС', 'IDEAL',  # Plintus va Aksessuarlar
}


def strip_category_keywords(name):
    """Remove category-contextual words that are redundant in navigation."""
    for kw in CATEGORY_STRIP_WORDS:
        # Replace the keyword wherever it appears (case-insensitive)
        pattern = re.compile(re.escape(kw), re.IGNORECASE)
        name = pattern.sub('', name)
    # Clean up resulting double spaces and leading punctuation
    name = re.sub(r'\s+', ' ', name).strip()
    name = re.sub(r'^[\s\-\u2013\u2014/\\:,.]+', '', name)
    return name


def generate_display_name(raw_name, producer_cyrillic):
    """Create a clean display name for mobile UI.

    Pipeline:
    1. Strip producer name
    2. Strip known brand prefixes
    3. Strip weight/volume/size (already shown separately)
    4. Strip redundant category keywords
    5. Normalize Cyrillic to Title Case
    6. Transliterate to Latin
    7. Clean up quotes, extra spaces
    8. Truncate to 40 chars if needed
    """
    name = raw_name.strip()

    # 1-2. Strip producer and brand
    name = strip_producer(name, producer_cyrillic)
    name = strip_brand_prefix(name)

    # 3. Strip weight/volume/size info
    name = strip_weight_volume(name)

    # 4. Strip category keywords
    name = strip_category_keywords(name)

    # 5. Remove excessive quotes and clean
    name = name.replace('"', '').replace("'", '')
    name = re.sub(r'\s+', ' ', name).strip()
    name = re.sub(r'^[\s\-\u2013\u2014/\\:,.]+', '', name)
    name = re.sub(r'[\s\-\u2013\u2014/\\:,.]+$', '', name)

    # 6. Normalize to Title Case (before transliteration for clean multi-char maps)
    name = cyrillic_title_case(name)

    # 7. Transliterate to Latin
    name = transliterate(name)

    # 8. Final cleanup
    name = re.sub(r'\s+', ' ', name).strip()

    # 9. Truncate if too long
    if len(name) > 40:
        cut = name[:37].rfind(' ')
        if cut > 15:
            name = name[:cut].rstrip('.,- ') + '...'
        else:
            name = name[:37].rstrip('.,- ') + '...'

    return name if name else transliterate(raw_name[:30])


def standardize_cyrillic_name(raw_name, producer_cyrillic):
    """Create a clean Russian name for future bilingual support.

    Same stripping as display name but WITHOUT transliteration.
    """
    name = raw_name.strip()
    name = strip_producer(name, producer_cyrillic)
    name = strip_brand_prefix(name)
    name = strip_weight_volume(name)
    name = strip_category_keywords(name)
    name = name.replace('"', '').replace("'", '')
    name = re.sub(r'\s+', ' ', name).strip()
    name = re.sub(r'^[\s\-\u2013\u2014/\\:,.]+', '', name)
    name = re.sub(r'[\s\-\u2013\u2014/\\:,.]+$', '', name)
    name = cyrillic_title_case(name)
    return name if name else raw_name[:40]


def import_from_catalog_clean(xlsx_path: str):
    """Import all products from the Catalog Clean sheet.

    Skips import if products already exist to keep IDs stable
    (cart persistence depends on stable product IDs).
    Use FORCE_REIMPORT=1 env var to force a fresh import.
    """
    init_db()
    conn = get_db()

    # Check if products already exist — skip import to keep IDs stable
    existing = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    force = os.getenv("FORCE_REIMPORT", "").strip()
    if existing > 0 and force != "1":
        print(f"Database already has {existing} products — skipping import (IDs stable).")
        print("Set FORCE_REIMPORT=1 to force a fresh import.")
        conn.close()
        return

    print(f"Loading {xlsx_path}...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if 'Catalog Clean' not in wb.sheetnames:
        print("ERROR: 'Catalog Clean' sheet not found!")
        return
    ws = wb['Catalog Clean']

    # Clear existing data
    conn.execute("DELETE FROM products")
    conn.execute("DELETE FROM producers")
    conn.execute("DELETE FROM categories")
    # Reset auto-increment so IDs start from 1 on fresh import
    conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('products','producers','categories')")

    cat_map = {}   # name -> id
    prod_map = {}  # name -> id
    imported = 0
    skipped = 0

    rows = list(ws.iter_rows(min_row=2, values_only=False))
    print(f"Processing {len(rows)} rows...")

    for row in rows:
        category = row[0].value   # A
        producer = row[1].value   # B
        name = row[2].value       # C
        weight_val = row[3].value if len(row) > 3 else None  # D
        unit = row[4].value if len(row) > 4 else 'sht'       # E
        price_uzs = row[5].value if len(row) > 5 else None    # F
        price_usd = row[6].value if len(row) > 6 else None    # G

        if not category or not producer or not name:
            skipped += 1
            continue

        category = str(category).strip()
        producer = str(producer).strip()
        name = str(name).strip()
        unit = str(unit).strip() if unit else 'sht'

        # Transliterate producer name to Latin for display
        producer_latin = transliterate(cyrillic_title_case(producer))

        # Ensure category exists (categories are already in Latin/Uzbek)
        if category not in cat_map:
            conn.execute(
                "INSERT OR IGNORE INTO categories (name, sort_order) VALUES (?, ?)",
                (category, len(cat_map) + 1)
            )
            cat_id = conn.execute(
                "SELECT id FROM categories WHERE name = ?", (category,)
            ).fetchone()[0]
            cat_map[category] = cat_id

        # Ensure producer exists - use transliterated name
        if producer_latin not in prod_map:
            conn.execute(
                "INSERT OR IGNORE INTO producers (name) VALUES (?)",
                (producer_latin,)
            )
            prod_id = conn.execute(
                "SELECT id FROM producers WHERE name = ?", (producer_latin,)
            ).fetchone()[0]
            prod_map[producer_latin] = prod_id

        # Parse prices
        p_usd = 0
        p_uzs = 0
        try:
            if price_usd is not None and price_usd != '' and price_usd != 0:
                p_usd = float(price_usd)
        except (ValueError, TypeError):
            pass
        try:
            if price_uzs is not None and price_uzs != '' and price_uzs != 0:
                p_uzs = float(price_uzs)
        except (ValueError, TypeError):
            pass

        # Parse weight
        weight = None
        try:
            if weight_val is not None and weight_val != '':
                weight = float(weight_val)
        except (ValueError, TypeError):
            pass

        # name = ORIGINAL Cyrillic (for Russian bilingual support)
        original_cyrillic = name

        # name_display = clean Latin for Uzbek display
        display_name = generate_display_name(name, producer)

        conn.execute(
            """INSERT INTO products
               (name, name_display, category_id, producer_id, unit,
                price_usd, price_uzs, weight, is_active)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)""",
            (original_cyrillic, display_name, cat_map[category], prod_map[producer_latin],
             unit, p_usd, p_uzs, weight)
        )
        imported += 1

    # Update denormalized counts
    conn.execute("""
        UPDATE categories SET product_count = (
            SELECT COUNT(*) FROM products WHERE products.category_id = categories.id AND is_active = 1
        )
    """)
    conn.execute("""
        UPDATE producers SET product_count = (
            SELECT COUNT(*) FROM products WHERE products.producer_id = producers.id AND is_active = 1
        )
    """)

    conn.commit()

    # Summary
    total_cats = conn.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
    total_prods = conn.execute("SELECT COUNT(*) FROM producers").fetchone()[0]
    total_products = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
    usd_count = conn.execute("SELECT COUNT(*) FROM products WHERE price_usd > 0").fetchone()[0]
    uzs_count = conn.execute("SELECT COUNT(*) FROM products WHERE price_uzs > 0").fetchone()[0]

    print(f"\nImport complete:")
    print(f"  Categories: {total_cats}")
    print(f"  Producers: {total_prods}")
    print(f"  Products: {total_products} (USD: {usd_count}, UZS: {uzs_count})")
    print(f"  Skipped: {skipped}")

    # Category breakdown
    rows = conn.execute("""
        SELECT c.name, c.product_count
        FROM categories c ORDER BY c.product_count DESC
    """).fetchall()
    print("\n  Category breakdown:")
    for r in rows:
        print(f"    {r['product_count']:>4}  {r['name']}")

    # Top producers
    rows = conn.execute("""
        SELECT name, product_count FROM producers ORDER BY product_count DESC LIMIT 10
    """).fetchall()
    print("\n  Top 10 producers:")
    for r in rows:
        print(f"    {r['product_count']:>4}  {r['name']}")

    # Show sample display names from various categories
    rows = conn.execute("""
        SELECT p.name, p.name_display, pr.name as producer_name, c.name as cat_name
        FROM products p
        JOIN producers pr ON p.producer_id = pr.id
        JOIN categories c ON p.category_id = c.id
        ORDER BY RANDOM()
        LIMIT 25
    """).fetchall()
    print("\n  Sample display names (random 25):")
    for r in rows:
        print(f"    [{r['cat_name'][:15]}] [{r['producer_name']}] {r['name'][:45]}")
        print(f"      -> {r['name_display']}")

    conn.close()
    wb.close()


if __name__ == "__main__":
    # Check multiple possible locations
    candidates = [
        os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'products.xlsx'),
        os.path.join(os.path.dirname(__file__), '..', '..', '..', '09.03.26 List of products (Inventory) - FINAL.xlsx'),
        '/sessions/clever-vibrant-hamilton/mnt/Catalogue:Telegram app/09.03.26 List of products (Inventory) - FINAL.xlsx',
        './data/products.xlsx',
    ]
    xlsx_path = None
    for c in candidates:
        if os.path.exists(c):
            xlsx_path = os.path.abspath(c)
            break
    if xlsx_path:
        import_from_catalog_clean(xlsx_path)
    else:
        print("ERROR: No xlsx file found. Checked:", [os.path.abspath(c) for c in candidates])
        sys.exit(1)
