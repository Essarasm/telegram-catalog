"""Classify products into lifecycle buckets based on supply history.

Reads `Inventory/Product_Lifecycle_2026-04-17.xlsx` (Session F's analysis) to learn
which products are Active / Aging / Stale / Never-supplied, then:
  1. Sets `products.lifecycle` for each matched product
  2. Applies the 7-item Стеллаж hit-list from `1C_Correction_Report.xlsx` (is_active=0)
  3. Generates `Catalog_Cleanup_<date>.xlsx` review file

Catalog browse (bot) shows only lifecycle IN ('active','aging') = ~1,490 products.
Dashboard (Session X) can filter lifecycle='active' = ~869 products.
Stale + never stay is_active=1 so they're searchable via fuzzy match (click-tracked).

Re-runnable: safe to call after 1C updates.
"""
import os
import sys
import sqlite3
import unicodedata
import re
import openpyxl
from datetime import date


PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REPO_ROOT = os.path.dirname(PROJECT_ROOT)


def _find_xlsx(name, fallback_name=None):
    """Try bundled (data/), then parent Inventory/, return first that exists."""
    candidates = [
        os.path.join(PROJECT_ROOT, 'data', name),
        os.path.join(REPO_ROOT, 'Inventory', name),
    ]
    if fallback_name:
        candidates += [
            os.path.join(PROJECT_ROOT, 'data', fallback_name),
            os.path.join(REPO_ROOT, 'Inventory', fallback_name),
        ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return candidates[0]  # return first (will fail with clear error if missing)


LIFECYCLE_XLSX = _find_xlsx('Product_Lifecycle.xlsx', 'Product_Lifecycle_2026-04-17.xlsx')
CORRECTION_XLSX = _find_xlsx('1C_Correction_Report.xlsx')
DB_PATH = os.environ.get('DATABASE_PATH', os.path.join(PROJECT_ROOT, 'data', 'catalog.db'))
OUT_XLSX = os.path.join(PROJECT_ROOT, f'Catalog_Cleanup_{date.today().isoformat()}.xlsx')

SHEET_LIFECYCLE_MAP = {
    'Active (869)': 'active',
    'Aging (621)': 'aging',
    'Stale (248)': 'stale',
    'Never supplied (815)': 'never',
}


def canon(s: str) -> str:
    if not isinstance(s, str):
        return ''
    s = unicodedata.normalize('NFC', s).upper().strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def load_lifecycle_classifications(xlsx_path):
    """Returns {canonical_name: lifecycle} dict."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out = {}
    for sheet_name, lifecycle in SHEET_LIFECYCLE_MAP.items():
        ws = wb[sheet_name]
        # Row 1 = header; name in col B
        for row in ws.iter_rows(min_row=2, values_only=True):
            name_raw = row[1] if len(row) > 1 else None
            if not isinstance(name_raw, str):
                continue
            out[canon(name_raw)] = lifecycle
    wb.close()
    return out


def load_hitlist(xlsx_path):
    """Returns list of canonical names to deactivate (is_active=0)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["4. O'chirish kerak"]
    names = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        # col B (index 1) = '1C dagi nomi'
        if len(row) > 1 and isinstance(row[1], str) and row[1].strip():
            names.append(canon(row[1]))
    wb.close()
    return names


def main(dry_run=False):
    print(f'Reading lifecycle classifications from {LIFECYCLE_XLSX}...')
    lifecycle_map = load_lifecycle_classifications(LIFECYCLE_XLSX)
    print(f'  {len(lifecycle_map)} classified products in xlsx')

    print(f'Reading hit-list from {CORRECTION_XLSX}...')
    hitlist = load_hitlist(CORRECTION_XLSX)
    print(f'  {len(hitlist)} products to hard-deactivate')

    print(f'Opening DB {DB_PATH}...')
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # Before snapshot
    before = list(conn.execute(
        "SELECT id, name, name_display, is_active, lifecycle FROM products"
    ).fetchall())
    print(f'  {len(before)} products in DB')

    # Build canonical-name -> db product map
    db_by_canon = {}
    for p in before:
        db_by_canon.setdefault(canon(p['name']), []).append(p)

    # Classify
    changes = []  # list of (id, name, old_lifecycle, new_lifecycle, old_is_active, new_is_active, reason)
    matched_classifications = 0
    unmatched_xlsx = []

    for canon_name, new_lifecycle in lifecycle_map.items():
        db_products = db_by_canon.get(canon_name, [])
        if not db_products:
            unmatched_xlsx.append(canon_name)
            continue
        for p in db_products:
            if p['lifecycle'] != new_lifecycle:
                changes.append((p['id'], p['name'], p['lifecycle'], new_lifecycle,
                                p['is_active'], p['is_active'], 'lifecycle_updated'))
            matched_classifications += 1

    # Apply hit-list (is_active=0, regardless of lifecycle)
    hitlist_matched = 0
    for canon_name in hitlist:
        db_products = db_by_canon.get(canon_name, [])
        for p in db_products:
            if p['is_active']:
                changes.append((p['id'], p['name'], p['lifecycle'], p['lifecycle'],
                                1, 0, 'hitlist_стеллаж'))
                hitlist_matched += 1

    # Unclassified products (in DB but not in xlsx): default to 'never' as a safe signal
    db_canon_set = set(db_by_canon.keys())
    xlsx_canon_set = set(lifecycle_map.keys())
    unclassified_in_db = db_canon_set - xlsx_canon_set
    for canon_name in unclassified_in_db:
        for p in db_by_canon[canon_name]:
            # If already lifecycle is set (non-default), skip
            if p['lifecycle'] != 'active':
                continue
            changes.append((p['id'], p['name'], p['lifecycle'], 'never',
                            p['is_active'], p['is_active'], 'not_in_xlsx'))

    print(f'\n=== CLASSIFICATION SUMMARY ===')
    print(f'  Classifications from xlsx: {len(lifecycle_map)}')
    print(f'  Matched to DB products: {matched_classifications}')
    print(f'  Unmatched in xlsx (not in DB): {len(unmatched_xlsx)}')
    print(f'  Unclassified in DB (fallback=never): {len(unclassified_in_db)}')
    print(f'  Hit-list matches (will is_active=0): {hitlist_matched}')
    print(f'  Total changes: {len(changes)}')

    # Preview distribution
    preview_counts = {'active': 0, 'aging': 0, 'stale': 0, 'never': 0}
    for p in before:
        canon_name = canon(p['name'])
        target = lifecycle_map.get(canon_name, 'never' if canon_name in unclassified_in_db else p['lifecycle'])
        preview_counts[target] = preview_counts.get(target, 0) + 1
    print(f'\n=== POST-RUN DISTRIBUTION (predicted) ===')
    for lc, n in preview_counts.items():
        print(f'  {lc}: {n}')
    print(f'  Catalog visible (active+aging): {preview_counts["active"] + preview_counts["aging"]}')

    if dry_run:
        print('\n[DRY RUN] No DB changes written.')
    else:
        print('\nApplying changes to DB...')
        for pid, name, old_lc, new_lc, old_act, new_act, reason in changes:
            if old_lc != new_lc:
                conn.execute("UPDATE products SET lifecycle = ? WHERE id = ?", (new_lc, pid))
            if old_act != new_act:
                conn.execute("UPDATE products SET is_active = ? WHERE id = ?", (new_act, pid))
        conn.commit()
        print(f'  Committed {len(changes)} changes.')

    # Generate review xlsx
    print(f'\nWriting review xlsx to {OUT_XLSX}...')
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

    def make_sheet(name, headers, rows):
        ws = out_wb.create_sheet(name)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
        for r, data_row in enumerate(rows, 2):
            for c, v in enumerate(data_row, 1):
                ws.cell(row=r, column=c, value=v)
        for c, w in enumerate([7, 45, 15, 15, 12, 12, 20], 1):
            ws.column_dimensions[chr(64 + c)].width = w
        return ws

    # Sheet 1: Summary
    ws = out_wb.create_sheet('Summary')
    ws.append(['Catalog Cleanup Report', ''])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append(['Generated', date.today().isoformat()])
    ws.append([])
    ws.append(['Total products in DB', len(before)])
    ws.append(['Lifecycle classifications from xlsx', len(lifecycle_map)])
    ws.append(['Matched to DB', matched_classifications])
    ws.append(['Unmatched in xlsx (not in DB)', len(unmatched_xlsx)])
    ws.append(['Unclassified in DB (→ never)', len(unclassified_in_db)])
    ws.append(['Hit-list hard-deactivations', hitlist_matched])
    ws.append(['Total DB changes', len(changes)])
    ws.append([])
    ws.append(['Final lifecycle distribution', ''])
    for lc, n in preview_counts.items():
        ws.append([f'  {lc}', n])
    ws.append(['  Catalog visible (active + aging)', preview_counts['active'] + preview_counts['aging']])
    ws.append(['  Dashboard focus (active only)', preview_counts['active']])
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    # Sheet 2: Changes
    rows = [(pid, name, old_lc, new_lc, old_act, new_act, reason)
            for pid, name, old_lc, new_lc, old_act, new_act, reason in changes]
    make_sheet('Changes', ['id', 'name (1C)', 'old lifecycle', 'new lifecycle',
                           'old is_active', 'new is_active', 'reason'], rows)

    # Sheet 3: Will-be-hidden (lifecycle = stale or never, is_active stays 1)
    hidden_rows = []
    for p in before:
        canon_name = canon(p['name'])
        target = lifecycle_map.get(canon_name, 'never' if canon_name in unclassified_in_db else p['lifecycle'])
        if target in ('stale', 'never') and p['is_active']:
            hidden_rows.append((p['id'], p['name'], target, '', 1, 1, 'hidden_searchable'))
    make_sheet(f'Hidden (searchable, {len(hidden_rows)})',
               ['id', 'name (1C)', 'lifecycle', '', 'is_active (stays 1)', '', 'note'], hidden_rows)

    # Sheet 4: Hard-deactivated (hitlist)
    ds_rows = [(pid, name, old_lc, new_lc, old_act, new_act, reason)
               for pid, name, old_lc, new_lc, old_act, new_act, reason in changes
               if reason == 'hitlist_стеллаж']
    make_sheet(f'Hard-deactivated ({len(ds_rows)})',
               ['id', 'name (1C)', 'lifecycle', '', 'old is_active', 'new is_active', 'reason'], ds_rows)

    # Sheet 5: Unmatched xlsx rows (products in xlsx not found in DB)
    make_sheet(f'Unmatched xlsx ({len(unmatched_xlsx)})',
               ['canonical name'], [(n,) for n in sorted(unmatched_xlsx)[:1000]])

    out_wb.save(OUT_XLSX)
    print(f'  Saved.')

    conn.close()
    return {
        'changes': len(changes),
        'distribution': preview_counts,
        'catalog_visible': preview_counts['active'] + preview_counts['aging'],
        'dashboard_focus': preview_counts['active'],
    }


if __name__ == '__main__':
    dry_run = '--dry-run' in sys.argv
    startup_mode = '--startup' in sys.argv  # silent skip on missing xlsx; no review xlsx
    try:
        if startup_mode and not (os.path.exists(LIFECYCLE_XLSX) and os.path.exists(CORRECTION_XLSX)):
            print(f'[classify_lifecycle] startup mode: xlsx not found, skipping.')
            sys.exit(0)
        result = main(dry_run=dry_run)
        print(f'\nResult: {result}')
    except Exception as e:
        if startup_mode:
            print(f'[classify_lifecycle] ERROR in startup mode, continuing boot: {e}')
            sys.exit(0)
        raise
