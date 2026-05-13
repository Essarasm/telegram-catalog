"""Fill `allowed_clients.viloyat/tuman/moljal` from Client Master xlsx Contacts sheet.

Fill-only — per `feedback_master_fill_only.md`, never overwrites a non-empty value
already present in the DB. Joins by `Original Ism (1C)` → `allowed_clients.client_id_1c`.

Local run:
    python3 telegram-catalog/tools/sync_master_tuman_to_db.py \\
        --xlsx "../Client Master 13.05.26.xlsx" \\
        --db   "telegram-catalog/data/snapshots/prod-2026-05-13.db"

Prod (Railway) — production volume DB at /data/catalog.db:
    railway ssh "python tools/sync_master_tuman_to_db.py --xlsx /tmp/cm.xlsx --db /data/catalog.db"
    (after uploading the xlsx to /tmp/ — separate step)
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True, help="Path to Client Master xlsx")
    p.add_argument("--db",   required=True, help="Path to catalog.db")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    db_path = Path(args.db).expanduser().resolve()
    if not xlsx_path.exists():
        sys.exit(f"xlsx not found: {xlsx_path}")
    if not db_path.exists():
        sys.exit(f"db not found: {db_path}")

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Contacts" not in wb.sheetnames:
        sys.exit("No Contacts sheet in xlsx")
    ws = wb["Contacts"]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows))
    idx = {h: i for i, h in enumerate(hdr)}
    needed = {"Original Ism (1C)", "Viloyat", "Shahar/Tuman", "Mo'ljal"}
    missing = needed - set(idx.keys())
    if missing:
        sys.exit(f"Contacts sheet missing columns: {missing}")

    contacts: dict[str, tuple[str, str, str]] = {}
    for row in rows:
        ism = row[idx["Original Ism (1C)"]]
        if not isinstance(ism, str) or not ism.strip():
            continue
        vil = row[idx["Viloyat"]]
        tum = row[idx["Shahar/Tuman"]]
        mol = row[idx["Mo'ljal"]]
        contacts[ism.strip()] = (
            (vil or "").strip() if isinstance(vil, str) else "",
            (tum or "").strip() if isinstance(tum, str) else "",
            (mol or "").strip() if isinstance(mol, str) else "",
        )
    wb.close()
    print(f"[contacts] {len(contacts)} rows with Original Ism (1C)")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    ac_rows = conn.execute(
        """
        SELECT id, client_id_1c, COALESCE(viloyat,'') AS viloyat,
               COALESCE(tuman,'') AS tuman, COALESCE(moljal,'') AS moljal
        FROM allowed_clients
        WHERE COALESCE(status,'active')='active' AND client_id_1c IS NOT NULL
        """
    ).fetchall()
    print(f"[allowed_clients] {len(ac_rows)} active rows with client_id_1c")

    filled = {"viloyat": 0, "tuman": 0, "moljal": 0}
    rows_touched = 0
    rows_skipped_already_filled = 0
    rows_no_master = 0

    cursor = conn.cursor()
    for r in ac_rows:
        master = contacts.get(r["client_id_1c"])
        if not master:
            rows_no_master += 1
            continue
        mv, mt, mm = master
        updates: list[str] = []
        params: list = []
        if not r["viloyat"] and mv:
            updates.append("viloyat = ?"); params.append(mv); filled["viloyat"] += 1
        if not r["tuman"] and mt:
            updates.append("tuman = ?"); params.append(mt); filled["tuman"] += 1
        if not r["moljal"] and mm:
            updates.append("moljal = ?"); params.append(mm); filled["moljal"] += 1
        if not updates:
            rows_skipped_already_filled += 1
            continue
        params.append(r["id"])
        if not args.dry_run:
            cursor.execute(f"UPDATE allowed_clients SET {', '.join(updates)} WHERE id = ?", params)
        rows_touched += 1

    if not args.dry_run:
        conn.commit()
    conn.close()

    print(f"[result] rows touched: {rows_touched}")
    print(f"         fields filled: viloyat={filled['viloyat']} tuman={filled['tuman']} moljal={filled['moljal']}")
    print(f"         rows skipped — already filled: {rows_skipped_already_filled}")
    print(f"         rows skipped — no master row: {rows_no_master}")
    if args.dry_run:
        print("         (DRY RUN — no changes written)")


if __name__ == "__main__":
    main()
