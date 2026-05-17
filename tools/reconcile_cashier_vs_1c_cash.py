"""One-off reconciliation: cashier intake (Muqaddas's Telegram input) vs Alisher's 1C cash upload.

Outputs an XLSX with per-client aggregates plus the underlying rows so each mismatch can be
walked through manually with Alisher to confirm whether the comparison surfaces real typos.

Usage:
    python tools/reconcile_cashier_vs_1c_cash.py \
        --db data/snapshots/prod-2026-05-15.db \
        --start 2026-04-30 --end 2026-05-14 \
        --out Cashier_vs_1C_Reconciliation_2026-05-15.xlsx
"""

import argparse
import re
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ILLEGAL_XLSX_CHARS = re.compile(r"[\000-\010\013\014\016-\037]")


def clean(v):
    if isinstance(v, str):
        return ILLEGAL_XLSX_CHARS.sub("", v)
    return v


def append_row(ws, row):
    ws.append([clean(v) for v in row])

FX_FALLBACK = 12000  # project fallback per memory
RATIO_LO, RATIO_HI = 10000, 14000  # currency-swap detection band


def fetch_aggregates(conn, start, end):
    """Per-client cashier-cash + 1C-cash totals over the window."""
    cur = conn.cursor()
    cur.execute(
        """
        WITH cashier AS (
          SELECT client_id,
            SUM(CASE WHEN currency='UZS' THEN amount ELSE 0 END) AS uzs,
            SUM(CASE WHEN currency='USD' THEN amount ELSE 0 END) AS usd,
            COUNT(*) AS n
          FROM intake_payments
          WHERE status='confirmed'
            AND channel IN ('cash_direct','cash_via_agent')
            AND DATE(submitted_at) BETWEEN ? AND ?
          GROUP BY client_id
        ),
        onec AS (
          SELECT client_id,
            SUM(CASE WHEN currency='UZS' THEN amount_local ELSE 0 END) AS uzs,
            SUM(CASE WHEN currency='USD' THEN amount_currency ELSE 0 END) AS usd,
            COUNT(*) AS n
          FROM client_payments
          WHERE doc_date BETWEEN ? AND ? AND client_id IS NOT NULL
          GROUP BY client_id
        )
        SELECT
          ac.id, ac.client_id_1c,
          COALESCE(c.uzs, 0), COALESCE(c.usd, 0), COALESCE(c.n, 0),
          COALESCE(o.uzs, 0), COALESCE(o.usd, 0), COALESCE(o.n, 0)
        FROM allowed_clients ac
        LEFT JOIN cashier c ON ac.id = c.client_id
        LEFT JOIN onec o ON ac.id = o.client_id
        WHERE c.client_id IS NOT NULL OR o.client_id IS NOT NULL
        """,
        (start, end, start, end),
    )
    return cur.fetchall()


def fetch_cashier_rows(conn, start, end):
    """All confirmed cash intake entries in window, ordered for grouping."""
    cur = conn.cursor()
    cur.execute(
        """
        SELECT ip.client_id, ac.client_id_1c, DATE(ip.submitted_at) AS d, ip.submitted_at,
               ip.channel, ip.currency, ip.amount, ip.submitter_telegram_id, ip.notes
        FROM intake_payments ip
        JOIN allowed_clients ac ON ac.id = ip.client_id
        WHERE ip.status='confirmed'
          AND ip.channel IN ('cash_direct','cash_via_agent')
          AND DATE(ip.submitted_at) BETWEEN ? AND ?
        ORDER BY ac.client_id_1c, ip.submitted_at
        """,
        (start, end),
    )
    return cur.fetchall()


def fetch_onec_rows(conn, start, end):
    """All 1C cash entries with resolved client in window, ordered for grouping."""
    cur = conn.cursor()
    cur.execute(
        """
        SELECT cp.client_id, ac.client_id_1c, cp.doc_date, cp.doc_number_1c,
               cp.currency,
               CASE WHEN cp.currency='UZS' THEN cp.amount_local ELSE cp.amount_currency END AS amt,
               cp.amount_local, cp.amount_currency, cp.fx_rate, cp.author,
               cp.received_from, cp.client_name_1c
        FROM client_payments cp
        JOIN allowed_clients ac ON ac.id = cp.client_id
        WHERE cp.doc_date BETWEEN ? AND ?
          AND cp.client_id IS NOT NULL
        ORDER BY ac.client_id_1c, cp.doc_date, cp.doc_number_1c
        """,
        (start, end),
    )
    return cur.fetchall()


def _fmt_amt(amount, currency):
    if currency == "USD":
        return f"${amount:,.2f}"
    return f"{int(round(amount)):,} UZS"


def format_cashier_entry(row):
    """Single cashier row → 'YYYY-MM-DD HH:MM • amt cur • channel'."""
    _, _, _, submitted_at, channel, currency, amount, _, _ = row
    ts = (submitted_at or "")[:16]  # trim seconds
    ch = "agent" if channel == "cash_via_agent" else "direct"
    return f"{ts} • {_fmt_amt(amount, currency)} • {ch}"


def format_onec_entry(row):
    """Single 1C row → 'YYYY-MM-DD • amt cur • doc #N'."""
    _, _, doc_date, doc_num, currency, amt, _, _, _, _, _, _ = row
    return f"{doc_date} • {_fmt_amt(amt, currency)} • doc #{doc_num}"


def group_by_client(rows):
    """rows[0] must be client_id. Returns {client_id: [row, row, ...]}."""
    grouped = {}
    for r in rows:
        grouped.setdefault(r[0], []).append(r)
    return grouped


def classify(k_uzs, k_usd, a_uzs, a_usd, k_n, a_n):
    d_uzs = k_uzs - a_uzs
    d_usd = k_usd - a_usd
    # currency-swap: cashier-up-UZS with cashier-down-USD (or vice versa), ratio ~12,000
    if abs(d_uzs) >= 1_000_000 and abs(d_usd) >= 50 and d_uzs * d_usd < 0:
        ratio = abs(d_uzs / d_usd) if d_usd else 0
        if RATIO_LO <= ratio <= RATIO_HI:
            return "currency_swap", ratio
    if k_n > 0 and a_n == 0:
        return "lag_or_missing_in_1c", None
    if k_n == 0 and a_n > 0:
        return "outside_cashier_scope", None
    if abs(d_uzs) < 1000 and abs(d_usd) < 0.5:
        return "match", None
    return "other_mismatch", None


HDR_FILL = PatternFill("solid", fgColor="305496")
HDR_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
BUCKET_COLOR = {
    "currency_swap": PatternFill("solid", fgColor="F4B084"),
    "lag_or_missing_in_1c": PatternFill("solid", fgColor="FFE699"),
    "outside_cashier_scope": PatternFill("solid", fgColor="C6E0B4"),
    "other_mismatch": PatternFill("solid", fgColor="FFC7CE"),
    "match": None,
}


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def autosize(ws):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", required=True)
    ap.add_argument("--start", required=True)
    ap.add_argument("--end", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    conn = sqlite3.connect(args.db)
    rows = fetch_aggregates(conn, args.start, args.end)

    enriched = []
    for cid, client_1c, k_uzs, k_usd, k_n, a_uzs, a_usd, a_n in rows:
        cls, ratio = classify(k_uzs, k_usd, a_uzs, a_usd, k_n, a_n)
        enriched.append(
            {
                "id": cid,
                "client": client_1c,
                "k_uzs": k_uzs, "k_usd": k_usd, "k_n": k_n,
                "a_uzs": a_uzs, "a_usd": a_usd, "a_n": a_n,
                "d_uzs": k_uzs - a_uzs,
                "d_usd": k_usd - a_usd,
                "abs_usd_eq": abs(k_uzs - a_uzs) + abs(k_usd - a_usd) * FX_FALLBACK,
                "class": cls,
                "ratio": ratio,
            }
        )

    bucket1 = [e for e in enriched if e["k_n"] > 0 and e["a_n"] > 0]
    bucket2 = [e for e in enriched if e["k_n"] > 0 and e["a_n"] == 0]
    bucket3 = [e for e in enriched if e["k_n"] == 0 and e["a_n"] > 0]

    bucket1.sort(key=lambda e: e["abs_usd_eq"], reverse=True)
    bucket2.sort(key=lambda e: e["abs_usd_eq"], reverse=True)
    bucket3.sort(key=lambda e: e["abs_usd_eq"], reverse=True)

    all_cashier_rows = fetch_cashier_rows(conn, args.start, args.end)
    all_onec_rows = fetch_onec_rows(conn, args.start, args.end)
    cashier_by_client = group_by_client(all_cashier_rows)
    onec_by_client = group_by_client(all_onec_rows)

    def cashier_entries_str(cid):
        return "\n".join(format_cashier_entry(r) for r in cashier_by_client.get(cid, []))

    def onec_entries_str(cid):
        return "\n".join(format_onec_entry(r) for r in onec_by_client.get(cid, []))

    mismatch_ids = {e["id"] for e in bucket1 if e["class"] != "match"}
    cashier_detail = [r for r in all_cashier_rows if r[0] in mismatch_ids]
    onec_detail = [r for r in all_onec_rows if r[0] in mismatch_ids]

    wb = Workbook()

    # --- Summary
    ws = wb.active
    ws.title = "Summary"
    append_row(ws, ["Reconciliation window", f"{args.start} → {args.end}"])
    append_row(ws, ["DB snapshot", args.db])
    append_row(ws, ["FX fallback used for ranking", FX_FALLBACK])
    append_row(ws, [])
    append_row(ws, ["Bucket", "Clients", "Cashier UZS", "1C UZS", "Δ UZS", "Cashier USD", "1C USD", "Δ USD"])
    def bucket_row(name, items):
        ku = sum(e["k_uzs"] for e in items)
        au = sum(e["a_uzs"] for e in items)
        kd = sum(e["k_usd"] for e in items)
        ad = sum(e["a_usd"] for e in items)
        append_row(ws, [name, len(items), ku, au, ku - au, kd, ad, kd - ad])
    bucket_row("1 — Both sides recorded", bucket1)
    bucket_row("  └─ classified as 'match'", [e for e in bucket1 if e["class"] == "match"])
    bucket_row("  └─ currency_swap candidates", [e for e in bucket1 if e["class"] == "currency_swap"])
    bucket_row("  └─ other_mismatch", [e for e in bucket1 if e["class"] == "other_mismatch"])
    bucket_row("  └─ lag_or_missing_in_1c (in B1)", [e for e in bucket1 if e["class"] == "lag_or_missing_in_1c"])
    bucket_row("2 — Cashier only", bucket2)
    bucket_row("3 — 1C only (outside cashier scope?)", bucket3)
    append_row(ws, [])
    append_row(ws, ["Legend"])
    append_row(ws, ["currency_swap", "Δ UZS and Δ USD opposite signs, |Δ UZS / Δ USD| ≈ 12,000 → likely Alisher converted UZS to USD when typing into 1C"])
    append_row(ws, ["lag_or_missing_in_1c", "Cashier has entries, 1C has 0 — Alisher hasn't entered yet, or entered under a different client name"])
    append_row(ws, ["outside_cashier_scope", "1C has entries, cashier has 0 — Alisher entered directly (historical, off-channel, or genuine cashier miss)"])
    append_row(ws, ["other_mismatch", "Both sides have data but neither match nor a clean currency_swap — needs case-by-case review"])
    append_row(ws, ["match", "Totals agree within ±1,000 UZS and ±$0.50"])

    # --- Bucket 1 detail
    ws = wb.create_sheet("Bucket 1 — Mismatches")
    headers = [
        "Client (1C)", "Pattern", "Ratio (UZS/USD)",
        "Cashier UZS", "1C UZS", "Δ UZS",
        "Cashier USD", "1C USD", "Δ USD",
        "Cashier #rows", "1C #rows",
        "Cashier entries (date time • amt • channel)",
        "1C entries (date • amt • doc#)",
    ]
    append_row(ws, headers)
    style_header(ws, len(headers))
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for e in bucket1:
        if e["class"] == "match":
            continue
        row = [
            e["client"], e["class"],
            round(e["ratio"], 0) if e["ratio"] else None,
            e["k_uzs"], e["a_uzs"], e["d_uzs"],
            round(e["k_usd"], 2), round(e["a_usd"], 2), round(e["d_usd"], 2),
            e["k_n"], e["a_n"],
            cashier_entries_str(e["id"]),
            onec_entries_str(e["id"]),
        ]
        append_row(ws, row)
        r_idx = ws.max_row
        fill = BUCKET_COLOR.get(e["class"])
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c).fill = fill
        ws.cell(row=r_idx, column=12).alignment = wrap_align
        ws.cell(row=r_idx, column=13).alignment = wrap_align
        max_lines = max(e["k_n"], e["a_n"], 1)
        ws.row_dimensions[r_idx].height = min(max(15 * max_lines + 3, 18), 400)
    autosize(ws)
    ws.column_dimensions["L"].width = 48
    ws.column_dimensions["M"].width = 48

    # --- Cashier detail (mismatched clients only)
    ws = wb.create_sheet("Detail — Cashier rows")
    headers = ["Client (1C)", "Date", "Submitted at", "Channel", "Currency", "Amount", "Submitter TG", "Notes"]
    append_row(ws, headers)
    style_header(ws, len(headers))
    for r in cashier_detail:
        _, client_1c, d, submitted_at, channel, currency, amount, submitter, notes = r
        append_row(ws, [client_1c, d, submitted_at, channel, currency, amount, submitter, notes])
    autosize(ws)

    # --- 1C detail (mismatched clients only)
    ws = wb.create_sheet("Detail — 1C rows")
    headers = ["Client (1C)", "Doc date", "Doc #", "Currency", "Amount", "Amount UZS leg",
               "Amount USD leg", "FX rate", "Author", "Received from (1C)", "Subconto1 (1C)"]
    append_row(ws, headers)
    style_header(ws, len(headers))
    for r in onec_detail:
        _, client_1c, doc_date, doc_num, currency, amt, amt_local, amt_curr, fx, author, recv_from, sub1 = r
        append_row(ws, [client_1c, doc_date, doc_num, currency, amt, amt_local, amt_curr, fx, author, recv_from, sub1])
    autosize(ws)

    # --- Bucket 2 (cashier-only)
    ws = wb.create_sheet("Bucket 2 — Cashier only")
    headers = ["Client (1C)", "Cashier UZS", "Cashier USD", "Cashier #rows",
               "Cashier entries (date time • amt • channel)"]
    append_row(ws, headers)
    style_header(ws, len(headers))
    for e in bucket2:
        append_row(ws, [e["client"], e["k_uzs"], round(e["k_usd"], 2), e["k_n"],
                        cashier_entries_str(e["id"])])
        r_idx = ws.max_row
        ws.cell(row=r_idx, column=5).alignment = wrap_align
        ws.row_dimensions[r_idx].height = min(max(15 * e["k_n"] + 3, 18), 400)
    autosize(ws)
    ws.column_dimensions["E"].width = 48

    # --- Bucket 3 (1C-only)
    ws = wb.create_sheet("Bucket 3 — 1C only")
    headers = ["Client (1C)", "1C UZS", "1C USD", "1C #rows",
               "1C entries (date • amt • doc#)"]
    append_row(ws, headers)
    style_header(ws, len(headers))
    for e in bucket3:
        append_row(ws, [e["client"], e["a_uzs"], round(e["a_usd"], 2), e["a_n"],
                        onec_entries_str(e["id"])])
        r_idx = ws.max_row
        ws.cell(row=r_idx, column=5).alignment = wrap_align
        ws.row_dimensions[r_idx].height = min(max(15 * e["a_n"] + 3, 18), 400)
    autosize(ws)
    ws.column_dimensions["E"].width = 48

    autosize(wb["Summary"])

    out = Path(args.out)
    wb.save(out)
    print(f"Wrote {out} ({out.stat().st_size:,} bytes)")
    print(f"  Bucket 1 (both sides): {len(bucket1)} clients")
    print(f"    currency_swap candidates: {sum(1 for e in bucket1 if e['class']=='currency_swap')}")
    print(f"    other_mismatch:          {sum(1 for e in bucket1 if e['class']=='other_mismatch')}")
    print(f"    lag_or_missing_in_1c:    {sum(1 for e in bucket1 if e['class']=='lag_or_missing_in_1c')}")
    print(f"    match (within tolerance):{sum(1 for e in bucket1 if e['class']=='match')}")
    print(f"  Bucket 2 (cashier only):   {len(bucket2)} clients")
    print(f"  Bucket 3 (1C only):        {len(bucket3)} clients")


if __name__ == "__main__":
    main()
