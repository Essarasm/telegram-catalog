#!/usr/bin/env python3
"""Seed the product_aliases table from multiple data sources.

Sources (in priority order):
1. DB products table — every product's own 1C name is an alias for itself
2. Rassvet_Master Ibrat.xlsx "Katalog" sheet — 1C name → Ibrat ID → match to DB product
3. DB supply_order_items — 10,263 historical line items with 1C names matched to products
4. inventory/products.xlsx "Catalog Clean" sheet — Asl nomi → cleaned Latin name

Run: python3 tools/seed_product_aliases.py [--dry-run]
"""
import os
import sys
import sqlite3
import unicodedata

# Add project root to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

DATABASE_PATH = os.getenv("DATABASE_PATH", "data/catalog.db")
INVENTORY_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "inventory")
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data")

DRY_RUN = "--dry-run" in sys.argv


def normalize(name):
    """Normalize a product name for matching: lowercase, NFC, strip."""
    if not name:
        return ""
    return unicodedata.normalize("NFC", name.strip().lower())


def get_db():
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def seed_from_db_products(conn):
    """Source 1: Every product's 1C name (products.name) is an alias for itself."""
    rows = conn.execute(
        "SELECT id, name FROM products WHERE is_active = 1 AND name IS NOT NULL"
    ).fetchall()
    count = 0
    for r in rows:
        alias = normalize(r["name"])
        if not alias:
            continue
        try:
            conn.execute(
                "INSERT OR IGNORE INTO product_aliases (alias_name, alias_name_lower, product_id, source) "
                "VALUES (?, ?, ?, 'db_product')",
                (r["name"].strip(), alias, r["id"]),
            )
            count += 1
        except Exception:
            pass
    return count


def seed_from_rassvet_master(conn):
    """Source 2: Rassvet_Master Ibrat.xlsx — 1C names with Ibrat IDs mapped to DB products."""
    master_path = os.path.join(INVENTORY_DIR, "Rassvet_Master Ibrat.xlsx")
    if not os.path.exists(master_path):
        print(f"  SKIP: {master_path} not found")
        return 0

    import openpyxl
    wb = openpyxl.load_workbook(master_path, read_only=True)
    ws = wb["Katalog"]

    # Build a lookup: 1C name → DB product (by exact name match)
    db_products = conn.execute(
        "SELECT id, name FROM products WHERE name IS NOT NULL"
    ).fetchall()
    name_to_id = {}
    for p in db_products:
        n = normalize(p["name"])
        if n:
            name_to_id[n] = p["id"]

    count = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        asl_nomi = row[3]  # Column D: Asl nomi (1C)
        latin_name = row[4]  # Column E: Ilovadagi nomi (Latin)
        if not asl_nomi:
            continue

        asl_lower = normalize(asl_nomi)
        product_id = name_to_id.get(asl_lower)

        if not product_id:
            skipped += 1
            continue

        # Add the 1C name as alias (may already exist from source 1)
        try:
            conn.execute(
                "INSERT OR IGNORE INTO product_aliases (alias_name, alias_name_lower, product_id, source) "
                "VALUES (?, ?, ?, 'rassvet_master')",
                (asl_nomi.strip(), asl_lower, product_id),
            )
            count += 1
        except Exception:
            pass

        # Also add the Latin display name as an alias
        if latin_name and latin_name.strip():
            lat_lower = normalize(latin_name)
            if lat_lower and lat_lower != asl_lower:
                try:
                    conn.execute(
                        "INSERT OR IGNORE INTO product_aliases (alias_name, alias_name_lower, product_id, source) "
                        "VALUES (?, ?, ?, 'rassvet_master_latin')",
                        (latin_name.strip(), lat_lower, product_id),
                    )
                    count += 1
                except Exception:
                    pass

    wb.close()
    if skipped:
        print(f"  ({skipped} Ibrat rows had no DB match — expected for deactivated products)")
    return count


def seed_from_supply_history(conn):
    """Source 3: supply_order_items — historical 1C names from 15-month supply docs."""
    try:
        rows = conn.execute(
            """SELECT DISTINCT soi.product_name_1c, p.id as product_id
               FROM supply_order_items soi
               JOIN products p ON p.id = soi.product_id
               WHERE soi.product_name_1c IS NOT NULL AND soi.product_id IS NOT NULL"""
        ).fetchall()
    except Exception as e:
        print(f"  SKIP: {e}")
        return 0

    count = 0
    for r in rows:
        alias = normalize(r["product_name_1c"])
        if not alias:
            continue
        try:
            conn.execute(
                "INSERT OR IGNORE INTO product_aliases (alias_name, alias_name_lower, product_id, source) "
                "VALUES (?, ?, ?, 'supply_history')",
                (r["product_name_1c"].strip(), alias, r["product_id"]),
            )
            count += 1
        except Exception:
            pass
    return count


def seed_from_catalog_clean(conn):
    """Source 4: data/products.xlsx 'Catalog Clean' — Asl nomi column."""
    clean_path = os.path.join(DATA_DIR, "products.xlsx")
    if not os.path.exists(clean_path):
        # Try inventory folder copy
        clean_path = os.path.join(INVENTORY_DIR, "products.xlsx")
    if not os.path.exists(clean_path):
        print(f"  SKIP: products.xlsx not found")
        return 0

    import openpyxl
    wb = openpyxl.load_workbook(clean_path, read_only=True)
    if "Catalog Clean" not in wb.sheetnames:
        print(f"  SKIP: 'Catalog Clean' sheet not found")
        wb.close()
        return 0

    ws = wb["Catalog Clean"]

    # Build name→id lookup
    db_products = conn.execute(
        "SELECT id, name FROM products WHERE name IS NOT NULL"
    ).fetchall()
    name_to_id = {}
    for p in db_products:
        n = normalize(p["name"])
        if n:
            name_to_id[n] = p["id"]

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Column index depends on sheet layout
        # Catalog Clean: col L (index 11) = Asl nomi, col C (index 2) = Mahsulot nomi
        asl_nomi = row[11] if len(row) > 11 else None
        product_name = row[2] if len(row) > 2 else None

        if not asl_nomi:
            continue

        asl_lower = normalize(asl_nomi)
        product_id = name_to_id.get(asl_lower)
        if not product_id:
            continue

        # Add the cleaned Latin product name as alias
        if product_name and product_name.strip():
            clean_lower = normalize(product_name)
            if clean_lower and clean_lower != asl_lower:
                try:
                    conn.execute(
                        "INSERT OR IGNORE INTO product_aliases (alias_name, alias_name_lower, product_id, source) "
                        "VALUES (?, ?, ?, 'catalog_clean')",
                        (product_name.strip(), clean_lower, product_id),
                    )
                    count += 1
                except Exception:
                    pass

    wb.close()
    return count


def main():
    if DRY_RUN:
        print("DRY RUN — no changes will be written\n")

    conn = get_db()

    # Ensure tables exist
    conn.execute("""
        CREATE TABLE IF NOT EXISTS product_aliases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alias_name TEXT NOT NULL,
            alias_name_lower TEXT NOT NULL,
            product_id INTEGER NOT NULL,
            source TEXT DEFAULT 'manual',
            confirmed INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (product_id) REFERENCES products(id)
        )
    """)
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_product_aliases_name
        ON product_aliases(alias_name_lower)
    """)

    # Check existing count
    existing = conn.execute("SELECT COUNT(*) FROM product_aliases").fetchone()[0]
    print(f"Existing aliases: {existing}\n")

    print("Source 1: DB products (1C names)...")
    n1 = seed_from_db_products(conn)
    print(f"  +{n1} aliases\n")

    print("Source 2: Rassvet_Master Ibrat.xlsx...")
    n2 = seed_from_rassvet_master(conn)
    print(f"  +{n2} aliases\n")

    print("Source 3: Supply history (10K+ line items)...")
    n3 = seed_from_supply_history(conn)
    print(f"  +{n3} aliases\n")

    print("Source 4: Catalog Clean (Latin product names)...")
    n4 = seed_from_catalog_clean(conn)
    print(f"  +{n4} aliases\n")

    if DRY_RUN:
        conn.rollback()
        print("DRY RUN — rolled back")
    else:
        conn.commit()

    final = conn.execute("SELECT COUNT(*) FROM product_aliases").fetchone()[0]
    by_source = conn.execute(
        "SELECT source, COUNT(*) as c FROM product_aliases GROUP BY source ORDER BY c DESC"
    ).fetchall()

    print(f"\n{'='*40}")
    print(f"Total aliases: {final} (was {existing})")
    print(f"New aliases: {final - existing}")
    print(f"\nBy source:")
    for r in by_source:
        print(f"  {r['source']:<25} {r['c']:>5}")

    conn.close()


if __name__ == "__main__":
    main()
