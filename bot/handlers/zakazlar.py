"""/zakazlar — per-supplier reorder view.

Inventory-group command. Lists active suppliers as inline-keyboard buttons;
clicking one returns BOTH:
  (a) a simple chat message with status emoji + name + qoldiq + buyurtma
      (read-top-down to place an order — non-tech-friendly), and
  (b) an xlsx attachment with 4 sheets for desktop analysis:
      Tushuntirish (trilingual guide) / Buyurtma / Hammasi / Yig'ma.

Formula: 30d demand × seasonal YoY multiplier × (lead_time + 7d review) × 1.5 safety.
Lead time is the median inter-delivery gap from supply history (product-level
→ supplier-level → 14d global fallback). Demand-signal augmentation adds lost
demand from stockout days. No prices (per memory feedback_order_prep_no_prices).
"""
from __future__ import annotations

import io
import logging
from typing import List

from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)

from bot.shared import is_admin
from backend.services.reorder import (
    DEFAULT_REVIEW_PERIOD_DAYS,
    DEFAULT_SAFETY_FACTOR,
    DEFAULT_WINDOW_DAYS,
    list_supplier_full,
    list_suppliers_with_products,
)

router = Router()
logger = logging.getLogger(__name__)


STATUS_EMOJI = {
    "stockout": "🔴",
    "order_now": "🟠",
    "order_soon": "🟡",
    "ok": "🟢",
    "no_recent_demand": "⚪",
}
STATUS_LABEL_UZ = {
    "stockout": "Tugagan",
    "order_now": "Tezda",
    "order_soon": "Yaqin orada",
    "ok": "Yetarli",
    "no_recent_demand": "Talab yo'q",
}


def _supplier_keyboard(suppliers: List[dict]) -> InlineKeyboardMarkup:
    rows = []
    pair: List[InlineKeyboardButton] = []
    for s in suppliers:
        label = f"{s['name_1c']} ({s['oos_count']}/{s['product_count']})"
        if len(label) > 60:
            label = label[:57] + "..."
        cb = f"zakaz:s:{s['id']}" if s['id'] is not None else "zakaz:s:none"
        pair.append(InlineKeyboardButton(text=label, callback_data=cb))
        if len(pair) == 2:
            rows.append(pair)
            pair = []
    if pair:
        rows.append(pair)
    rows.append([InlineKeyboardButton(text="❌ Yopish", callback_data="zakaz:close")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _format_simple_text(supplier_label: str, items: List[dict]) -> str:
    """Status-grouped list, top-down for hand-placing an order."""
    if not items:
        return (f"📦 {supplier_label}\n\n"
                f"Hech qanday mahsulot uchun buyurtma kerak emas.\n"
                f"(Oxirgi {DEFAULT_WINDOW_DAYS} kun sotuvi + mavsumiy multiplikator "
                f"+ yetkazib berish muddati formulasiga ko'ra qoldiq yetarli.)")

    total_buy = sum(it["suggested_buy"] for it in items)
    lines = [
        f"📦 {supplier_label}",
        f"Buyurtma kerak: {len(items)} ta mahsulot · jami {total_buy:,} dona",
        "",
    ]
    for i, it in enumerate(items, 1):
        emoji = STATUS_EMOJI.get(it["status"], "·")
        name = it["name"]
        if len(name) > 52:
            name = name[:49] + "..."
        lines.append(f"{i}. {emoji} {name}")
        doc = it["days_of_cover"]
        cover_txt = f" · ~{int(doc)} kun qoldi" if doc is not None and doc < 999 else ""
        lines.append(f"   Qoldiq: {int(it['stock'])} → Buyurtma: {it['suggested_buy']}{cover_txt}")
    lines.append("")
    lines.append("🔴 Tugagan  🟠 Tezda  🟡 Yaqin orada")
    lines.append("💡 Batafsil ma'lumot — biriktirilgan Excel fayldan oching.")
    return "\n".join(lines)


def _chunk_text(text: str, limit: int = 3900) -> List[str]:
    if len(text) <= limit:
        return [text]
    chunks, cur = [], ""
    for line in text.split("\n"):
        if len(cur) + len(line) + 1 > limit:
            chunks.append(cur)
            cur = line
        else:
            cur = (cur + "\n" + line) if cur else line
    if cur:
        chunks.append(cur)
    return chunks


def _build_xlsx(supplier_label: str, reorder_items: List[dict],
                full_items: List[dict]) -> bytes:
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2E4F73")
    red_fill = PatternFill("solid", fgColor="FFD6D6")
    amber_fill = PatternFill("solid", fgColor="FFF3CC")
    yellow_fill = PatternFill("solid", fgColor="FFF9CC")
    green_fill = PatternFill("solid", fgColor="D8F0D8")
    grey_fill = PatternFill("solid", fgColor="EEEEEE")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_FILL = {
        "stockout": red_fill,
        "order_now": amber_fill,
        "order_soon": yellow_fill,
        "ok": green_fill,
        "no_recent_demand": grey_fill,
    }

    COL_DEFS = [
        ("#", "rank", 5,
         "Rank: top of list = most urgent (sorted by status then days-of-cover).",
         "Ранг: вверху — самые срочные.",
         "Tartib: yuqorida eng shoshilinch."),
        ("Status", "status", 14,
         "🔴 stockout · 🟠 order now · 🟡 order soon · 🟢 ok · ⚪ no recent demand.",
         "🔴 нет в наличии · 🟠 срочно · 🟡 скоро · 🟢 норма · ⚪ нет спроса.",
         "🔴 tugagan · 🟠 tezda · 🟡 yaqin orada · 🟢 yetarli · ⚪ talab yo'q."),
        ("Mahsulot nomi", "name", 42,
         "Full 1C product name (Cyrillic). Copy verbatim when ordering.",
         "Полное название из 1С (кириллица). Копировать дословно.",
         "1C-dagi to'liq nom (kirill). Buyurtmada aynan shu matnni nusxalang."),
        ("Stock", "stock", 8,
         "Current warehouse stock from latest 1C balances upload (donalar).",
         "Текущий складской остаток (шт.).",
         "Hozirgi ombor qoldig'i (dona)."),
        (f"Sotilgan ({DEFAULT_WINDOW_DAYS}d)", "sold_window", 12,
         f"Real-orders units in last {DEFAULT_WINDOW_DAYS}d. Demand signals shown separately.",
         f"Реальные продажи за {DEFAULT_WINDOW_DAYS} дней.",
         f"Oxirgi {DEFAULT_WINDOW_DAYS} kun haqiqiy sotuvi (dona)."),
        ("Lost (signals)", "demand_signal_qty", 11,
         "Demand signals from out-of-stock orders — added to demand for forecast.",
         "Сигналы спроса при нулевом остатке — добавлены к спросу.",
         "Qoldiqsiz buyurtma signallari — talabga qo'shilgan."),
        ("Kunlik", "seasoned_daily", 9,
         "Seasoned daily rate: (sold + lost) / window × seasonal multiplier.",
         "Среднесуточный спрос: (sold + lost) / window × сезонный коэф.",
         "O'rtacha kunlik talab: (sotilgan + signallar) / oyna × mavsumiy."),
        ("Mavsum", "seasonal_mult", 8,
         "YoY multiplier = (last-year same-month daily) / (last-year prior-60d daily). 1.0 = no signal.",
         "Сезонный коэф. = продажи (этот месяц прошлого года) / (предыдущие 60 дней). 1.0 = нет данных.",
         "Mavsumiy koef. = (o'tgan yili shu oy) / (oldingi 60 kun). 1.0 = ma'lumot yo'q."),
        ("Yetkazish", "lead_time_days", 11,
         "Lead time (days) = median gap between supply deliveries. Product-level if ≥3 events, else supplier-level, else 14d.",
         "Срок поставки (дни) = медиана между поставками. По товару если ≥3 событий, иначе по поставщику, иначе 14д.",
         "Yetkazib berish muddati (kun) = supply orasidagi mediana. ≥3 hodisa bo'lsa mahsulotga, aks holda supplierga, aks holda 14."),
        ("Buyurtma", "suggested_buy", 11,
         f"Suggested order = ceil(seasoned_daily × (lead + {DEFAULT_REVIEW_PERIOD_DAYS}d review) × {DEFAULT_SAFETY_FACTOR} safety) − stock.",
         f"Рекомендуемый заказ = ceil(daily × (lead + {DEFAULT_REVIEW_PERIOD_DAYS}д) × {DEFAULT_SAFETY_FACTOR}) − остаток.",
         f"Tavsiya buyurtma = ceil(kunlik × (lead + {DEFAULT_REVIEW_PERIOD_DAYS}) × {DEFAULT_SAFETY_FACTOR}) − qoldiq."),
        ("~Cover", "days_of_cover", 9,
         "Days of stock remaining at current daily rate.",
         "Сколько дней хватит при текущем спросе.",
         "Hozirgi qoldiq necha kun yetadi."),
        ("Oxirgi sotuv", "last_sale", 11,
         "Most recent date this SKU was sold. >60 days = verify still active.",
         "Дата последней продажи.",
         "Oxirgi sotilgan sana."),
    ]

    today = date.today().isoformat()
    wb = Workbook()

    ws_g = wb.active
    ws_g.title = "Tushuntirish (Guide)"

    ws_g["A1"] = f"{supplier_label} — Buyurtma Tahlili / Анализ заказа / Order Analysis"
    ws_g["A1"].font = Font(bold=True, size=15)
    ws_g.merge_cells("A1:D1")
    n_full = len(full_items)
    n_buy = len(reorder_items)
    ws_g["A2"] = (f"Sana / Дата / Date: {today} | "
                  f"Window: {DEFAULT_WINDOW_DAYS}d | "
                  f"Review: {DEFAULT_REVIEW_PERIOD_DAYS}d | "
                  f"Safety: {DEFAULT_SAFETY_FACTOR}× | "
                  f"{n_full} ta mahsulot · {n_buy} buyurtma kerak")
    ws_g["A2"].font = Font(italic=True, color="666666")
    ws_g.merge_cells("A2:D2")

    def section(ws, row, title):
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, size=13, color="FFFFFF")
        c.fill = head_fill
        c.alignment = left
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 22
        return row + 1

    r = 4
    r = section(ws_g, r, "1. Bu hisobot nima haqida / О чём этот отчёт / What this report does")
    overview = [
        ("UZ", f"'{supplier_label}' yetkazib beruvchidagi {n_full} ta aktiv mahsulot uchun "
               f"buyurtma tavsiyasi. Hisob asoslari: oxirgi {DEFAULT_WINDOW_DAYS} kun haqiqiy sotuvi, "
               f"qoldiqsiz davrda kelgan talab signallari, o'tgan yilgi shu oyning mavsumiy koef., "
               f"hamda supply tarixidan hisoblangan yetkazish muddati."),
        ("RU", f"Рекомендация по заказу для {n_full} активных товаров поставщика "
               f"'{supplier_label}'. Учтены: продажи за {DEFAULT_WINDOW_DAYS} дн., сигналы спроса "
               f"при нулевом остатке, сезонный коэф. (год к году) и расчётный срок поставки."),
        ("EN", f"Reorder recommendation for {n_full} active products mapped to "
               f"'{supplier_label}'. Inputs: last-{DEFAULT_WINDOW_DAYS}-day real sales, "
               f"out-of-stock demand signals, year-over-year seasonal multiplier, "
               f"lead time derived from supply delivery history."),
    ]
    for lang, text in overview:
        ws_g.cell(row=r, column=1, value=lang).font = bold
        ws_g.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ws_g.cell(row=r, column=2, value=text).alignment = left_top
        ws_g.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws_g.row_dimensions[r].height = 60
        r += 1

    r += 1
    r = section(ws_g, r, "2. Hisoblash usuli / Метод расчёта / Calculation method")
    method_lines = [
        ("Window", f"{DEFAULT_WINDOW_DAYS} kun",
         f"Sotuv tahlili davri.",
         f"Период анализа продаж."),
        ("Lost demand", "demand_signals",
         "Qoldiqsiz buyurtma signallari talabga qo'shiladi.",
         "Сигналы спроса при OOS добавлены к спросу."),
        ("Seasonal (YoY)", "last-year same-month / prior-60d",
         "Mavsumiy koef. — o'tgan yili shu oy bilan oldingi 60 kun nisbati. <30 dona bo'lsa 1.0.",
         "Сезонный коэф. — отношение к предыдущим 60 дням. <30 шт. → 1.0."),
        ("Lead time", "median gap (≥3 events)",
         "Mahsulot darajasida ≥3 supply hodisasi → mediana. Aks holda supplier yoki 14 kun.",
         "Медиана между поставками. Если событий <3 → по поставщику или 14д."),
        ("Review period", f"{DEFAULT_REVIEW_PERIOD_DAYS} kun",
         "Buyurtmalar orasidagi qarash davri.",
         "Период между ревизиями заказа."),
        ("Safety factor", f"{DEFAULT_SAFETY_FACTOR}×",
         "Xatolik buferi (spros pikida tushib qolmaslik uchun).",
         "Запас на ошибку прогноза."),
        ("Reorder point", "daily × lead × safety",
         "Shundan past tushganda buyurtma berish kerak.",
         "Точка перезаказа."),
        ("Target", "daily × (lead + review) × safety",
         "Buyurtmadan keyin qancha qoldiq bo'lishi kerakligi.",
         "Целевой остаток после поставки."),
        ("Buyurtma", "ceil(target − stock)",
         "Manfiy bo'lsa 0.",
         "Если меньше 0 → 0."),
    ]
    headers = ["Atama / Параметр / Term", "Qiymat / Значение / Value", "UZ", "RU"]
    for ci, h in enumerate(headers, start=1):
        c = ws_g.cell(row=r, column=ci, value=h)
        c.font = head_font
        c.fill = head_fill
        c.alignment = center
        c.border = border
    ws_g.row_dimensions[r].height = 24
    r += 1
    for term, val, uz, ru in method_lines:
        ws_g.cell(row=r, column=1, value=term).font = bold
        ws_g.cell(row=r, column=1).alignment = left_top
        ws_g.cell(row=r, column=2, value=val).alignment = left_top
        ws_g.cell(row=r, column=3, value=uz).alignment = left_top
        ws_g.cell(row=r, column=4, value=ru).alignment = left_top
        for ci in range(1, 5):
            ws_g.cell(row=r, column=ci).border = border
        ws_g.row_dimensions[r].height = 42
        r += 1

    r += 1
    r = section(ws_g, r, "3. Status belgilari / Статусы / Status markers")
    decision = [
        ("🔴 stockout", "stock=0 + demand>0", red_fill,
         "OUT OF STOCK with active demand. Sales lost until restocked.",
         "Нет в наличии при активном спросе. Теряем продажи.",
         "Sotuvda yo'q va talab bor. Sotuv yo'qotamiz."),
        ("🟠 order_now", "stock < reorder_point", amber_fill,
         "Below reorder point — order now to avoid imminent stockout.",
         "Ниже точки перезаказа — заказать срочно.",
         "Perezakaz nuqtasidan past — tezda buyurtma bering."),
        ("🟡 order_soon", "stock < target", yellow_fill,
         "Above reorder point but below post-delivery target.",
         "Выше точки перезаказа, но ниже целевого остатка.",
         "Maqsad qiymatidan past — yaqin orada to'ldirish."),
        ("🟢 ok", "stock ≥ target", green_fill,
         "Sufficient cover. Skip.",
         "Достаточно. Не заказывать.",
         "Yetarli. Buyurtma bermang."),
        ("⚪ no_recent_demand", "no sales in window", grey_fill,
         "No sales in window. Verify with uncle before reordering.",
         "Нет продаж в окне. Уточнить у дяди.",
         "Oynada sotuv yo'q. Amakidan tekshiring."),
    ]
    headers = ["Belgi / Маркер / Marker", "Shart / Условие / Condition", "EN", "UZ / RU"]
    for ci, h in enumerate(headers, start=1):
        c = ws_g.cell(row=r, column=ci, value=h)
        c.font = head_font
        c.fill = head_fill
        c.alignment = center
        c.border = border
    ws_g.row_dimensions[r].height = 22
    r += 1
    for marker, cond, fill, en, ru, uz in decision:
        c1 = ws_g.cell(row=r, column=1, value=marker)
        c1.font = bold
        c1.alignment = center
        c1.fill = fill
        ws_g.cell(row=r, column=2, value=cond).alignment = left_top
        ws_g.cell(row=r, column=3, value=en).alignment = left_top
        ws_g.cell(row=r, column=4, value=f"UZ: {uz}\nRU: {ru}").alignment = left_top
        for ci in range(1, 5):
            ws_g.cell(row=r, column=ci).border = border
        ws_g.row_dimensions[r].height = 50
        r += 1

    r += 1
    r = section(ws_g, r, "4. Cheklovlar / Ограничения / Caveats")
    cav = [
        ("UZ",
         "• Yetkazib berish muddati supply tarixidan olinadi — buyurtma berilgan sana emas, kelgan sana. Real lead time bundan bir oz uzun bo'lishi mumkin.\n"
         "• Mavsumiy koef. faqat o'tgan yili shu oyda ≥30 dona sotilgan mahsulotlarga qo'llaniladi; yangi mahsulotlar uchun 1.0 ishlatiladi.\n"
         "• Stockout tarixi vaqt qatori sifatida saqlanmaydi — demand_signals faqat hozirgi qoldiqsiz davrlarda yozilgan, eski oylarni qoplamaydi.\n"
         "• Statistik xatolik buferi (1.5×) qo'lda sozlanadi; haqiqiy variatsiyani hisoblanmaydi."),
        ("RU",
         "• Срок поставки — медиана интервалов между поступлениями (а не «заказали–получили»). Реальный лид может быть длиннее.\n"
         "• Сезонный коэф. применяется только если в прошлом году за этот месяц продано ≥30 шт. Для новых SKU → 1.0.\n"
         "• История OOS как временной ряд не хранится — demand_signals покрывают только недавние эпизоды.\n"
         "• Буфер на ошибку прогноза (1.5×) задаётся вручную."),
        ("EN",
         "• Lead time is the median inter-delivery gap, not the order-to-arrival span. Actual lead time may be longer.\n"
         "• Seasonal multiplier applies only when last-year same-month sales ≥30 units. New SKUs use 1.0.\n"
         "• Historical stock-out series is not preserved — demand_signals only capture recent OOS episodes.\n"
         "• Safety factor (1.5×) is hand-tuned, not a statistical safety stock."),
    ]
    for lang, text in cav:
        ws_g.cell(row=r, column=1, value=lang).font = bold
        ws_g.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ws_g.cell(row=r, column=2, value=text).alignment = left_top
        ws_g.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws_g.row_dimensions[r].height = 75
        r += 1

    for col, w in [("A", 22), ("B", 32), ("C", 48), ("D", 48)]:
        ws_g.column_dimensions[col].width = w

    ws = wb.create_sheet("Buyurtma (Order)")
    ws["A1"] = f"{supplier_label} — Buyurtma kerak ({len(reorder_items)} ta mahsulot)"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(f"A1:{get_column_letter(len(COL_DEFS))}1")
    ws["A2"] = "Hover the column header for explanation. See 'Tushuntirish' sheet for full reference."
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells(f"A2:{get_column_letter(len(COL_DEFS))}2")

    for ci, (label, key, width, en, ru, uz) in enumerate(COL_DEFS, start=1):
        c = ws.cell(row=4, column=ci, value=label)
        c.font = head_font
        c.fill = head_fill
        c.alignment = center
        c.border = border
        cm = Comment(f"EN: {en}\n\nRU: {ru}\n\nUZ: {uz}", "Reorder Analysis")
        cm.width = 360
        cm.height = 140
        c.comment = cm
        ws.column_dimensions[get_column_letter(ci)].width = width

    _write_data_rows(ws, reorder_items, COL_DEFS, start_row=5, border=border,
                     left=left, right=right, center=center, bold=bold,
                     status_fill=STATUS_FILL)

    if reorder_items:
        r_idx = 5 + len(reorder_items)
        total_buy = sum(it["suggested_buy"] for it in reorder_items)
        total_stock = sum(int(it["stock"]) for it in reorder_items)
        ws.cell(row=r_idx, column=3, value=f"JAMI ({len(reorder_items)} ta)").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=r_idx, column=3).alignment = right
        ws.cell(row=r_idx, column=4, value=total_stock).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=r_idx, column=10, value=total_buy).font = Font(bold=True, color="FFFFFF")
        for ci in range(1, len(COL_DEFS) + 1):
            ws.cell(row=r_idx, column=ci).border = border
            ws.cell(row=r_idx, column=ci).fill = head_fill
    ws.freeze_panes = "A5"

    ws_all = wb.create_sheet(f"Hammasi ({len(full_items)})")
    ws_all["A1"] = f"{supplier_label} — barcha aktiv mahsulotlar ({len(full_items)} ta)"
    ws_all["A1"].font = Font(bold=True, size=14)
    ws_all.merge_cells(f"A1:{get_column_letter(len(COL_DEFS) + 1)}1")

    all_cols = COL_DEFS + [("Lifecycle", "lifecycle", 10,
                             "Popularity classifier: active / aging / stale / never.",
                             "Классификатор: active / aging / stale / never.",
                             "Faollik: active / aging / stale / never.")]
    for ci, (label, key, width, en, ru, uz) in enumerate(all_cols, start=1):
        c = ws_all.cell(row=3, column=ci, value=label)
        c.font = head_font
        c.fill = head_fill
        c.alignment = center
        c.border = border
        cm = Comment(f"EN: {en}\n\nRU: {ru}\n\nUZ: {uz}", "Reorder Analysis")
        cm.width = 360
        cm.height = 140
        c.comment = cm
        ws_all.column_dimensions[get_column_letter(ci)].width = width

    _write_data_rows(ws_all, full_items, all_cols, start_row=4, border=border,
                     left=left, right=right, center=center, bold=bold,
                     status_fill=STATUS_FILL)
    ws_all.freeze_panes = "A4"

    ws_y = wb.create_sheet("Yig'ma (Summary)")
    ws_y["A1"] = f"{supplier_label} — Yig'ma / Summary"
    ws_y["A1"].font = Font(bold=True, size=14)
    ws_y.merge_cells("A1:C1")

    status_counts = {k: 0 for k in STATUS_EMOJI}
    for it in full_items:
        status_counts[it["status"]] = status_counts.get(it["status"], 0) + 1
    total_buy = sum(it["suggested_buy"] for it in full_items)
    total_stock = sum(int(it["stock"]) for it in full_items)
    total_sold = sum(int(it["sold_window"]) for it in full_items)
    total_lost = sum(int(it["demand_signal_qty"]) for it in full_items)

    lead_sources = {"product": 0, "supplier": 0, "global": 0}
    seasonal_sources = {"yoy": 0, "fallback": 0}
    for it in full_items:
        lead_sources[it["lead_time_source"]] = lead_sources.get(it["lead_time_source"], 0) + 1
        seasonal_sources[it["seasonal_source"]] = seasonal_sources.get(it["seasonal_source"], 0) + 1

    rows_data = [
        ("Jami mahsulotlar / Всего / Total", len(full_items)),
        ("", ""),
        ("🔴 Stockout (tugagan + talab)", status_counts.get("stockout", 0)),
        ("🟠 Order now (perezakaz)", status_counts.get("order_now", 0)),
        ("🟡 Order soon (yaqin orada)", status_counts.get("order_soon", 0)),
        ("🟢 Ok (yetarli)", status_counts.get("ok", 0)),
        ("⚪ No demand (talab yo'q)", status_counts.get("no_recent_demand", 0)),
        ("", ""),
        ("Jami qoldiq / Stock total (dona)", total_stock),
        (f"Oxirgi {DEFAULT_WINDOW_DAYS}d sotuv / Sales window", total_sold),
        ("Qoldiqsiz signal (lost demand)", total_lost),
        ("Tavsiya jami buyurtma / Total order", total_buy),
        ("", ""),
        ("Lead time — product-level", lead_sources["product"]),
        ("Lead time — supplier-level fallback", lead_sources["supplier"]),
        ("Lead time — global 14d fallback", lead_sources["global"]),
        ("", ""),
        ("Seasonal — YoY applied", seasonal_sources["yoy"]),
        ("Seasonal — 1.0 fallback", seasonal_sources["fallback"]),
    ]
    r = 3
    for label, val in rows_data:
        if label:
            ws_y.cell(row=r, column=1, value=label).alignment = left
            ws_y.cell(row=r, column=2, value=val).alignment = right
            if isinstance(val, int) and val > 0:
                ws_y.cell(row=r, column=2).font = bold
        r += 1
    ws_y.column_dimensions["A"].width = 50
    ws_y.column_dimensions["B"].width = 14

    r += 1
    ws_y.cell(row=r, column=1, value="Method:").font = bold
    ws_y.cell(row=r, column=2,
              value=f"({DEFAULT_WINDOW_DAYS}d sales + lost) × YoY × (lead + {DEFAULT_REVIEW_PERIOD_DAYS}d review) × {DEFAULT_SAFETY_FACTOR} − stock")
    ws_y.cell(row=r + 1, column=1, value="Generated:").font = bold
    ws_y.cell(row=r + 1, column=2, value=today)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_data_rows(ws, items, cols, start_row, border, left, right, center,
                     bold, status_fill):
    """Generic writer for Buyurtma + Hammasi sheets. Renders one row per item."""
    key_to_col = {c[1]: i + 1 for i, c in enumerate(cols)}
    for i, it in enumerate(items, 1):
        r = start_row + i - 1
        fill = status_fill.get(it["status"])
        ws.cell(row=r, column=1, value=i).alignment = center
        emoji = STATUS_EMOJI.get(it["status"], "")
        label = STATUS_LABEL_UZ.get(it["status"], it["status"])
        ws.cell(row=r, column=2, value=f"{emoji} {label}").alignment = center
        ws.cell(row=r, column=3, value=it["name"]).alignment = left
        ws.cell(row=r, column=4, value=int(it["stock"])).alignment = right
        ws.cell(row=r, column=5, value=int(it["sold_window"])).alignment = right
        ws.cell(row=r, column=6, value=int(it["demand_signal_qty"])).alignment = right
        ws.cell(row=r, column=7, value=round(it["seasoned_daily"], 2)).alignment = right
        ws.cell(row=r, column=8, value=it["seasonal_mult"]).alignment = right
        lt_txt = f"{it['lead_time_days']}"
        if it["lead_time_source"] == "supplier":
            lt_txt += "*"
        elif it["lead_time_source"] == "global":
            lt_txt += "†"
        ws.cell(row=r, column=9, value=lt_txt).alignment = right
        buy_cell = ws.cell(row=r, column=10, value=it["suggested_buy"])
        buy_cell.alignment = right
        buy_cell.font = bold
        doc = it["days_of_cover"]
        ws.cell(row=r, column=11, value=("∞" if doc is None else doc)).alignment = right
        ws.cell(row=r, column=12, value=it["last_sale"]).alignment = center
        if "lifecycle" in key_to_col:
            ws.cell(row=r, column=13, value=it["lifecycle"]).alignment = center
        for ci in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=ci)
            if fill is not None:
                cell.fill = fill
            cell.border = border


@router.message(Command("zakazlar"))
async def cmd_zakazlar(message: Message):
    if not is_admin(message):
        return
    suppliers = list_suppliers_with_products()
    if not suppliers:
        await message.reply(
            "Hech qanday yetkazib beruvchi topilmadi.\n"
            "(Mahsulotlarda `latest_supplier_id` belgilanmagan — /supply yuklanishi kerak.)"
        )
        return
    total_oos = sum(s["oos_count"] for s in suppliers)
    total_prod = sum(s["product_count"] for s in suppliers)
    text = (
        f"📦 Yetkazib beruvchini tanlang:\n"
        f"({len(suppliers)} ta supplier · {total_prod} ta mahsulot · "
        f"{total_oos} ta tugagan)\n\n"
        f"Format: <b>NAME (tugagan/jami)</b>"
    )
    await message.reply(text, parse_mode="HTML",
                        reply_markup=_supplier_keyboard(suppliers))


@router.callback_query(F.data.startswith("zakaz:s:"))
async def cb_supplier_pick(cb: CallbackQuery):
    await cb.answer("Hisoblanmoqda...")

    suffix = cb.data.removeprefix("zakaz:s:")
    supplier_id = None if suffix == "none" else int(suffix)

    if supplier_id is None:
        label = "(noma'lum supplier)"
    else:
        from backend.database import get_db
        conn = get_db()
        try:
            row = conn.execute(
                "SELECT name_1c FROM suppliers WHERE id = ?", (supplier_id,)
            ).fetchone()
            label = row["name_1c"] if row else f"#{supplier_id}"
        finally:
            conn.close()

    full_items = list_supplier_full(supplier_id)
    reorder_items = [x for x in full_items if x["suggested_buy"] > 0]

    text = _format_simple_text(label, reorder_items)
    for chunk in _chunk_text(text):
        await cb.message.answer(chunk)

    xlsx_bytes = _build_xlsx(label, reorder_items, full_items)
    from datetime import date as _date
    safe_name = "".join(c if ord(c) < 128 else "_" for c in label).strip("_") or "supplier"
    fname = f"buyurtma_{safe_name}_{_date.today().isoformat()}.xlsx"
    caption = (f"📊 {label} — to'liq tahlil ({len(full_items)} mahsulot, "
               f"{len(reorder_items)} buyurtma kerak)")
    await cb.message.answer_document(
        BufferedInputFile(xlsx_bytes, filename=fname[:120]),
        caption=caption,
    )


@router.callback_query(F.data == "zakaz:close")
async def cb_close(cb: CallbackQuery):
    await cb.answer()
    try:
        await cb.message.delete()
    except Exception:
        pass
